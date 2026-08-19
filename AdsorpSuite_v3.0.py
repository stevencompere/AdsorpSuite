# -*- coding: utf-8 -*-
"""
===============================================================================
 AdsorpSuite  v3.0
 Suite complete d'analyse d'isothermes d'adsorption  (tout-en-un)
===============================================================================

 Onglets :
   1. Donnees        : saisie manuelle / collage Excel / import .xlsx-.csv
                       + IMPORT EN LOT de plusieurs fichiers (v3)
   2. Conversion     : quantite en EXCES  ->  quantite TOTALE (n_tot)
                       + conversion de TOUS les jeux en une fois (v3)
   3. Ajustement     : Langmuir, Langmuir bi-site, Freundlich, Sips, Toth,
                       Dubinin-Radushkevich (D-R), Dubinin-Astakhov (D-A)
                       + ajustement en lot et matrice de comparaison (v3)
   4. IAST           : melanges binaires et ternaires, tous modeles
                       + melanges types en un clic (v3)
   5. Graphiques     : figure personnalisable + export PNG/PDF/SVG/EPS/TIFF
   6. Export         : classeur Excel multi-feuilles, projet .adsp (JSON)
                       + export Origin (Long Name / Units) (v3)

-------------------------------------------------------------------------------
 NOUVEAUTES DE LA VERSION 3.0
-------------------------------------------------------------------------------
 EXACTITUDE
   * Densites de gaz de reference tabulees (NIST Chemistry WebBook) pour CO2,
     CH4 et H2 a 303 K, choisies par defaut pour la conversion exces -> total.
     Peng-Robinson s'ecarte de 1 a 3 % de ces valeurs, et cet ecart se propage
     directement dans n_total. Repli automatique sur Peng-Robinson hors table.
   * La verification interne compte desormais 39 controles (contre 30).

 ERGONOMIE
   * PLUS DE PERTE DE SAISIE : le jeu de donnees en cours est enregistre
     automatiquement quand on change de jeu, quand on quitte l'onglet, et avant
     tout calcul. Un temoin "modifications non enregistrees / a jour" indique
     l'etat en permanence.
   * Champ de filtre au-dessus de la liste des jeux de donnees.
   * Import en lot : selection multiple de fichiers, deduction automatique du
     gaz et de l'echantillon d'apres le nom, recapitulatif modifiable avant
     validation (double-clic pour corriger une ligne).
   * Conversion en lot : tous les jeux "Exces" convertis d'un coup.
   * Ajustement en lot : tous les jeux x tous les modeles coches, avec barre de
     progression et rapport d'echecs.
   * Matrice de comparaison des modeles : tableau croise jeu x modele
     (R2, RMSE, AICc) avec le meilleur modele signale par jeu.
   * Melanges types IAST en un clic (CO2/CH4 50:50, 25:75, 75:25, CH4/H2 50:50,
     ternaire), avec association automatique des ajustements par gaz.
   * Export Origin : classeur au format large, ligne 1 = Long Name,
     ligne 2 = Units, donnees a partir de la ligne 3.
   * Raccourcis : Maj+F5 = action en lot de l'onglet actif, Ctrl+I = import
     multiple, Ctrl+E = export Origin.

-------------------------------------------------------------------------------
 Dependances : numpy, scipy, pandas, matplotlib, openpyxl
               (optionnel : CoolProp pour les densites de gaz reel)

 Lancement   : python AdsorpSuite.py
-------------------------------------------------------------------------------
 Auteur : genere pour Steven Compere - analyse d'adsorption sur carbones ZTC
===============================================================================
"""

from __future__ import annotations

import os
import sys
import json
import math
import copy
import datetime
import traceback

import numpy as np
import pandas as pd

import matplotlib
try:
    matplotlib.use("TkAgg")
except Exception:
    pass
from matplotlib.figure import Figure
try:
    from matplotlib.backends.backend_tkagg import (FigureCanvasTkAgg,
                                                   NavigationToolbar2Tk)
except Exception:
    FigureCanvasTkAgg = None
    NavigationToolbar2Tk = None

from scipy.optimize import curve_fit, brentq
from scipy.integrate import quad
from scipy.special import gammaincc, gamma as gamma_fn

import tkinter as tk
from tkinter import ttk, filedialog, messagebox, colorchooser

try:
    from CoolProp.CoolProp import PropsSI
    HAS_COOLPROP = True
except Exception:
    HAS_COOLPROP = False


# =============================================================================
# 1. CONSTANTES ET DONNEES PHYSIQUES
# =============================================================================

APP_NAME = "AdsorpSuite"
APP_VERSION = "3.0"

R_J = 8.314462618          # J / (mol.K)
_TRAPZ = getattr(np, "trapezoid", None) or np.trapz   # numpy 1.x et 2.x
V_MOLAR_STP = 22.413969    # L/mol  a 273.15 K et 1 atm  (=> cm3/mmol)

# Base de donnees des gaz : Tc (K), Pc (bar), facteur acentrique, M (g/mol)
GAS_DB = {
    "CO2":     dict(Tc=304.1282, Pc=73.773,  w=0.22394,   M=44.0095),
    "CH4":     dict(Tc=190.564,  Pc=45.992,  w=0.01142,   M=16.0425),
    "H2":      dict(Tc=33.145,   Pc=12.964,  w=-0.219,    M=2.01588),
    "N2":      dict(Tc=126.192,  Pc=33.958,  w=0.0372,    M=28.0134),
    "O2":      dict(Tc=154.581,  Pc=50.430,  w=0.0222,    M=31.9988),
    "Ar":      dict(Tc=150.687,  Pc=48.630,  w=-0.00219,  M=39.948),
    "CO":      dict(Tc=132.860,  Pc=34.940,  w=0.0497,    M=28.0101),
    "He":      dict(Tc=5.1953,   Pc=2.2746,  w=-0.385,    M=4.0026),
    "Ne":      dict(Tc=44.4918,  Pc=26.786,  w=-0.0387,   M=20.1797),
    "Kr":      dict(Tc=209.48,   Pc=55.250,  w=-0.000894, M=83.798),
    "Xe":      dict(Tc=289.733,  Pc=58.420,  w=0.00363,   M=131.293),
    "C2H6":    dict(Tc=305.322,  Pc=48.722,  w=0.0995,    M=30.069),
    "C2H4":    dict(Tc=282.35,   Pc=50.418,  w=0.0866,    M=28.054),
    "C3H8":    dict(Tc=369.89,   Pc=42.512,  w=0.1521,    M=44.0956),
    "n-C4H10": dict(Tc=425.125,  Pc=37.960,  w=0.201,     M=58.122),
    "NH3":     dict(Tc=405.40,   Pc=113.33,  w=0.25601,   M=17.0305),
    "H2S":     dict(Tc=373.10,   Pc=90.000,  w=0.1005,    M=34.081),
    "SO2":     dict(Tc=430.64,   Pc=78.840,  w=0.2557,    M=64.064),
    "H2O":     dict(Tc=647.096,  Pc=220.640, w=0.3443,    M=18.01528),
}
GAS_LIST = list(GAS_DB.keys())

# Noms CoolProp correspondants
COOLPROP_NAMES = {
    "CO2": "CarbonDioxide", "CH4": "Methane", "H2": "Hydrogen",
    "N2": "Nitrogen", "O2": "Oxygen", "Ar": "Argon", "CO": "CarbonMonoxide",
    "He": "Helium", "Ne": "Neon", "Kr": "Krypton", "Xe": "Xenon",
    "C2H6": "Ethane", "C2H4": "Ethylene", "C3H8": "n-Propane",
    "n-C4H10": "n-Butane", "NH3": "Ammonia", "H2S": "HydrogenSulfide",
    "SO2": "SulfurDioxide", "H2O": "Water",
}

# Facteurs de conversion vers le bar
P_UNITS = {
    "bar": 1.0, "mbar": 1e-3, "Pa": 1e-5, "kPa": 1e-2, "MPa": 10.0,
    "atm": 1.01325, "torr": 1.0/750.06168, "mmHg": 1.0/750.06168,
    "psi": 0.0689475729,
}

Q_UNITS = ["mmol/g", "mol/kg", "cm3(STP)/g", "mg/g", "wt%", "g/g", "mmol/cm3"]


def q_to_mmol_g(q, unit, M_gmol, rho_bulk=None):
    """Convertit une quantite adsorbee vers mmol/g."""
    q = np.asarray(q, dtype=float)
    if unit in ("mmol/g", "mol/kg"):
        return q
    if unit == "cm3(STP)/g":
        return q / V_MOLAR_STP
    if unit == "mg/g":
        return q / M_gmol
    if unit == "wt%":
        # g adsorbat / 100 g adsorbant  ->  mmol/g
        return q * 10.0 / M_gmol
    if unit == "g/g":
        return q * 1000.0 / M_gmol
    if unit == "mmol/cm3":
        if not rho_bulk:
            return q
        return q / rho_bulk
    return q


def mmol_g_to_unit(q, unit, M_gmol, rho_bulk=None):
    """Conversion inverse de q_to_mmol_g."""
    q = np.asarray(q, dtype=float)
    if unit in ("mmol/g", "mol/kg"):
        return q
    if unit == "cm3(STP)/g":
        return q * V_MOLAR_STP
    if unit == "mg/g":
        return q * M_gmol
    if unit == "wt%":
        return q * M_gmol / 10.0
    if unit == "g/g":
        return q * M_gmol / 1000.0
    if unit == "mmol/cm3":
        if not rho_bulk:
            return q
        return q * rho_bulk
    return q



# -----------------------------------------------------------------------------
# Densites molaires de reference (NIST Chemistry WebBook) le long d'isothermes
# usuelles. Sert d'alternative "exacte" a l'EOS cubique pour la conversion
# exces -> total, ou l'erreur sur rho se propage directement sur n_total.
#   cle   : (gaz, T en K)
#   valeur: (tableau P en bar, tableau rho en mol/L == mmol/cm3)
# Sources : Span & Wagner 1996 (CO2), Setzmann & Wagner 1991 (CH4),
#           Leachman et al. 2009 (H2 normal).
# CO2 est SOUS-critique a 303 K (Tc = 304,128 K) : la table s'arrete a
# Psat = 71,89 bar, au-dela le fluide est liquide.
# -----------------------------------------------------------------------------

_NIST_RAW = {
 ("CO2", 303.0): (
  "0 1 2 3 4 5 6 7 8 9 10 11 12 13 14 15 16 17 18 19 20 21 22 23 24 25 26 27 "
  "28 29 30 31 32 33 34 35 36 37 38 39 40 41 42 43 44 45 46 47 48 49 50 51 52 "
  "53 54 55 56 57 58 59 60 61 62 63 64 65 66 67 68 69 70 71",
  "0 0.0398818 0.0801462 0.120802 0.161860 0.203329 0.245220 0.287544 0.330313 "
  "0.373538 0.417231 0.461406 0.506075 0.551254 0.596956 0.643198 0.689995 "
  "0.737365 0.785324 0.833893 0.883090 0.932936 0.983453 1.03466 1.08659 "
  "1.13927 1.19271 1.24695 1.30203 1.35796 1.41479 1.47256 1.53129 1.59104 "
  "1.65185 1.71376 1.77683 1.84111 1.90666 1.97354 2.04183 2.11159 2.18292 "
  "2.25589 2.33061 2.40719 2.48573 2.56637 2.64925 2.73454 2.82241 2.91307 "
  "3.00673 3.10366 3.20416 3.30856 3.41727 3.53073 3.64951 3.77425 3.90576 "
  "4.04500 4.19320 4.35192 4.52320 4.70980 4.91558 5.14622 5.41063 5.72425 "
  "6.11813 6.67526"),
 ("CH4", 303.0): (
  " ".join(str(i) for i in range(0, 101)),
  "0 0.0397584 0.0796467 0.119665 0.159814 0.200094 0.240505 0.281047 0.321721 "
  "0.362527 0.403466 0.444536 0.485740 0.527077 0.568547 0.610150 0.651887 "
  "0.693758 0.735763 0.777902 0.820176 0.862584 0.905126 0.947804 0.990616 "
  "1.03356 1.07664 1.11986 1.16321 1.20670 1.25032 1.29408 1.33797 1.38200 "
  "1.42616 1.47045 1.51488 1.55945 1.60414 1.64898 1.69394 1.73904 1.78427 "
  "1.82963 1.87513 1.92076 1.96652 2.01241 2.05843 2.10458 2.15086 2.19727 "
  "2.24380 2.29047 2.33726 2.38417 2.43122 2.47838 2.52567 2.57309 2.62062 "
  "2.66828 2.71605 2.76395 2.81196 2.86009 2.90833 2.95669 3.00517 3.05375 "
  "3.10245 3.15125 3.20016 3.24918 3.29830 3.34753 3.39686 3.44629 3.49582 "
  "3.54544 3.59516 3.64497 3.69487 3.74487 3.79495 3.84511 3.89537 3.94570 "
  "3.99611 4.04660 4.09716 4.14780 4.19850 4.24928 4.30012 4.35102 4.40199 "
  "4.45302 4.50410 4.55523 4.60642"),
 ("H2", 303.0): (
  " ".join(str(i) for i in range(0, 101)),
  "0 0.0396708 0.0792956 0.118874 0.158407 0.197894 0.237335 0.276730 0.316080 "
  "0.355384 0.394642 0.433854 0.473021 0.512143 0.551219 0.590250 0.629235 "
  "0.668176 0.707070 0.745920 0.784725 0.823484 0.862199 0.900868 0.939492 "
  "0.978072 1.01661 1.05510 1.09354 1.13194 1.17030 1.20861 1.24687 1.28510 "
  "1.32327 1.36141 1.39950 1.43754 1.47554 1.51349 1.55141 1.58927 1.62710 "
  "1.66487 1.70261 1.74030 1.77795 1.81555 1.85311 1.89062 1.92809 1.96552 "
  "2.00290 2.04024 2.07754 2.11479 2.15200 2.18917 2.22629 2.26337 2.30040 "
  "2.33739 2.37434 2.41125 2.44811 2.48493 2.52170 2.55843 2.59512 2.63177 "
  "2.66837 2.70493 2.74145 2.77793 2.81436 2.85075 2.88709 2.92340 2.95966 "
  "2.99588 3.03205 3.06819 3.10428 3.14033 3.17633 3.21230 3.24822 3.28410 "
  "3.31994 3.35573 3.39148 3.42720 3.46287 3.49849 3.53408 3.56962 3.60512 "
  "3.64058 3.67600 3.71138 3.74671"),
}

NIST_TABLES = {}
for _k, (_pp, _rr) in _NIST_RAW.items():
    _P = np.array([float(x) for x in _pp.split()])
    _R = np.array([float(x) for x in _rr.split()])
    if len(_P) == len(_R):
        NIST_TABLES[_k] = (_P, _R)

NIST_GASES = sorted({g for (g, _T) in NIST_TABLES})
NIST_TEMPS = sorted({T for (_g, T) in NIST_TABLES})


def nist_available(gas, T, tol=0.6):
    """Renvoie la cle (gas, T_table) si une table de reference couvre ce cas."""
    for (g, Tt) in NIST_TABLES:
        if g == gas and abs(Tt - T) <= tol:
            return (g, Tt)
    return None


def nist_density_mmol_cm3(T, P_bar, gas):
    """Densite molaire de reference interpolee (mmol/cm3 == mol/L).
    Renvoie nan hors table (extrapolation volontairement interdite)."""
    key = nist_available(gas, T)
    if key is None:
        return float("nan")
    Pt, Rt = NIST_TABLES[key]
    P = float(P_bar)
    if P < 0 or P > Pt[-1] * 1.0000001:
        return float("nan")
    return float(np.interp(min(P, Pt[-1]), Pt, Rt))

# =============================================================================
# 2. EQUATION D'ETAT (PENG-ROBINSON) ET DENSITES DE GAZ
# =============================================================================

def pr_compressibility(T, P_bar, Tc, Pc_bar, omega):
    """Facteur de compressibilite Z de la phase vapeur (Peng-Robinson, 1976)."""
    if P_bar <= 0:
        return 1.0
    Tr = T / Tc
    if omega <= 0.49:
        kappa = 0.37464 + 1.54226 * omega - 0.26992 * omega ** 2
    else:
        kappa = 0.379642 + 1.48503 * omega - 0.164423 * omega ** 2 \
                + 0.016666 * omega ** 3
    alpha = (1.0 + kappa * (1.0 - math.sqrt(Tr))) ** 2
    Pc = Pc_bar * 1e5
    P = P_bar * 1e5
    a = 0.45723553 * R_J ** 2 * Tc ** 2 / Pc * alpha
    b = 0.07779607 * R_J * Tc / Pc
    A = a * P / (R_J * T) ** 2
    B = b * P / (R_J * T)
    coeffs = [1.0, -(1.0 - B), A - 3.0 * B ** 2 - 2.0 * B,
              -(A * B - B ** 2 - B ** 3)]
    roots = np.roots(coeffs)
    real = roots[np.abs(roots.imag) < 1e-8].real
    real = real[real > B]
    if real.size == 0:
        return 1.0
    return float(np.max(real))       # racine vapeur


def pr_molar_volume(T, P_bar, Tc, Pc_bar, omega, translate=True):
    """Volume molaire (m3/mol) par Peng-Robinson avec translation de Peneloux."""
    Z = pr_compressibility(T, P_bar, Tc, Pc_bar, omega)
    V = Z * R_J * T / (P_bar * 1e5)
    if translate:
        Zra = 0.29056 - 0.08775 * omega
        c = 0.40768 * (R_J * Tc / (Pc_bar * 1e5)) * (0.29441 - Zra)
        V = V - c
    return max(V, 1e-9)


def gas_density_mmol_cm3(T, P_bar, gas, eos="Peng-Robinson"):
    if eos == "Table NIST":
        v = nist_density_mmol_cm3(T, P_bar, gas)
        if np.isfinite(v):
            return v
        eos = "Peng-Robinson"   # repli automatique hors table

    """Densite molaire du gaz libre, en mmol/cm3.

    eos : "Gaz parfait" | "Peng-Robinson" | "CoolProp"
    """
    P_bar = float(P_bar)
    if P_bar <= 0:
        return 0.0
    if eos == "CoolProp" and HAS_COOLPROP and gas in COOLPROP_NAMES:
        try:
            rho = PropsSI("Dmolar", "T", T, "P", P_bar * 1e5,
                          COOLPROP_NAMES[gas])      # mol/m3
            return rho * 1e-3                        # -> mmol/cm3
        except Exception:
            pass
    if eos == "Gaz parfait" or gas not in GAS_DB:
        rho_mol_m3 = (P_bar * 1e5) / (R_J * T)
    else:
        g = GAS_DB[gas]
        rho_mol_m3 = 1.0 / pr_molar_volume(T, P_bar, g["Tc"], g["Pc"], g["w"])
    return rho_mol_m3 * 1e-3                        # mol/m3 -> mmol/cm3


def eos_z_effective(T, P_bar, gas, eos="Peng-Robinson"):
    """Facteur de compressibilite effectif reellement utilise pour rho."""
    if P_bar <= 0:
        return 1.0
    rho = gas_density_mmol_cm3(T, P_bar, gas, eos)      # mmol/cm3 = kmol/m3
    if rho <= 0:
        return 1.0
    return (P_bar * 1e5) / (rho * 1e3 * R_J * T)


def near_critical(gas, T, P_bar):
    """Signale un etat ou les EOS cubiques sont peu fiables (+/- 10-25 %)."""
    if gas not in GAS_DB:
        return False, ""
    g = GAS_DB[gas]
    Tr = T / g["Tc"]
    Pr = P_bar / g["Pc"]
    if 0.85 < Tr < 1.25 and Pr > 0.4:
        return True, ("Etat quasi-critique (Tr=%.2f, Pr=%.2f) : l'ecart de "
                      "Peng-Robinson sur la densite peut atteindre 10 a 25 %%. "
                      "Utilisez CoolProp (pip install CoolProp) ou la methode "
                      "\"Densite fournie\" avec des valeurs NIST/REFPROP."
                      % (Tr, Pr))
    return False, "Tr=%.2f, Pr=%.2f" % (Tr, Pr)


def excess_to_total(P_bar, n_exc, T, gas, method="Volume poreux",
                    V_pore=0.0, rho_ads=None, eos="Peng-Robinson",
                    rho_gas_user=None):
    """Conversion quantite en exces -> quantite totale (absolue).

    Methodes :
      "Volume poreux"     : n_tot = n_exc + V_pore * rho_gaz(T,P)
      "Densite adsorbee"  : n_tot = n_exc / (1 - rho_gaz/rho_ads)
      "Densite fournie"   : idem "Volume poreux" avec rho_gaz donne par l'utilisateur
    Unites : P (bar), n (mmol/g), V_pore (cm3/g), rho_ads (mmol/cm3)
    """
    P_bar = np.asarray(P_bar, dtype=float)
    n_exc = np.asarray(n_exc, dtype=float)

    if method == "Densite fournie" and rho_gas_user is not None:
        rho = np.asarray(rho_gas_user, dtype=float)
    else:
        rho = np.array([gas_density_mmol_cm3(T, p, gas, eos) for p in P_bar])

    if method == "Densite adsorbee":
        if not rho_ads or rho_ads <= 0:
            raise ValueError("Densite de phase adsorbee invalide.")
        denom = 1.0 - rho / rho_ads
        denom = np.where(np.abs(denom) < 1e-9, np.nan, denom)
        return n_exc / denom, rho
    # methodes volumetriques
    return n_exc + V_pore * rho, rho


# =============================================================================
# 3. MODELES D'ISOTHERMES
# =============================================================================
# Chaque modele est decrit par :
#   func(P, ctx, *params)   -> q (mmol/g)
#   pi  (P, ctx, *params)   -> pression d'etalement reduite pi*A/RT (mmol/g)
#                              (None => integration numerique)
#   pinv(pi, ctx, *params)  -> P0 tel que pi(P0) = pi  (None => inversion num.)
#   guess(P, q, ctx)        -> valeurs initiales
#   bounds(P, q, ctx)       -> (bornes_inf, bornes_sup)
#   needs                   -> parametres de contexte requis ('T', 'P0')
# ctx = {'T': temperature (K), 'P0': pression de saturation / reference (bar)}

_EPS = 1e-300


def _safe_P(P):
    return np.maximum(np.atleast_1d(np.asarray(P, dtype=float)), 0.0)


# ---------------------------------------------------------------- Langmuir --
def f_langmuir(P, ctx, qm, b):
    P = _safe_P(P)
    return qm * b * P / (1.0 + b * P)


def pi_langmuir(P, ctx, qm, b):
    P = _safe_P(P)
    return qm * np.log1p(b * P)


def pinv_langmuir(pi, ctx, qm, b):
    if qm <= 0 or b <= 0:
        return np.nan
    return np.expm1(min(pi / qm, 700.0)) / b


def g_langmuir(P, q, ctx):
    qm = max(np.max(q) * 1.05, 1e-6)
    b = _initial_slope(P, q) / qm if qm > 0 else 0.1
    return [qm, max(b, 1e-4)]


def b_langmuir(P, q, ctx):
    return ([0.0, 0.0], [np.inf, np.inf])


# ------------------------------------------------- Langmuir bi-site (DSL) ---
def f_dsl(P, ctx, qm1, b1, qm2, b2):
    P = _safe_P(P)
    return qm1 * b1 * P / (1.0 + b1 * P) + qm2 * b2 * P / (1.0 + b2 * P)


def pi_dsl(P, ctx, qm1, b1, qm2, b2):
    P = _safe_P(P)
    return qm1 * np.log1p(b1 * P) + qm2 * np.log1p(b2 * P)


def g_dsl(P, q, ctx):
    qm = max(np.max(q) * 1.05, 1e-6)
    b = max(_initial_slope(P, q) / qm, 1e-4)
    return [0.5 * qm, 10.0 * b, 0.6 * qm, 0.1 * b]


def b_dsl(P, q, ctx):
    return ([0, 0, 0, 0], [np.inf] * 4)


# -------------------------------------------------------------- Freundlich --
def f_freundlich(P, ctx, KF, nF):
    P = _safe_P(P)
    return KF * np.power(P, nF, where=P > 0, out=np.zeros_like(P))


def pi_freundlich(P, ctx, KF, nF):
    P = _safe_P(P)
    return KF * np.power(P, nF, where=P > 0, out=np.zeros_like(P)) / nF


def pinv_freundlich(pi, ctx, KF, nF):
    if KF <= 0 or nF <= 0:
        return np.nan
    return (nF * pi / KF) ** (1.0 / nF)


def g_freundlich(P, q, ctx):
    try:
        m, c = np.polyfit(np.log(P[P > 0]), np.log(np.maximum(q[P > 0], 1e-12)), 1)
        return [float(np.exp(c)), float(np.clip(m, 0.05, 3.0))]
    except Exception:
        return [max(np.max(q), 1e-3), 0.5]


def b_freundlich(P, q, ctx):
    return ([0.0, 1e-4], [np.inf, 5.0])


# -------------------------------------------------------------------- Sips --
def f_sips(P, ctx, qm, b, ns):
    P = _safe_P(P)
    bp = b * P
    bpn = np.power(bp, ns, where=bp > 1e-300, out=np.zeros_like(bp))
    return qm * bpn / (1.0 + bpn)


def pi_sips(P, ctx, qm, b, ns):
    P = _safe_P(P)
    bp = b * P
    bpn = np.power(bp, ns, where=bp > 1e-300, out=np.zeros_like(bp))
    return (qm / ns) * np.log1p(bpn)


def pinv_sips(pi, ctx, qm, b, ns):
    if qm <= 0 or b <= 0 or ns <= 0:
        return np.nan
    val = np.expm1(min(ns * pi / qm, 700.0))
    return (val ** (1.0 / ns)) / b


def g_sips(P, q, ctx):
    qm = max(np.max(q) * 1.05, 1e-6)
    b = max(_initial_slope(P, q) / qm, 1e-4)
    return [qm, b, 0.9]


def b_sips(P, q, ctx):
    return ([0.0, 0.0, 1e-3], [np.inf, np.inf, 5.0])


# -------------------------------------------------------------------- Toth --
def f_toth(P, ctx, qm, b, t):
    P = _safe_P(P)
    bp = b * P
    bpt = np.power(bp, t, where=bp > 1e-300, out=np.zeros_like(bp))
    denom = np.power(1.0 + bpt, 1.0 / t)
    return qm * bp / denom


def g_toth(P, q, ctx):
    qm = max(np.max(q) * 1.05, 1e-6)
    b = max(_initial_slope(P, q) / qm, 1e-4)
    return [qm, b, 0.9]


def b_toth(P, q, ctx):
    return ([0.0, 0.0, 1e-3], [np.inf, np.inf, 3.0])


# ------------------------------------------- Dubinin-Astakhov / Radushkevich
def _da_core(P, T, P0, qm, E, n):
    """q = qm * exp( -( R T ln(P0/P) / E )^n ),  q = qm si P >= P0."""
    P = np.atleast_1d(np.asarray(P, dtype=float))
    out = np.zeros_like(P)
    valid = P > 0
    below = valid & (P < P0)
    above = valid & (P >= P0)
    if np.any(below):
        A = R_J * T * np.log(P0 / P[below])
        u = np.clip((A / E) ** n, 0.0, 700.0)
        out[below] = qm * np.exp(-u)
    out[above] = qm
    return out


def _pi_da_core(P, T, P0, qm, E, n):
    """pi* = qm E /(R T n) * Gamma(1/n) * Q(1/n, u^n),  u = R T ln(P0/P)/E."""
    P = np.atleast_1d(np.asarray(P, dtype=float))
    out = np.zeros_like(P)
    pref = qm * E / (R_J * T * n) * gamma_fn(1.0 / n)
    below = (P > 0) & (P < P0)
    above = P >= P0
    if np.any(below):
        u = R_J * T * np.log(P0 / P[below]) / E
        out[below] = pref * gammaincc(1.0 / n, np.clip(u ** n, 0.0, 700.0))
    if np.any(above):
        out[above] = pref + qm * np.log(P[above] / P0)
    return out


def f_da(P, ctx, qm, E, n):
    return _da_core(P, ctx["T"], ctx["P0"], qm, E, n)


def pi_da(P, ctx, qm, E, n):
    return _pi_da_core(P, ctx["T"], ctx["P0"], qm, E, n)


def g_da(P, q, ctx):
    return [max(np.max(q) * 1.05, 1e-6), 15000.0, 2.0]


def b_da(P, q, ctx):
    return ([0.0, 100.0, 0.3], [np.inf, 1e6, 6.0])


def f_da_p0(P, ctx, qm, E, n, P0):
    return _da_core(P, ctx["T"], P0, qm, E, n)


def pi_da_p0(P, ctx, qm, E, n, P0):
    return _pi_da_core(P, ctx["T"], P0, qm, E, n)


def g_da_p0(P, q, ctx):
    return [max(np.max(q) * 1.05, 1e-6), 15000.0, 2.0, float(np.max(P)) * 1.5]


def b_da_p0(P, q, ctx):
    return ([0.0, 100.0, 0.3, float(np.max(P)) * 1.001],
            [np.inf, 1e6, 6.0, np.inf])


def f_dr(P, ctx, qm, K):
    E = 1.0 / math.sqrt(max(K, 1e-30))
    return _da_core(P, ctx["T"], ctx["P0"], qm, E, 2.0)


def pi_dr(P, ctx, qm, K):
    E = 1.0 / math.sqrt(max(K, 1e-30))
    return _pi_da_core(P, ctx["T"], ctx["P0"], qm, E, 2.0)


def g_dr(P, q, ctx):
    T, P0 = ctx["T"], ctx["P0"]
    try:
        m = (P > 0) & (P < P0) & (q > 0)
        eps2 = (R_J * T * np.log(P0 / P[m])) ** 2
        s, c = np.polyfit(eps2, np.log(q[m]), 1)
        return [max(float(np.exp(c)), 1e-6), max(-float(s), 1e-12)]
    except Exception:
        return [max(np.max(q) * 1.05, 1e-6), 1e-9]


def b_dr(P, q, ctx):
    return ([0.0, 0.0], [np.inf, np.inf])


def f_dr_p0(P, ctx, qm, K, P0):
    E = 1.0 / math.sqrt(max(K, 1e-30))
    return _da_core(P, ctx["T"], P0, qm, E, 2.0)


def pi_dr_p0(P, ctx, qm, K, P0):
    E = 1.0 / math.sqrt(max(K, 1e-30))
    return _pi_da_core(P, ctx["T"], P0, qm, E, 2.0)


def g_dr_p0(P, q, ctx):
    g = g_dr(P, q, ctx)
    return g + [float(np.max(P)) * 1.5]


def b_dr_p0(P, q, ctx):
    return ([0.0, 0.0, float(np.max(P)) * 1.001], [np.inf, np.inf, np.inf])


def _initial_slope(P, q):
    """Estimation de la constante de Henry (pente initiale)."""
    P = np.asarray(P, float)
    q = np.asarray(q, float)
    idx = np.argsort(P)
    P, q = P[idx], q[idx]
    n = max(2, min(4, len(P)))
    try:
        s = np.polyfit(P[:n], q[:n], 1)[0]
        if s > 0:
            return float(s)
    except Exception:
        pass
    with np.errstate(divide="ignore", invalid="ignore"):
        r = q[P > 0] / P[P > 0]
    r = r[np.isfinite(r)]
    return float(np.max(r)) if r.size else 0.1


class Model:
    def __init__(self, key, label, pnames, punits, func, guess, bounds,
                 pi=None, pinv=None, needs=(), formula="", note=""):
        self.key = key
        self.label = label
        self.pnames = pnames
        self.punits = punits
        self.func = func
        self.guess = guess
        self.bounds = bounds
        self.pi = pi
        self.pinv = pinv
        self.needs = needs
        self.formula = formula
        self.note = note

    @property
    def nparam(self):
        return len(self.pnames)

    def derived(self, params, ctx):
        """Grandeurs derivees utiles (dictionnaire nom -> (valeur, unite))."""
        d = {}
        if self.key in ("dr", "dr_p0"):
            K = params[1]
            if K > 0:
                E = 1.0 / math.sqrt(K)
                d["E (energie caract.)"] = (E / 1000.0, "kJ/mol")
        if self.key in ("da", "da_p0"):
            d["E (energie caract.)"] = (params[1] / 1000.0, "kJ/mol")
        if self.key in ("langmuir", "sips", "toth"):
            d["K_H (Henry, approx.)"] = (params[0] * params[1], "mmol/g/bar")
        if self.key == "dsl":
            d["K_H (Henry)"] = (params[0] * params[1] + params[2] * params[3],
                                "mmol/g/bar")
            d["q_max total"] = (params[0] + params[2], "mmol/g")
        return d


MODELS = {}


def _reg(m):
    MODELS[m.key] = m


_reg(Model("langmuir", "Langmuir", ["q_max", "b"], ["mmol/g", "1/bar"],
           f_langmuir, g_langmuir, b_langmuir, pi_langmuir, pinv_langmuir,
           formula="q = q_max b P / (1 + b P)"))

_reg(Model("dsl", "Langmuir bi-site (DSL)",
           ["q_m1", "b1", "q_m2", "b2"],
           ["mmol/g", "1/bar", "mmol/g", "1/bar"],
           f_dsl, g_dsl, b_dsl, pi_dsl, None,
           formula="q = q_m1 b1 P/(1+b1 P) + q_m2 b2 P/(1+b2 P)"))

_reg(Model("freundlich", "Freundlich", ["K_F", "n_F"],
           ["mmol/g/bar^n", "-"],
           f_freundlich, g_freundlich, b_freundlich,
           pi_freundlich, pinv_freundlich,
           formula="q = K_F P^n_F",
           note="Pas de limite de saturation ; IAST valable seulement si "
                "n_F < 1 et loin de la saturation."))

_reg(Model("sips", "Sips (Langmuir-Freundlich)", ["q_max", "b", "n_s"],
           ["mmol/g", "1/bar", "-"],
           f_sips, g_sips, b_sips, pi_sips, pinv_sips,
           formula="q = q_max (bP)^n_s / (1 + (bP)^n_s)"))

_reg(Model("toth", "Toth", ["q_max", "b", "t"], ["mmol/g", "1/bar", "-"],
           f_toth, g_toth, b_toth, None, None,
           formula="q = q_max b P / (1 + (bP)^t)^(1/t)",
           note="Pression d'etalement calculee par integration numerique."))

_reg(Model("dr", "Dubinin-Radushkevich (P0 fixe)", ["q_m", "K"],
           ["mmol/g", "mol2/J2"],
           f_dr, g_dr, b_dr, pi_dr, None, needs=("T", "P0"),
           formula="q = q_m exp(-K eps^2),  eps = R T ln(P0/P)",
           note="Modele de remplissage de micropores : constante de Henry "
                "nulle => usage de l'IAST a tres basse pression a eviter."))

_reg(Model("dr_p0", "Dubinin-Radushkevich (P0 ajuste)", ["q_m", "K", "P0"],
           ["mmol/g", "mol2/J2", "bar"],
           f_dr_p0, g_dr_p0, b_dr_p0, pi_dr_p0, None, needs=("T",),
           formula="q = q_m exp(-K eps^2),  eps = R T ln(P0/P)"))

_reg(Model("da", "Dubinin-Astakhov (P0 fixe)", ["q_m", "E", "n"],
           ["mmol/g", "J/mol", "-"],
           f_da, g_da, b_da, pi_da, None, needs=("T", "P0"),
           formula="q = q_m exp(-(R T ln(P0/P)/E)^n)",
           note="Constante de Henry nulle => prudence avec l'IAST."))

_reg(Model("da_p0", "Dubinin-Astakhov (P0 ajuste)", ["q_m", "E", "n", "P0"],
           ["mmol/g", "J/mol", "-", "bar"],
           f_da_p0, g_da_p0, b_da_p0, pi_da_p0, None, needs=("T",),
           formula="q = q_m exp(-(R T ln(P0/P)/E)^n)"))

MODEL_ORDER = ["langmuir", "dsl", "freundlich", "sips", "toth",
               "dr", "dr_p0", "da", "da_p0"]


# =============================================================================
# 4. AJUSTEMENT NON LINEAIRE ET STATISTIQUES
# =============================================================================

WEIGHT_MODES = ["Aucune (moindres carres)", "1/q (poids intermediaire)",
                "1/q^2 (erreur relative)"]


def _perturb(p0, lb, ub, rng, span=0.8):
    """Point de depart aleatoire autour de p0 (jitter multiplicatif borne)."""
    out = []
    for i, v in enumerate(p0):
        base = v if abs(v) > 1e-12 else 1.0
        x = base * 10.0 ** rng.uniform(-span, span)
        lo = lb[i] if np.isfinite(lb[i]) else -np.inf
        hi = ub[i] if np.isfinite(ub[i]) else np.inf
        if np.isfinite(lo) and np.isfinite(hi):
            x = float(np.clip(x, lo + 1e-12, hi - 1e-12))
            if not np.isfinite(x):
                x = 0.5 * (lo + hi)
        elif np.isfinite(lo):
            x = max(x, lo + 1e-12)
        elif np.isfinite(hi):
            x = min(x, hi - 1e-12)
        out.append(float(x))
    return out


def fit_isotherm(model, P, q, ctx, p0=None, lower=None, upper=None,
                 weights="Aucune (moindres carres)", maxfev=50000,
                 n_starts=1, n_boot=0, seed=12345, progress=None):
    """Ajuste un modele sur (P, q). Retourne un dictionnaire de resultats.

    n_starts : nombre de departs (1 = depart unique ; >1 = multi-depart
               aleatoire pour echapper aux minima locaux).
    n_boot   : nombre de reechantillonnages bootstrap des residus pour les
               intervalles de confiance (0 = desactive).
    """
    P = np.asarray(P, dtype=float)
    q = np.asarray(q, dtype=float)
    m = np.isfinite(P) & np.isfinite(q) & (P > 0)
    P, q = P[m], q[m]
    if len(P) < model.nparam + 1:
        raise ValueError("Pas assez de points valides (%d) pour %d parametres."
                         % (len(P), model.nparam))

    if p0 is None:
        p0 = model.guess(P, q, ctx)
    lb, ub = model.bounds(P, q, ctx)
    if lower is not None:
        lb = [lb[i] if lower[i] is None else lower[i] for i in range(len(lb))]
    if upper is not None:
        ub = [ub[i] if upper[i] is None else upper[i] for i in range(len(ub))]
    p0 = [float(np.clip(p0[i], lb[i] + 1e-15, ub[i] - 1e-15))
          if np.isfinite(ub[i]) else float(max(p0[i], lb[i] + 1e-15))
          for i in range(len(p0))]

    if weights.startswith("1/q^2"):
        sigma = np.maximum(np.abs(q), 1e-9)
    elif weights.startswith("1/q "):
        sigma = np.sqrt(np.maximum(np.abs(q), 1e-9))
    else:
        sigma = None

    def f(x, *pars):
        return np.asarray(model.func(x, ctx, *pars), dtype=float).ravel()

    def _ssr(pars):
        r = (q - f(P, *pars))
        if sigma is not None:
            r = r / sigma
        return float(np.sum(r ** 2))

    def _run(start):
        po, pc = curve_fit(f, P, q, p0=start, bounds=(lb, ub), sigma=sigma,
                           absolute_sigma=False, maxfev=maxfev)
        return po, pc, _ssr(po)

    rng = np.random.default_rng(seed)
    popt, pcov, best = None, None, np.inf
    n_ok, n_ko = 0, 0
    for it in range(max(1, int(n_starts))):
        start = p0 if it == 0 else _perturb(p0, lb, ub, rng)
        try:
            po, pc, ss = _run(start)
        except Exception:
            n_ko += 1
            continue
        n_ok += 1
        if ss < best * (1.0 - 1e-12):
            popt, pcov, best = po, pc, ss
    if popt is None:
        raise RuntimeError("aucun depart n'a converge (%d essais)"
                           % max(1, int(n_starts)))
    with np.errstate(invalid="ignore"):
        try:
            perr = np.sqrt(np.diag(pcov))
        except Exception:
            perr = np.full(len(popt), np.nan)

    # ------- matrice de correlation des parametres -------------------------
    k_ = len(popt)
    corr = np.full((k_, k_), np.nan)
    try:
        with np.errstate(invalid="ignore"):
            sd = np.sqrt(np.diag(pcov))
        if np.all(np.isfinite(sd)) and np.all(sd > 0):
            corr = pcov / np.outer(sd, sd)
            corr = np.clip(corr, -1.0, 1.0)
    except Exception:
        pass
    off = [abs(corr[i, j]) for i in range(k_) for j in range(i + 1, k_)
           if np.isfinite(corr[i, j])]
    rmax = float(max(off)) if off else np.nan

    qpred = f(P, *popt)
    res = q - qpred
    N, k = len(q), len(popt)
    SSR = float(np.sum(res ** 2))
    SST = float(np.sum((q - np.mean(q)) ** 2))
    R2 = 1.0 - SSR / SST if SST > 1e-15 else (1.0 if SSR < 1e-15 else 0.0)
    R2adj = (1.0 - (1.0 - R2) * (N - 1) / (N - k - 1)) if N - k - 1 > 0 else np.nan
    RMSE = math.sqrt(SSR / N)
    MSSR = SSR / N
    with np.errstate(divide="ignore", invalid="ignore"):
        ARE = np.abs(res / np.where(np.abs(q) > 1e-12, q, np.nan))
    ARE = float(np.nanmean(ARE) * 100.0)
    chi2 = float(np.nansum(res ** 2 / np.where(np.abs(qpred) > 1e-12,
                                               qpred, np.nan)))
    AIC = N * math.log(max(SSR / N, 1e-300)) + 2 * k
    AICc = AIC + (2 * k * (k + 1) / (N - k - 1)) if N - k - 1 > 0 else np.nan
    BIC = N * math.log(max(SSR / N, 1e-300)) + k * math.log(N)

    # ------- intervalles de confiance par bootstrap des residus ------------
    ci_lo = ci_hi = None
    boot_ok = 0
    if n_boot and n_boot >= 20:
        qhat = f(P, *popt)
        draws = []
        for ib in range(int(n_boot)):
            qs = qhat + rng.choice(res, size=len(res), replace=True)
            try:
                pb, _ = curve_fit(f, P, qs, p0=popt, bounds=(lb, ub),
                                  sigma=sigma, absolute_sigma=False,
                                  maxfev=maxfev)
                draws.append(pb)
            except Exception:
                pass
            if progress is not None and (ib + 1) % 25 == 0:
                progress(ib + 1, int(n_boot))
        if len(draws) >= 10:
            D = np.asarray(draws, dtype=float)
            boot_ok = len(draws)
            ci_lo = np.percentile(D, 2.5, axis=0).tolist()
            ci_hi = np.percentile(D, 97.5, axis=0).tolist()
            perr_boot = D.std(axis=0, ddof=1).tolist()
        else:
            perr_boot = None
    else:
        perr_boot = None

    return dict(model_key=model.key, params=[float(v) for v in popt],
                errors=[float(v) for v in perr], ctx=dict(ctx),
                corr=corr.tolist(), r_max=rmax,
                ci_lo=ci_lo, ci_hi=ci_hi, err_boot=perr_boot,
                n_boot=boot_ok, n_starts=int(max(1, n_starts)),
                starts_ok=n_ok, starts_ko=n_ko,
                P_exp_min=float(np.min(P)), P_exp_max=float(np.max(P)),
                stats=dict(N=N, R2=R2, R2adj=R2adj, SSR=SSR, MSSR=MSSR,
                           RMSE=RMSE, ARE=ARE, chi2=chi2, AIC=AIC,
                           AICc=AICc, BIC=BIC),
                P=P.tolist(), q=q.tolist(), qpred=qpred.tolist(),
                residuals=res.tolist(), weights=weights)


# =============================================================================
# 5. PRESSION D'ETALEMENT ET IAST
# =============================================================================

class SpreadingPressure:
    """pi* = (pi A)/(R T) = integrale de 0 a P de q(p)/p dp   [mmol/g]."""

    def __init__(self, model, ctx, params, P_lo=1e-10, P_hi=1e5, ngrid=2000,
                 P_exp_max=None):
        self.model = model
        self.ctx = dict(ctx) if ctx else {}
        self.params = [float(v) for v in params]
        # pression experimentale maximale du corps pur : sert a signaler
        # l'extrapolation implicite du modele lors du calcul IAST
        self.P_exp_max = (float(P_exp_max)
                          if P_exp_max and np.isfinite(P_exp_max) and P_exp_max > 0
                          else None)
        self._table = None
        if model.pi is None:
            self._build_table(P_lo, P_hi, ngrid)

    # -- integrande en y = ln(p) : q(e^y) --------------------------------
    def _integrand(self, y):
        v = self.model.func(np.exp(y), self.ctx, *self.params)
        return float(np.asarray(v, dtype=float).ravel()[0])

    def _build_table(self, P_lo, P_hi, ngrid):
        ygrid = np.linspace(math.log(P_lo), math.log(P_hi), ngrid)
        pig = np.zeros(ngrid)
        acc = quad(self._integrand, ygrid[0] - 34.0, ygrid[0], limit=200)[0]
        pig[0] = acc
        for i in range(1, ngrid):
            acc += quad(self._integrand, ygrid[i - 1], ygrid[i], limit=200)[0]
            pig[i] = acc
        self._table = (ygrid, np.maximum.accumulate(pig))

    def value(self, P):
        P = np.atleast_1d(np.asarray(P, dtype=float))
        if self.model.pi is not None:
            out = np.asarray(self.model.pi(P, self.ctx, *self.params),
                             dtype=float)
            return np.atleast_1d(out)
        ygrid, pig = self._table
        y = np.log(np.maximum(P, 1e-300))
        out = np.interp(y, ygrid, pig)
        # correction locale exacte : integration depuis le noeud inferieur
        inside = (y > ygrid[0]) & (y <= ygrid[-1])
        if np.any(inside):
            idx = np.searchsorted(ygrid, y[inside]) - 1
            idx = np.clip(idx, 0, len(ygrid) - 1)
            vals = []
            for k, i0 in enumerate(idx):
                y1 = float(np.atleast_1d(y[inside])[k])
                vals.append(pig[i0] + quad(self._integrand, ygrid[i0], y1,
                                           limit=100)[0])
            out[inside] = vals
        hi = y > ygrid[-1]
        if np.any(hi):
            base = pig[-1]
            out[hi] = [base + quad(self._integrand, ygrid[-1], float(v),
                                   limit=200)[0] for v in np.atleast_1d(y[hi])]
        lo = y < ygrid[0]
        if np.any(lo):
            out[lo] = pig[0] * np.exp(y[lo] - ygrid[0])
        return out

    def inverse(self, pi):
        """Retourne P0 tel que pi*(P0) = pi."""
        pi = float(pi)
        if pi <= 0:
            return 1e-300
        if self.model.pinv is not None:
            v = self.model.pinv(pi, self.ctx, *self.params)
            if np.isfinite(v) and v > 0:
                return float(v)
        if self.model.pi is None:
            ygrid, pig = self._table
            if pi <= pig[-1]:
                lp = float(np.interp(pi, pig, ygrid))
            else:
                slope = max((pig[-1] - pig[-2]) / (ygrid[-1] - ygrid[-2]), 1e-12)
                lp = float(ygrid[-1] + (pi - pig[-1]) / slope)
            # affinage de Newton : d(pi*)/d ln P = q(P)
            for _ in range(2):
                Pk = math.exp(lp)
                dq = float(self.q(Pk)[0])
                if dq <= 1e-14:
                    break
                lp -= (float(self.value(Pk)[0]) - pi) / dq
            return float(math.exp(lp))
        # inversion numerique d'une expression analytique de pi
        g = lambda t: float(self.value(math.exp(t))[0]) - pi
        t_hi, t_lo = 0.0, 0.0
        it = 0
        while g(t_hi) < 0 and it < 400:
            t_hi += 1.0
            it += 1
        it = 0
        while g(t_lo) > 0 and it < 400:
            t_lo -= 1.0
            it += 1
        try:
            return float(math.exp(brentq(g, t_lo, t_hi, xtol=1e-12,
                                         rtol=1e-12, maxiter=300)))
        except Exception:
            return np.nan

    def q(self, P):
        return np.atleast_1d(np.asarray(
            self.model.func(P, self.ctx, *self.params), dtype=float))


def iast_point(P_total, y, comps):
    """IAST pour un point de pression totale.

    comps : liste de SpreadingPressure (un par constituant)
    y     : fractions molaires en phase gaz (somme = 1)
    Retourne un dict : P0, x, q, q_total, pi, converged
    """
    n = len(comps)
    y = np.asarray(y, dtype=float)
    p_part = y * P_total
    act = [i for i in range(n) if y[i] > 1e-12]

    out = dict(P0=np.full(n, np.nan), x=np.zeros(n), q=np.zeros(n),
               q_total=np.nan, pi=np.nan, converged=False, message="")

    if P_total <= 0 or not act:
        out.update(converged=True, q_total=0.0, pi=0.0)
        return out

    def sum_x(t):
        pi = math.exp(t)
        s = 0.0
        for i in act:
            P0 = comps[i].inverse(pi)
            if not np.isfinite(P0) or P0 <= 0:
                return np.inf
            s += p_part[i] / P0
        return s

    g = lambda t: sum_x(t) - 1.0

    # encadrement de ln(pi)
    t_hi = 0.0
    it = 0
    while g(t_hi) > 0 and it < 500:
        t_hi += 0.5
        it += 1
    t_lo = t_hi
    it = 0
    while g(t_lo) < 0 and it < 500:
        t_lo -= 0.5
        it += 1
    if g(t_lo) < 0 or g(t_hi) > 0:
        out["message"] = "Encadrement de la pression d'etalement impossible."
        return out
    try:
        t = brentq(g, t_lo, t_hi, xtol=1e-13, rtol=1e-13, maxiter=400)
    except Exception as e:
        out["message"] = "Echec du solveur : %s" % e
        return out

    pi = math.exp(t)
    P0 = np.full(n, np.nan)
    x = np.zeros(n)
    for i in act:
        P0[i] = comps[i].inverse(pi)
        x[i] = p_part[i] / P0[i]
    sx = x.sum()
    if sx > 1e-12:
        x = x / sx

    inv = 0.0
    ok = True
    for i in act:
        q0 = float(comps[i].q(P0[i])[0])
        if q0 <= 1e-12:
            ok = False
            break
        inv += x[i] / q0
    if not ok or inv <= 0:
        out["message"] = "Quantite adsorbee pure nulle a P0."
        return out
    q_tot = 1.0 / inv
    out.update(P0=P0, x=x, q=x * q_tot, q_total=q_tot, pi=pi, converged=True)
    return out


def iast_scan(P_totals, y, comps, names):
    """Balayage en pression totale. Retourne un DataFrame."""
    rows = []
    n = len(comps)
    for Pt in np.atleast_1d(np.asarray(P_totals, dtype=float)):
        r = iast_point(Pt, y, comps)
        ratios = []
        for i in range(n):
            pm = getattr(comps[i], "P_exp_max", None)
            ratios.append(r["P0"][i] / pm if (pm and np.isfinite(r["P0"][i]))
                          else np.nan)
        rmax = np.nanmax(ratios) if np.any(np.isfinite(ratios)) else np.nan
        row = {"P_total (bar)": Pt, "pi* (mmol/g)": r["pi"],
               "q_total (mmol/g)": r["q_total"],
               "Convergence": bool(r["converged"]),
               "Extrapolation max (P0/P_exp)": rmax}
        for i, nm in enumerate(names):
            row["y_%s" % nm] = y[i]
            row["p_%s (bar)" % nm] = y[i] * Pt
            row["P0_%s (bar)" % nm] = r["P0"][i]
            row["P0/P_exp_%s" % nm] = ratios[i]
            row["x_%s" % nm] = r["x"][i]
            row["q_%s (mmol/g)" % nm] = r["q"][i]
            # quantite pure a la pression partielle (reference sans competition)
            try:
                row["q_%s pur (mmol/g)" % nm] = float(comps[i].q(y[i] * Pt)[0])
            except Exception:
                row["q_%s pur (mmol/g)" % nm] = np.nan
        for i in range(n):
            for j in range(n):
                if i >= j:
                    continue
                si = selectivity(r["x"], y, i, j)
                row["S_%s/%s" % (names[i], names[j])] = si
                row["S_%s/%s" % (names[j], names[i])] = (
                    1.0 / si if (np.isfinite(si) and abs(si) > 1e-30)
                    else (0.0 if si == np.inf else np.nan))
        rows.append(row)
    return pd.DataFrame(rows)


def iast_scan_composition(P_total, y1_values, comps, names, y_rest=None):
    """Balayage en composition a pression totale fixee (binaire ou ternaire).

    y_rest : pour un ternaire, rapport fixe entre les constituants 2 et 3.
    """
    rows = []
    n = len(comps)
    for y1 in np.atleast_1d(np.asarray(y1_values, dtype=float)):
        y = np.zeros(n)
        y[0] = y1
        if n == 2:
            y[1] = 1.0 - y1
        else:
            rest = 1.0 - y1
            if y_rest is None:
                y[1:] = rest / (n - 1)
            else:
                w = np.asarray(y_rest, dtype=float)
                w = w / w.sum()
                y[1:] = rest * w
        r = iast_point(P_total, y, comps)
        ratios = []
        for i in range(n):
            pm = getattr(comps[i], "P_exp_max", None)
            ratios.append(r["P0"][i] / pm if (pm and np.isfinite(r["P0"][i]))
                          else np.nan)
        row = {"P_total (bar)": P_total, "q_total (mmol/g)": r["q_total"],
               "pi* (mmol/g)": r["pi"], "Convergence": bool(r["converged"]),
               "Extrapolation max (P0/P_exp)":
                   (np.nanmax(ratios) if np.any(np.isfinite(ratios)) else np.nan)}
        for i, nm in enumerate(names):
            row["y_%s" % nm] = y[i]
            row["x_%s" % nm] = r["x"][i]
            row["q_%s (mmol/g)" % nm] = r["q"][i]
            row["P0/P_exp_%s" % nm] = ratios[i]
        for i in range(n):
            for j in range(i + 1, n):
                row["S_%s/%s" % (names[i], names[j])] = selectivity(r["x"], y,
                                                                    i, j)
        rows.append(row)
    return pd.DataFrame(rows)


def selectivity(x, y, i, j):
    """S_i/j = (x_i/x_j) / (y_i/y_j)."""
    try:
        if y[i] <= 0 or y[j] <= 0:
            return np.nan
        if x[j] <= 1e-15:
            return np.inf if x[i] > 1e-15 else np.nan
        return float((x[i] / y[i]) / (x[j] / y[j]))
    except Exception:
        return np.nan


# =============================================================================
# 6. UTILITAIRES D'INTERFACE
# =============================================================================

def to_float(s):
    """Conversion robuste texte -> float (accepte la virgule decimale)."""
    if s is None:
        return np.nan
    if isinstance(s, (int, float, np.floating, np.integer)):
        return float(s)
    t = str(s).strip().replace(" ", "").replace(" ", "")
    if not t:
        return np.nan
    if "," in t and "." in t:
        t = t.replace(",", "")           # virgule = separateur de milliers
    else:
        t = t.replace(",", ".")
    t = t.replace("−", "-")
    try:
        return float(t)
    except ValueError:
        return np.nan


def to_float_or(x, default):
    """to_float avec valeur de repli (NaN et texte invalide -> default)."""
    v = to_float(x)
    try:
        return default if (v is None or not np.isfinite(v)) else float(v)
    except (TypeError, ValueError):
        return default


def fmt(v, nd=5):
    if v is None:
        return ""
    try:
        v = float(v)
    except (TypeError, ValueError):
        return str(v)
    if not np.isfinite(v):
        return "inf" if v > 0 else ("-inf" if v < 0 else "nan")
    if v != 0 and (abs(v) < 1e-3 or abs(v) >= 1e5):
        return ("%%.%de" % nd) % v
    return ("%%.%df" % nd) % v


class EditableTable(ttk.Frame):
    """Grille editable facon tableur (double-clic, Ctrl+V, Ctrl+C, Suppr)."""

    def __init__(self, master, columns, height=14, widths=None, nrows=0):
        super().__init__(master)
        self.columns = list(columns)
        self._editor = None
        vsb = ttk.Scrollbar(self, orient="vertical")
        hsb = ttk.Scrollbar(self, orient="horizontal")
        self.tree = ttk.Treeview(self, columns=self.columns, show="headings",
                                 height=height, selectmode="extended",
                                 yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        vsb.config(command=self.tree.yview)
        hsb.config(command=self.tree.xview)
        for i, c in enumerate(self.columns):
            w = (widths[i] if widths and i < len(widths) else 110)
            self.tree.heading(c, text=c)
            self.tree.column(c, width=w, anchor="e", stretch=True)
        self.tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        self.rowconfigure(0, weight=1)
        self.columnconfigure(0, weight=1)

        self.tree.bind("<Double-1>", self._begin_edit)
        self.tree.bind("<Return>", self._begin_edit)
        self.tree.bind("<Control-v>", self._paste)
        self.tree.bind("<Control-V>", self._paste)
        self.tree.bind("<Control-c>", self._copy)
        self.tree.bind("<Control-C>", self._copy)
        self.tree.bind("<Delete>", lambda e: self.delete_selected())
        for _ in range(nrows):
            self.add_row()

    # ------------------------------------------------------------ edition --
    def _begin_edit(self, event=None):
        self._close_editor()
        if event is not None and getattr(event, "x", None) is not None:
            row = self.tree.identify_row(event.y)
            col = self.tree.identify_column(event.x)
        else:
            sel = self.tree.selection()
            row = sel[0] if sel else None
            col = "#1"
        if not row or not col:
            return "break"
        try:
            cidx = int(col.replace("#", "")) - 1
        except ValueError:
            return "break"
        if cidx < 0 or cidx >= len(self.columns):
            return "break"
        bbox = self.tree.bbox(row, col)
        if not bbox:
            return "break"
        x, y, w, h = bbox
        val = self.tree.set(row, self.columns[cidx])
        ed = tk.Entry(self.tree, justify="right")
        ed.insert(0, val)
        ed.select_range(0, "end")
        ed.place(x=x, y=y, width=w, height=h)
        ed.focus_set()
        self._editor = (ed, row, cidx)
        ed.bind("<Return>", lambda e: self._commit(move=1))
        ed.bind("<Tab>", lambda e: self._commit(move=2))
        ed.bind("<Escape>", lambda e: self._close_editor())
        ed.bind("<FocusOut>", lambda e: self._commit(move=0))
        return "break"

    def _fire_change(self):
        cb = getattr(self, "on_change", None)
        if callable(cb):
            try:
                cb()
            except Exception:
                pass

    def _commit(self, move=0):
        if not self._editor:
            return "break"
        ed, row, cidx = self._editor
        try:
            value = ed.get()
        except Exception:
            value = ""
        self._editor = None
        try:
            ed.destroy()
        except Exception:
            pass
        try:
            self.tree.set(row, self.columns[cidx], value)
        except Exception:
            return "break"
        if move == 1:
            nxt = self.tree.next(row)
            if not nxt:
                nxt = self.add_row()
            self.tree.selection_set(nxt)
            self.tree.focus(nxt)
            self.tree.see(nxt)
        elif move == 2 and cidx + 1 < len(self.columns):
            self.tree.selection_set(row)
        return "break"
        self._fire_change()

    def _close_editor(self):
        if self._editor:
            try:
                self._editor[0].destroy()
            except Exception:
                pass
            self._editor = None
        return "break"

    # ------------------------------------------------- copier / coller -----
    def _copy(self, event=None):
        rows = self.tree.selection() or self.tree.get_children()
        txt = "\n".join("\t".join(self.tree.set(r, c) for c in self.columns)
                        for r in rows)
        self.clipboard_clear()
        self.clipboard_append(txt)
        return "break"

    def _paste(self, event=None):
        try:
            data = self.clipboard_get()
        except Exception:
            return "break"
        block = parse_clipboard_block(data)
        if not block:
            return "break"
        sel = self.tree.selection()
        items = list(self.tree.get_children())
        start = items.index(sel[0]) if sel and sel[0] in items else len(items)
        for r, vals in enumerate(block):
            idx = start + r
            while idx >= len(self.tree.get_children()):
                self.add_row()
            item = self.tree.get_children()[idx]
            for c, v in enumerate(vals[:len(self.columns)]):
                self.tree.set(item, self.columns[c], v)
        return "break"

    # ------------------------------------------------------------ donnees --
    def add_row(self, values=None):
        vals = list(values) if values else [""] * len(self.columns)
        vals = (vals + [""] * len(self.columns))[:len(self.columns)]
        return self.tree.insert("", "end", values=vals)
        self._fire_change()

    def delete_selected(self):
        for r in self.tree.selection():
            self.tree.delete(r)
        self._fire_change()

    def clear(self):
        for r in self.tree.get_children():
            self.tree.delete(r)
        self._fire_change()

    def get_rows(self):
        return [[self.tree.set(r, c) for c in self.columns]
                for r in self.tree.get_children()]

    def set_rows(self, rows):
        self.clear()
        for r in rows:
            self.add_row(["" if v is None else
                          (fmt(v, 6) if isinstance(v, float) else str(v))
                          for v in r])
        self._fire_change()

    def get_numeric(self, ncols=None):
        ncols = ncols or len(self.columns)
        out = []
        for row in self.get_rows():
            vals = [to_float(row[i]) for i in range(ncols)]
            if all(np.isnan(v) for v in vals):
                continue
            out.append(vals)
        return np.array(out, dtype=float) if out else np.zeros((0, ncols))

    def sort_by(self, col_index=0):
        rows = self.get_rows()
        rows.sort(key=lambda r: (np.isnan(to_float(r[col_index])),
                                 to_float(r[col_index])))
        self.set_rows(rows)
        self._fire_change()

def parse_clipboard_block(data):
    """Decoupe un bloc colle depuis Excel/texte en liste de listes."""
    lines = [l for l in str(data).replace("\r\n", "\n").replace("\r", "\n").split("\n")]
    block = []
    for line in lines:
        if not line.strip():
            continue
        if "\t" in line:
            cells = line.split("\t")
        elif ";" in line:
            cells = line.split(";")
        elif "," in line and line.count(",") > 1:
            cells = line.split(",")
        else:
            cells = line.split()
        block.append([c.strip() for c in cells])
    return block


class ScrollFrame(ttk.Frame):
    """Cadre defilant vertical et horizontal, avec molette de souris.

    Le contenu n'est jamais tronque : si le panneau est plus etroit que les
    widgets, une barre de defilement horizontale apparait.
    """

    def __init__(self, master, **kw):
        super().__init__(master, **kw)
        canvas = tk.Canvas(self, highlightthickness=0, borderwidth=0)
        try:
            canvas.configure(background=UI["bg"])
        except Exception:
            pass
        vsb = ttk.Scrollbar(self, orient="vertical", command=canvas.yview)
        hsb = ttk.Scrollbar(self, orient="horizontal", command=canvas.xview)
        self.inner = ttk.Frame(canvas)
        self._win = canvas.create_window((0, 0), window=self.inner, anchor="nw")

        def _content(_=None):
            try:
                canvas.configure(scrollregion=canvas.bbox("all"))
            except Exception:
                pass

        def _resize(e):
            try:
                canvas.itemconfigure(self._win,
                                     width=max(e.width,
                                               self.inner.winfo_reqwidth()))
            except Exception:
                pass
            _content()

        self.inner.bind("<Configure>", _content)
        canvas.bind("<Configure>", _resize)
        canvas.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        canvas.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        self.rowconfigure(0, weight=1)
        self.columnconfigure(0, weight=1)

        def _wheel(e):
            try:
                d = -1 if getattr(e, "delta", 0) > 0 else 1
                canvas.yview_scroll(d, "units")
            except Exception:
                pass
            return "break"

        def _bind(_=None):
            canvas.bind_all("<MouseWheel>", _wheel)

        def _unbind(_=None):
            try:
                canvas.unbind_all("<MouseWheel>")
            except Exception:
                pass

        canvas.bind("<Enter>", _bind)
        canvas.bind("<Leave>", _unbind)
        self.canvas = canvas


def set_sash(pan, index, pos):
    """Fixe la position d'un separateur une fois la fenetre affichee."""
    def _do():
        try:
            pan.sashpos(index, int(pos))
        except Exception:
            pass
    try:
        pan.after(150, _do)
        pan.after(600, _do)
    except Exception:
        pass


class FigureWindow(tk.Toplevel):
    """Fenetre independante et redimensionnable pour agrandir un graphique."""

    def __init__(self, master, title, render, size=(12.0, 7.5), app=None):
        super().__init__(master)
        self.title(title)
        self.geometry("1250x820")
        self.app = app
        self.render = render
        try:
            self.configure(bg=UI["bg"])
        except Exception:
            pass
        bar = ttk.Frame(self, style="Bar.TFrame", padding=(10, 6))
        bar.pack(fill="x")
        ttk.Label(bar, text=title, style="Head.TLabel").pack(side="left")
        ttk.Button(bar, text="Actualiser", style="Ghost.TButton",
                   command=self.redraw).pack(side="right", padx=4)
        ttk.Button(bar, text="Enregistrer l'image...", style="Primary.TButton",
                   command=self.save).pack(side="right")
        body = ttk.Frame(self, padding=8)
        body.pack(fill="both", expand=True)
        holder = ttk.Frame(body, style="Card.TFrame", padding=8)
        holder.pack(fill="both", expand=True)
        self.fig = Figure(figsize=size, dpi=100, facecolor=UI["card"])
        self.canvas = FigureCanvasTkAgg(self.fig, master=holder)
        self.canvas.get_tk_widget().pack(fill="both", expand=True)
        if NavigationToolbar2Tk is not None:
            NavigationToolbar2Tk(self.canvas, holder).update()
        self.redraw()

    def redraw(self):
        try:
            self.render(self.fig)
            self.canvas.draw_idle()
        except Exception as e:
            messagebox.showerror("Graphique", str(e), parent=self)

    def save(self):
        path = filedialog.asksaveasfilename(
            defaultextension=".png", initialfile="figure.png",
            filetypes=[("PNG", "*.png"), ("PDF vectoriel", "*.pdf"),
                       ("SVG vectoriel", "*.svg"), ("TIFF", "*.tif"),
                       ("JPEG", "*.jpg")])
        if not path:
            return
        try:
            self.fig.savefig(path, dpi=300, bbox_inches="tight",
                             facecolor="white")
        except Exception as e:
            messagebox.showerror("Export", str(e), parent=self)
            return
        if self.app:
            self.app.log("Figure exportee : %s" % path)
        messagebox.showinfo("Export", "Figure enregistree :\n%s" % path,
                            parent=self)


def zoom_button(parent, render, title, app=None, **pack):
    """Bouton normalise 'Agrandir le graphique'."""
    b = ttk.Button(parent, text="Agrandir le graphique",
                   style="Ghost.TButton",
                   command=lambda: FigureWindow(parent.winfo_toplevel(), title,
                                                render, app=app))
    tip(b, "Ouvre le graphique dans une fenetre separee et redimensionnable, "
           "avec zoom, deplacement et export haute resolution")
    return b


MARKERS = ["o", "s", "^", "v", "D", "<", ">", "p", "*", "h", "x", "+", "None"]
LINESTYLES = ["-", "--", "-.", ":", "None"]
PALETTE = ["#1f77b4", "#d62728", "#2ca02c", "#ff7f0e", "#9467bd", "#8c564b",
           "#e377c2", "#7f7f7f", "#bcbd22", "#17becf", "#000000"]
LEGEND_LOCS = ["best", "upper right", "upper left", "lower left",
               "lower right", "center right", "center left",
               "upper center", "lower center", "center"]


# =============================================================================
# 6bis. THEME, WIDGETS ET GESTION DE SESSION  (v2)
# =============================================================================

UI = dict(
    bg="#eef2f6", card="#ffffff", ink="#1b2733", muted="#63707e",
    accent="#0f6fa8", accent_hi="#1583c4", accent_lo="#0b5580",
    ok="#1c8b5a", ok_hi="#22a86d", warn="#b06a00", danger="#b3261e",
    danger_hi="#cf3a31", line="#cfd8e0", zebra="#f7fafc", sel="#d6e9f7",
)
FONT = "Segoe UI"


def apply_theme(root):
    """Theme clair moderne, coherent sur l'ensemble de l'application."""
    st = ttk.Style(root)
    try:
        st.theme_use("clam")
    except Exception:
        pass
    try:
        root.configure(bg=UI["bg"])
    except Exception:
        pass
    base = (FONT, 10)
    st.configure(".", font=base, background=UI["bg"], foreground=UI["ink"])
    st.configure("TFrame", background=UI["bg"])
    st.configure("Card.TFrame", background=UI["card"], relief="flat")
    st.configure("Bar.TFrame", background=UI["card"])
    st.configure("TLabel", background=UI["bg"], foreground=UI["ink"])
    st.configure("Card.TLabel", background=UI["card"], foreground=UI["ink"])
    st.configure("Title.TLabel", font=(FONT, 15, "bold"), background=UI["card"],
                 foreground=UI["ink"])
    st.configure("Head.TLabel", font=(FONT, 10, "bold"), background=UI["card"],
                 foreground=UI["accent_lo"])
    st.configure("Muted.TLabel", font=(FONT, 9), background=UI["card"],
                 foreground=UI["muted"])
    st.configure("MutedBg.TLabel", font=(FONT, 9), background=UI["bg"],
                 foreground=UI["muted"])
    st.configure("Status.TLabel", font=(FONT, 9), background=UI["card"],
                 foreground=UI["muted"])
    st.configure("TLabelframe", background=UI["card"], borderwidth=1,
                 relief="solid", bordercolor=UI["line"])
    st.configure("TLabelframe.Label", font=(FONT, 10, "bold"),
                 background=UI["card"], foreground=UI["accent_lo"])
    st.configure("TButton", padding=(10, 6), relief="flat",
                 background="#e3eaf1", foreground=UI["ink"], borderwidth=0)
    st.map("TButton", background=[("active", "#d2dde7"),
                                  ("pressed", "#c2d0dd"),
                                  ("disabled", "#eceff2")])
    for name, col, hi in (("Primary", UI["accent"], UI["accent_hi"]),
                          ("Success", UI["ok"], UI["ok_hi"]),
                          ("Danger", UI["danger"], UI["danger_hi"])):
        st.configure("%s.TButton" % name, background=col, foreground="white",
                     padding=(14, 7), font=(FONT, 10, "bold"), borderwidth=0,
                     relief="flat")
        st.map("%s.TButton" % name,
               background=[("active", hi), ("pressed", hi),
                           ("disabled", "#b9c4cf")],
               foreground=[("disabled", "#eef2f6")])
    st.configure("Ghost.TButton", background=UI["card"], foreground=UI["accent"],
                 padding=(9, 5), borderwidth=0, relief="flat")
    st.map("Ghost.TButton", background=[("active", UI["sel"])])
    st.configure("Tool.TButton", background=UI["card"], foreground=UI["ink"],
                 padding=(11, 7), borderwidth=0, relief="flat",
                 font=(FONT, 10))
    st.map("Tool.TButton", background=[("active", UI["sel"]),
                                       ("pressed", UI["sel"])])
    st.configure("Step.TLabel", background=UI["card"], foreground=UI["muted"],
                 font=(FONT, 10), padding=(14, 7))
    st.configure("StepOn.TLabel", background=UI["accent"], foreground="white",
                 font=(FONT, 10, "bold"), padding=(14, 7))
    st.configure("StepDone.TLabel", background="#dceaf5",
                 foreground=UI["accent_lo"], font=(FONT, 10), padding=(14, 7))
    st.configure("Treeview", background="white", fieldbackground="white",
                 foreground=UI["ink"], rowheight=25, borderwidth=0,
                 font=(FONT, 9))
    st.configure("Treeview.Heading", font=(FONT, 9, "bold"), padding=(6, 5),
                 background="#e3eaf1", foreground=UI["ink"], relief="flat")
    st.map("Treeview", background=[("selected", UI["accent"])],
           foreground=[("selected", "white")])
    st.configure("TNotebook", background=UI["bg"], borderwidth=0, tabmargins=0)
    st.configure("TNotebook.Tab", padding=(0, 0), background=UI["bg"],
                 borderwidth=0)
    try:
        st.layout("Hidden.TNotebook.Tab", [])   # onglets masques : navigation
    except Exception:                            # par la barre d'etapes
        pass
    st.configure("TCombobox", padding=4)
    st.configure("TEntry", padding=4, fieldbackground="white")
    st.configure("TCheckbutton", background=UI["card"], foreground=UI["ink"])
    st.configure("CheckBg.TCheckbutton", background=UI["bg"])
    st.configure("TRadiobutton", background=UI["card"])
    st.configure("Horizontal.TProgressbar", background=UI["accent"],
                 troughcolor="#e3eaf1", borderwidth=0)
    st.configure("TPanedwindow", background=UI["bg"])
    st.configure("Sash", sashthickness=8)
    return st


class Tooltip:
    """Info-bulle legere pour n'importe quel widget."""

    def __init__(self, widget, text, delay=450):
        self.widget = widget
        self.text = text
        self.delay = delay
        self.tip = None
        self._id = None
        try:
            widget.bind("<Enter>", self._enter, add="+")
            widget.bind("<Leave>", self._leave, add="+")
            widget.bind("<ButtonPress>", self._leave, add="+")
        except Exception:
            pass

    def _enter(self, _=None):
        self._cancel()
        try:
            self._id = self.widget.after(self.delay, self._show)
        except Exception:
            pass

    def _cancel(self):
        if self._id:
            try:
                self.widget.after_cancel(self._id)
            except Exception:
                pass
            self._id = None

    def _show(self):
        if self.tip or not self.text:
            return
        try:
            x = self.widget.winfo_rootx() + 18
            y = self.widget.winfo_rooty() + self.widget.winfo_height() + 6
            self.tip = tk.Toplevel(self.widget)
            self.tip.wm_overrideredirect(True)
            self.tip.wm_geometry("+%d+%d" % (x, y))
            tk.Label(self.tip, text=self.text, justify="left",
                     background="#2b3a47", foreground="white",
                     font=(FONT, 9), padx=9, pady=6, wraplength=380).pack()
        except Exception:
            self.tip = None

    def _leave(self, _=None):
        self._cancel()
        if self.tip:
            try:
                self.tip.destroy()
            except Exception:
                pass
            self.tip = None


def tip(widget, text):
    Tooltip(widget, text)
    return widget


def card(parent, title=None, subtitle=None, **kw):
    """Bloc blanc avec titre : unite visuelle de base de l'interface."""
    outer = ttk.Frame(parent, style="Card.TFrame", padding=kw.pop("padding", 12))
    if title:
        head = ttk.Frame(outer, style="Card.TFrame")
        head.pack(fill="x", pady=(0, 8))
        ttk.Label(head, text=title, style="Head.TLabel").pack(side="left")
        if subtitle:
            ttk.Label(head, text="  " + subtitle,
                      style="Muted.TLabel").pack(side="left")
        sep = tk.Frame(outer, height=1, bg=UI["line"])
        sep.pack(fill="x", pady=(0, 10))
    body = ttk.Frame(outer, style="Card.TFrame")
    body.pack(fill="both", expand=True)
    outer.body = body
    return outer


def field(parent, label, widget_cls=None, width=14, row=0, col=0, hint=None,
          **kw):
    """Etiquette + champ alignes dans une grille."""
    ttk.Label(parent, text=label, style="Card.TLabel").grid(
        row=row, column=col, sticky="w", padx=(0, 6), pady=3)
    w = (widget_cls or ttk.Entry)(parent, width=width, **kw)
    w.grid(row=row, column=col + 1, sticky="w", pady=3)
    if hint:
        tip(w, hint)
    return w


class StepBar(ttk.Frame):
    """Barre d'etapes cliquable indiquant la progression du flux de travail."""

    def __init__(self, master, steps, command=None):
        super().__init__(master, style="Bar.TFrame", padding=(10, 6))
        self.command = command
        self.labels = []
        for i, txt in enumerate(steps):
            if i:
                ttk.Label(self, text="\u203a", style="Muted.TLabel").pack(
                    side="left", padx=2)
            lb = ttk.Label(self, text=" %d. %s " % (i + 1, txt),
                           style="Step.TLabel", cursor="hand2")
            lb.pack(side="left")
            lb.bind("<Button-1>", lambda e, k=i: self._click(k))
            self.labels.append(lb)

    def _click(self, i):
        if self.command:
            self.command(i)

    def set_active(self, idx):
        for i, lb in enumerate(self.labels):
            lb.configure(style="StepOn.TLabel" if i == idx else
                         ("StepDone.TLabel" if i < idx else "Step.TLabel"))


# ----------------------------------------------------------------------------
# Preferences persistantes et fichiers recents
# ----------------------------------------------------------------------------

def resource_path(name):
    """Chemin d'une ressource, y compris depuis un executable PyInstaller."""
    base = getattr(sys, "_MEIPASS", None) or os.path.dirname(
        os.path.abspath(__file__))
    return os.path.join(base, name)


def config_dir():
    """Dossier de configuration selon le systeme."""
    if sys.platform.startswith("win"):
        base = os.environ.get("APPDATA") or os.path.expanduser("~")
    elif sys.platform == "darwin":
        base = os.path.expanduser("~/Library/Application Support")
    else:
        base = os.environ.get("XDG_CONFIG_HOME") or os.path.expanduser("~/.config")
    d = os.path.join(base, APP_NAME)
    try:
        os.makedirs(d, exist_ok=True)
    except Exception:
        d = os.path.expanduser("~")
    return d


SETTINGS_FILE = "settings.json"


def load_settings():
    try:
        with open(os.path.join(config_dir(), SETTINGS_FILE), "r",
                  encoding="utf-8") as fh:
            return json.load(fh)
    except Exception:
        return {}


def save_settings(d):
    try:
        with open(os.path.join(config_dir(), SETTINGS_FILE), "w",
                  encoding="utf-8") as fh:
            json.dump(d, fh, indent=1, ensure_ascii=False)
    except Exception:
        pass


# ----------------------------------------------------------------------------
# Fichier de session complet (.adsp)
# ----------------------------------------------------------------------------

SESSION_EXT = ".adsp"


def collect_session(app):
    """Rassemble la totalite de l'etat de travail dans un dictionnaire."""
    iast = None
    df = getattr(app.tab_iast, "df", None)
    if df is not None:
        iast = dict(columns=list(df.columns),
                    rows=df.astype(object).where(pd.notnull(df), None
                                                 ).values.tolist(),
                    names=list(getattr(app.tab_iast, "names", [])),
                    specs=copy.deepcopy(getattr(app.tab_iast, "specs", [])),
                    mode=getattr(app.tab_iast, "mode", "P"))
    tp = app.tab_plot
    plot = {}
    try:
        plot = dict(title=tp.v_title.get(), xlab=tp.v_xlab.get(),
                    ylab=tp.v_ylab.get(), y2lab=tp.v_y2lab.get(),
                    xmin=tp.v_xmin.get(), xmax=tp.v_xmax.get(),
                    ymin=tp.v_ymin.get(), ymax=tp.v_ymax.get(),
                    logx=bool(tp.v_logx.get()), logy=bool(tp.v_logy.get()),
                    grid=bool(tp.v_grid.get()), leg=bool(tp.v_leg.get()),
                    loc=tp.cb_loc.get(), fs=tp.v_fs.get(), w=tp.v_w.get(),
                    h=tp.v_h.get(), dpi=tp.v_dpi.get(),
                    annotations=tp.tbl_ann.get_rows())
    except Exception:
        pass
    iast_rows = []
    try:
        for r in app.tab_iast.rows:
            iast_rows.append(dict(name=r.e_name.get(), y=r.e_y.get(),
                                  spec=copy.deepcopy(r.spec)))
    except Exception:
        pass
    return dict(
        format="%s session" % APP_NAME, version=APP_VERSION,
        saved=datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        datasets=copy.deepcopy(app.datasets),
        fits=copy.deepcopy(app.fits),
        series=[{k: (list(map(float, v)) if k in ("x", "y") else v)
                 for k, v in s.items()} for s in app.series],
        iast=iast, iast_rows=iast_rows, plot=plot,
        notes=getattr(app, "notes", ""),
        active_tab=int(getattr(app, "_last_tab", 0)))


def write_session(app, path):
    """Ecrit la session ; conserve une copie .bak de la version precedente."""
    data = collect_session(app)
    if os.path.exists(path):
        try:
            bak = path + ".bak"
            if os.path.exists(bak):
                os.remove(bak)
            os.replace(path, bak)
        except Exception:
            pass
    tmp = path + ".tmp"
    with open(tmp, "w", encoding="utf-8") as fh:
        json.dump(data, fh, indent=1, ensure_ascii=False, default=float)
    os.replace(tmp, path)
    return data


def read_session(path):
    with open(path, "r", encoding="utf-8") as fh:
        return json.load(fh)


def restore_session(app, data):
    """Restaure integralement l'etat de travail depuis un dictionnaire."""
    app.datasets = data.get("datasets", []) or []
    app.fits = data.get("fits", []) or []
    app.series = [{k: (np.asarray(v, float) if k in ("x", "y") else v)
                   for k, v in s.items()} for s in data.get("series", []) or []]
    app.notes = data.get("notes", "")
    # IAST
    it = app.tab_iast
    d = data.get("iast")
    if d and d.get("columns"):
        try:
            it.df = pd.DataFrame(d["rows"], columns=d["columns"])
            it.names = d.get("names", [])
            it.specs = d.get("specs", [])
            it.mode = d.get("mode", "P")
            it.fill_table(it.df)
            if it.names:
                it.draw(it.df, it.names)
        except Exception:
            it.df = None
    for i, r in enumerate(data.get("iast_rows", []) or []):
        if i >= len(it.rows):
            break
        try:
            it.rows[i].e_name.delete(0, "end")
            it.rows[i].e_name.insert(0, r.get("name", ""))
            it.rows[i].e_y.delete(0, "end")
            it.rows[i].e_y.insert(0, r.get("y", ""))
            it.rows[i].spec = r.get("spec")
            it.rows[i]._update_label()
        except Exception:
            pass
    # graphique
    p = data.get("plot") or {}
    tp = app.tab_plot
    try:
        for key, var in (("title", tp.v_title), ("xlab", tp.v_xlab),
                         ("ylab", tp.v_ylab), ("y2lab", tp.v_y2lab),
                         ("xmin", tp.v_xmin), ("xmax", tp.v_xmax),
                         ("ymin", tp.v_ymin), ("ymax", tp.v_ymax),
                         ("fs", tp.v_fs), ("w", tp.v_w), ("h", tp.v_h),
                         ("dpi", tp.v_dpi)):
            if key in p:
                var.set(p[key])
        for key, var in (("logx", tp.v_logx), ("logy", tp.v_logy),
                         ("grid", tp.v_grid), ("leg", tp.v_leg)):
            if key in p:
                var.set(bool(p[key]))
        if p.get("loc"):
            tp.cb_loc.set(p["loc"])
        if p.get("annotations"):
            tp.tbl_ann.set_rows(p["annotations"])
    except Exception:
        pass
    return dict(datasets=len(app.datasets), fits=len(app.fits),
                series=len(app.series),
                iast=(len(app.tab_iast.df) if getattr(app.tab_iast, "df", None)
                      is not None else 0))


# =============================================================================
# 7. ONGLET 1 : DONNEES
# =============================================================================

class ImportDialog(tk.Toplevel):
    """Selection de feuille, colonnes et unites lors d'un import."""

    def __init__(self, master, path):
        super().__init__(master)
        self.title("Import de donnees - %s" % os.path.basename(path))
        self.transient(master)
        self.grab_set()
        self.result = None
        self.path = path
        self.sheets = {}

        ext = os.path.splitext(path)[1].lower()
        try:
            if ext in (".csv", ".txt", ".dat"):
                try:
                    df = pd.read_csv(path, sep=None, engine="python",
                                     decimal=".")
                    if df.shape[1] < 2:
                        raise ValueError
                except Exception:
                    df = pd.read_csv(path, sep=None, engine="python",
                                     decimal=",")
                self.sheets = {"CSV": df}
            else:
                xl = pd.ExcelFile(path)
                self.sheets = {s: xl.parse(s) for s in xl.sheet_names}
        except Exception as e:
            messagebox.showerror("Import", "Lecture impossible :\n%s" % e,
                                 parent=master)
            self.destroy()
            return

        frm = ttk.Frame(self, padding=10)
        frm.pack(fill="both", expand=True)
        r = 0
        ttk.Label(frm, text="Feuille :").grid(row=r, column=0, sticky="w")
        self.cb_sheet = ttk.Combobox(frm, values=list(self.sheets.keys()),
                                     state="readonly", width=28)
        self.cb_sheet.grid(row=r, column=1, sticky="w", padx=4, pady=2)
        self.cb_sheet.current(0)
        self.cb_sheet.bind("<<ComboboxSelected>>", lambda e: self._sheet_changed())
        r += 1
        ttk.Label(frm, text="Colonne pression :").grid(row=r, column=0, sticky="w")
        self.cb_p = ttk.Combobox(frm, state="readonly", width=28)
        self.cb_p.grid(row=r, column=1, sticky="w", padx=4, pady=2)
        ttk.Label(frm, text="Unite :").grid(row=r, column=2, sticky="e")
        self.cb_up = ttk.Combobox(frm, values=list(P_UNITS.keys()),
                                  state="readonly", width=10)
        self.cb_up.grid(row=r, column=3, sticky="w", padx=4)
        self.cb_up.set("bar")
        r += 1
        ttk.Label(frm, text="Colonne quantite :").grid(row=r, column=0, sticky="w")
        self.cb_q = ttk.Combobox(frm, state="readonly", width=28)
        self.cb_q.grid(row=r, column=1, sticky="w", padx=4, pady=2)
        ttk.Label(frm, text="Unite :").grid(row=r, column=2, sticky="e")
        self.cb_uq = ttk.Combobox(frm, values=Q_UNITS, state="readonly", width=12)
        self.cb_uq.grid(row=r, column=3, sticky="w", padx=4)
        self.cb_uq.set("mmol/g")
        r += 1
        ttk.Label(frm, text="Gaz :").grid(row=r, column=0, sticky="w")
        self.cb_gas = ttk.Combobox(frm, values=GAS_LIST, width=12)
        self.cb_gas.grid(row=r, column=1, sticky="w", padx=4, pady=2)
        self.cb_gas.set("CO2")
        ttk.Label(frm, text="T (K) :").grid(row=r, column=2, sticky="e")
        self.e_T = ttk.Entry(frm, width=12)
        self.e_T.insert(0, "298.15")
        self.e_T.grid(row=r, column=3, sticky="w", padx=4)
        r += 1
        ttk.Label(frm, text="Nom du jeu :").grid(row=r, column=0, sticky="w")
        self.e_name = ttk.Entry(frm, width=30)
        self.e_name.insert(0, os.path.splitext(os.path.basename(path))[0])
        self.e_name.grid(row=r, column=1, sticky="w", padx=4, pady=2)
        ttk.Label(frm, text="Type :").grid(row=r, column=2, sticky="e")
        self.cb_kind = ttk.Combobox(frm, values=["Exces", "Total"],
                                    state="readonly", width=10)
        self.cb_kind.set("Exces")
        self.cb_kind.grid(row=r, column=3, sticky="w", padx=4)
        r += 1
        self.txt = tk.Text(frm, height=10, width=90, font=("Consolas", 8))
        self.txt.grid(row=r, column=0, columnspan=4, sticky="nsew", pady=6)
        frm.rowconfigure(r, weight=1)
        r += 1
        bar = ttk.Frame(frm)
        bar.grid(row=r, column=0, columnspan=4, sticky="e")
        ttk.Button(bar, text="Annuler", command=self.destroy).pack(side="right",
                                                                   padx=4)
        ttk.Button(bar, text="Importer", command=self._ok).pack(side="right")
        self._sheet_changed()
        self.wait_window(self)

    def _sheet_changed(self):
        df = self.sheets[self.cb_sheet.get()]
        cols = [str(c) for c in df.columns]
        self.cb_p["values"] = cols
        self.cb_q["values"] = cols
        gp = [c for c in cols if any(k in c.lower()
                                     for k in ("press", "p ", "p(", "bar"))]
        gq = [c for c in cols if any(k in c.lower()
                                     for k in ("adsorb", "quant", "n_", "uptake",
                                               "mmol", "excess", "exces"))]
        self.cb_p.set(gp[0] if gp else (cols[0] if cols else ""))
        self.cb_q.set(gq[0] if gq else (cols[1] if len(cols) > 1 else ""))
        self.txt.delete("1.0", "end")
        self.txt.insert("1.0", df.head(12).to_string())

    def _ok(self):
        df = self.sheets[self.cb_sheet.get()]
        try:
            P = pd.to_numeric(df[self.cb_p.get()], errors="coerce").values
            q = pd.to_numeric(df[self.cb_q.get()], errors="coerce").values
        except Exception as e:
            messagebox.showerror("Import", str(e), parent=self)
            return
        m = np.isfinite(P) & np.isfinite(q)
        self.result = dict(name=self.e_name.get().strip() or "Jeu importe",
                           gas=self.cb_gas.get().strip() or "CO2",
                           T=to_float_or(self.e_T.get(), 298.15),
                           kind=self.cb_kind.get(),
                           unit_p=self.cb_up.get(), unit_q=self.cb_uq.get(),
                           P_raw=P[m].tolist(), q_raw=q[m].tolist(),
                           note="Importe de %s" % os.path.basename(self.path))
        self.destroy()


def make_dataset(name, gas, T, kind, unit_p, unit_q, P_raw, q_raw, note=""):
    M = GAS_DB.get(gas, {}).get("M", 44.01)
    P_bar = np.asarray(P_raw, float) * P_UNITS.get(unit_p, 1.0)
    q_mm = q_to_mmol_g(np.asarray(q_raw, float), unit_q, M)
    return dict(name=name, gas=gas, T=float(T), kind=kind, note=note,
                unit_p=unit_p, unit_q=unit_q,
                P_raw=list(map(float, np.asarray(P_raw, float))),
                q_raw=list(map(float, np.asarray(q_raw, float))),
                P=list(map(float, P_bar)), q=list(map(float, q_mm)))



# -----------------------------------------------------------------------------
# Import en lot : deduction du gaz / de l'echantillon a partir du nom de fichier
# -----------------------------------------------------------------------------

_GAS_PATTERNS = [
    ("CO2", (r"co2", r"co_2", r"carbon.?diox")),
    ("CH4", (r"ch4", r"ch_4", r"methan")),
    ("H2",  (r"(?<![a-z0-9])h2(?![a-z0-9])", r"hydrog")),
    ("N2",  (r"(?<![a-z0-9])n2(?![a-z0-9])", r"azote", r"nitrog")),
    ("CO",  (r"(?<![a-z0-9])co(?![a-z0-9])",)),
    ("Ar",  (r"(?<![a-z0-9])ar(?![a-z0-9])", r"argon")),
    ("O2",  (r"(?<![a-z0-9])o2(?![a-z0-9])", r"oxyg")),
]


def guess_gas_from_name(name):
    low = os.path.splitext(os.path.basename(name))[0].lower()
    import re as _re
    for gas, pats in _GAS_PATTERNS:
        for pat in pats:
            if _re.search(pat, low):
                return gas
    return None


def guess_sample_from_name(name):
    """Renvoie la partie du nom de fichier qui n'est pas le gaz (ex. ZTC850)."""
    import re as _re
    base = os.path.splitext(os.path.basename(name))[0]
    g = guess_gas_from_name(name)
    if g:
        base = _re.sub(r"[-_ ]*" + _re.escape(g) + r"[-_ ]*", " ", base,
                       flags=_re.I)
    base = _re.sub(r"[-_]+", " ", base).strip()
    return base or os.path.splitext(os.path.basename(name))[0]


def read_two_columns(path, sheet=None):
    """Lit le premier bloc de deux colonnes numeriques d'un fichier.
    Renvoie (P, q, nom_feuille, entetes) ou leve une exception."""
    ext = os.path.splitext(path)[1].lower()
    if ext in (".xlsx", ".xls", ".xlsm"):
        xl = pd.ExcelFile(path)
        sheets = [sheet] if sheet else xl.sheet_names
        last = None
        for sh in sheets:
            df = xl.parse(sh, header=0)
            num = df.apply(pd.to_numeric, errors="coerce")
            good = [c for c in num.columns if num[c].notna().sum() >= 3]
            if len(good) >= 2:
                sub = num[good[:2]].dropna()
                if len(sub) >= 3:
                    return (sub.iloc[:, 0].to_numpy(float),
                            sub.iloc[:, 1].to_numpy(float), sh,
                            [str(good[0]), str(good[1])])
            last = sh
        raise ValueError("aucune paire de colonnes numeriques (feuille %s)"
                         % last)
    df = pd.read_csv(path, sep=None, engine="python")
    num = df.apply(pd.to_numeric, errors="coerce")
    good = [c for c in num.columns if num[c].notna().sum() >= 3]
    if len(good) < 2:
        raise ValueError("aucune paire de colonnes numeriques")
    sub = num[good[:2]].dropna()
    return (sub.iloc[:, 0].to_numpy(float), sub.iloc[:, 1].to_numpy(float),
            "-", [str(good[0]), str(good[1])])


class BatchImportDialog(tk.Toplevel):
    """Recapitulatif modifiable avant import en lot."""

    def __init__(self, master, rows, T_default=303.0):
        super().__init__(master)
        self.title("Import en lot - verifier avant de valider")
        self.transient(master)
        self.grab_set()
        self.result = None
        self.rows = rows
        frm = ttk.Frame(self, padding=12)
        frm.pack(fill="both", expand=True)

        ttk.Label(frm, text="%d fichier(s) lu(s). Verifiez le gaz, la "
                            "temperature et les unites, puis validez."
                  % len(rows)).pack(anchor="w", pady=(0, 8))

        top = ttk.Frame(frm)
        top.pack(fill="x", pady=(0, 8))
        ttk.Label(top, text="Appliquer a tous  -  T (K) :").pack(side="left")
        self.e_T = ttk.Entry(top, width=9)
        self.e_T.insert(0, fmt(T_default, 2))
        self.e_T.pack(side="left", padx=4)
        ttk.Label(top, text="Unite P :").pack(side="left", padx=(10, 2))
        self.cb_up = ttk.Combobox(top, state="readonly", width=10,
                                  values=list(P_UNITS.keys()))
        self.cb_up.set("bar")
        self.cb_up.pack(side="left")
        ttk.Label(top, text="Unite q :").pack(side="left", padx=(10, 2))
        self.cb_uq = ttk.Combobox(top, state="readonly", width=12,
                                  values=Q_UNITS)
        self.cb_uq.set("mmol/g")
        self.cb_uq.pack(side="left")
        ttk.Label(top, text="Type :").pack(side="left", padx=(10, 2))
        self.cb_kind = ttk.Combobox(top, state="readonly", width=8,
                                    values=["Exces", "Total"])
        self.cb_kind.set("Exces")
        self.cb_kind.pack(side="left")
        ttk.Button(top, text="Appliquer", command=self._apply_all
                   ).pack(side="left", padx=8)

        cols = ("Fichier", "Feuille", "Nom du jeu", "Gaz", "T (K)",
                "Points", "P min", "P max", "q max")
        self.tv = ttk.Treeview(frm, columns=cols, show="headings", height=12)
        for c, w in zip(cols, (200, 80, 170, 60, 70, 60, 70, 70, 80)):
            self.tv.heading(c, text=c)
            self.tv.column(c, width=w, anchor="w")
        self.tv.pack(fill="both", expand=True)
        self.tv.bind("<Double-1>", self._edit_cell)
        ttk.Label(frm, style="Muted.TLabel",
                  text="Double-clic sur 'Nom du jeu', 'Gaz' ou 'T (K)' pour "
                       "corriger une ligne.").pack(anchor="w", pady=(4, 0))
        self._fill()

        bar = ttk.Frame(frm)
        bar.pack(fill="x", pady=(10, 0))
        ttk.Button(bar, text="Annuler", command=self.destroy
                   ).pack(side="right", padx=4)
        ttk.Button(bar, text="Importer", style="Primary.TButton",
                   command=self._ok).pack(side="right")
        ttk.Button(bar, text="Retirer la ligne", style="Ghost.TButton",
                   command=self._drop).pack(side="left")
        self.geometry("1020x520")
        self.wait_window(self)

    def _fill(self):
        self.tv.delete(*self.tv.get_children())
        for i, r in enumerate(self.rows):
            self.tv.insert("", "end", iid=str(i), values=(
                os.path.basename(r["path"]), r["sheet"], r["name"], r["gas"],
                fmt(r["T"], 2), len(r["P"]), fmt(np.min(r["P"]), 4),
                fmt(np.max(r["P"]), 4), fmt(np.max(r["q"]), 4)))

    def _apply_all(self):
        T = to_float_or(self.e_T.get(), 303.0)
        for r in self.rows:
            r["T"] = T
            r["unit_p"] = self.cb_up.get()
            r["unit_q"] = self.cb_uq.get()
            r["kind"] = self.cb_kind.get()
        self._fill()

    def _drop(self):
        sel = self.tv.selection()
        if not sel:
            return
        for i in sorted((int(x) for x in sel), reverse=True):
            self.rows.pop(i)
        self._fill()

    def _edit_cell(self, event):
        item = self.tv.identify_row(event.y)
        col = self.tv.identify_column(event.x)
        if not item or col not in ("#3", "#4", "#5"):
            return
        i = int(item)
        key = {"#3": "name", "#4": "gas", "#5": "T"}[col]
        x, y, w, h = self.tv.bbox(item, col)
        if key == "gas":
            ed = ttk.Combobox(self.tv, values=GAS_LIST, state="readonly")
            ed.set(self.rows[i]["gas"])
        else:
            ed = ttk.Entry(self.tv)
            ed.insert(0, str(self.rows[i][key]) if key == "name"
                      else fmt(self.rows[i]["T"], 2))
        ed.place(x=x, y=y, width=w, height=h)
        ed.focus_set()

        def done(_=None):
            v = ed.get()
            if key == "T":
                self.rows[i]["T"] = to_float_or(v, self.rows[i]["T"])
            else:
                self.rows[i][key] = v
            ed.destroy()
            self._fill()
        ed.bind("<Return>", done)
        ed.bind("<FocusOut>", done)
        ed.bind("<<ComboboxSelected>>", done)

    def _ok(self):
        if not self.rows:
            self.destroy()
            return
        self.result = self.rows
        self.destroy()


class TabData(ttk.Frame):
    def __init__(self, master, app):
        super().__init__(master, padding=10)
        self.app = app
        self._dirty = False
        self._loading = False
        self._loaded_index = None
        pan = ttk.PanedWindow(self, orient="horizontal")
        pan.pack(fill="both", expand=True)
        left = ttk.Frame(pan)
        right = ttk.Frame(pan)
        pan.add(left, weight=0)
        pan.add(right, weight=3)

        # ------------------------------------------------ liste des jeux ---
        c = card(left, "Jeux de donnees", "selectionnez pour ouvrir")
        c.pack(fill="both", expand=True)
        b = c.body
        ff = ttk.Frame(b, style="Card.TFrame")
        ff.pack(fill="x", pady=(0, 6))
        ttk.Label(ff, text="Filtrer :", style="Card.TLabel").pack(side="left")
        self.var_filter = tk.StringVar()
        ent = ttk.Entry(ff, textvariable=self.var_filter, width=18)
        ent.pack(side="left", fill="x", expand=True, padx=4)
        self.var_filter.trace_add("write", lambda *a: self.refresh_list())
        tip(ttk.Button(ff, text="x", width=3, style="Ghost.TButton",
                       command=lambda: self.var_filter.set("")),
            "Effacer le filtre").pack(side="left")
        tvf = ttk.Frame(b, style="Card.TFrame")
        tvf.pack(fill="both", expand=True)
        self.lst = ttk.Treeview(tvf, show="tree", height=15, selectmode="browse")
        self.lst.column("#0", width=270)
        sc = ttk.Scrollbar(tvf, orient="vertical", command=self.lst.yview)
        self.lst.configure(yscrollcommand=sc.set)
        self.lst.pack(side="left", fill="both", expand=True)
        sc.pack(side="right", fill="y")
        self.lst.bind("<<TreeviewSelect>>", lambda e: self.on_select())

        bb = ttk.Frame(b, style="Card.TFrame")
        bb.pack(fill="x", pady=(12, 0))
        tip(ttk.Button(bb, text="Importer un fichier...", style="Primary.TButton",
                       command=self.import_file),
            "Lire un .xlsx, .xls ou .csv et choisir la feuille, les colonnes "
            "et les unites").grid(row=0, column=0, columnspan=2, sticky="ew",
                                  pady=2)
        tip(ttk.Button(bb, text="Importer un dossier / plusieurs fichiers...",
                       style="Primary.TButton", command=self.import_many),
            "Selectionner plusieurs fichiers d'un coup. Le gaz, l'echantillon "
            "et les unites sont deduits du nom de fichier "
            "(ex. ZTC850-CO2.xlsx) ; un recapitulatif est affiche avant import."
            ).grid(row=1, column=0, columnspan=2, sticky="ew", pady=2)
        tip(ttk.Button(bb, text="Nouveau jeu vide", command=self.new_dataset),
            "Creer un jeu vide et saisir ou coller les valeurs a la main"
            ).grid(row=2, column=0, columnspan=2, sticky="ew", pady=2)
        tip(ttk.Button(bb, text="Dupliquer", command=self.duplicate),
            "Copier le jeu selectionne").grid(row=3, column=0, sticky="ew",
                                              padx=(0, 3), pady=2)
        tip(ttk.Button(bb, text="Supprimer", style="Danger.TButton",
                       command=self.delete),
            "Supprimer definitivement le jeu selectionne").grid(
            row=3, column=1, sticky="ew", padx=(3, 0), pady=2)
        bb.columnconfigure(0, weight=1)
        bb.columnconfigure(1, weight=1)

        # -------------------------------------------------- metadonnees ----
        head = card(right, "Metadonnees",
                    "gaz et temperature servent aux modeles D-R / D-A et a l'IAST")
        head.pack(fill="x")
        g = head.body
        self.e_name = field(g, "Nom :", ttk.Entry, 28, 0, 0,
                            "Nom affiche dans tous les onglets")
        self.cb_gas = field(g, "Gaz :", ttk.Combobox, 12, 0, 2,
                            "Determine la masse molaire, Tc, Pc et l'acentricite",
                            values=GAS_LIST)
        self.cb_gas.set("CO2")
        self.e_T = field(g, "T (K) :", ttk.Entry, 12, 0, 4,
                         "Temperature de la mesure, en kelvin")
        self.e_T.insert(0, "298.15")
        self.cb_kind = field(g, "Type :", ttk.Combobox, 10, 0, 6,
                             "Quantite en exces (mesure brute) ou totale "
                             "(absolue)", values=["Exces", "Total"],
                             state="readonly")
        self.cb_kind.set("Exces")
        self.cb_up = field(g, "Unite P :", ttk.Combobox, 12, 1, 0,
                           "Unite de la colonne pression du tableau",
                           values=list(P_UNITS.keys()), state="readonly")
        self.cb_up.set("bar")
        self.cb_up.bind("<<ComboboxSelected>>", lambda e: self._update_headers())
        self.cb_uq = field(g, "Unite q :", ttk.Combobox, 12, 1, 2,
                           "Unite de la colonne quantite adsorbee",
                           values=Q_UNITS, state="readonly")
        self.cb_uq.set("mmol/g")
        self.cb_uq.bind("<<ComboboxSelected>>", lambda e: self._update_headers())
        self.e_note = field(g, "Commentaire :", ttk.Entry, 46, 1, 4,
                            "Note libre : origine de l'echantillon, appareil...")
        self.e_note.grid(columnspan=3)

        # -------------------------------------------- tableau et apercu ----
        mid = ttk.PanedWindow(right, orient="horizontal")
        mid.pack(fill="both", expand=True, pady=(10, 0))
        cl = ttk.Frame(mid)
        cr = ttk.Frame(mid)
        mid.add(cl, weight=1)
        mid.add(cr, weight=1)

        ct = card(cl, "Donnees experimentales",
                  "double-clic pour editer  -  Ctrl+V pour coller depuis Excel")
        ct.pack(fill="both", expand=True)
        self.tbl = EditableTable(ct.body, ["P (bar)", "q (mmol/g)"], height=16,
                                 widths=[150, 150], nrows=12)
        self.tbl.pack(fill="both", expand=True)
        tb = ttk.Frame(ct.body, style="Card.TFrame")
        tb.pack(fill="x", pady=(8, 0))
        for txt, cmd, hint in (
                ("+ Ligne", lambda: self.tbl.add_row(), "Ajouter une ligne vide"),
                ("- Ligne", self.tbl.delete_selected,
                 "Supprimer les lignes selectionnees (touche Suppr)"),
                ("Trier", lambda: self.tbl.sort_by(0),
                 "Trier par pression croissante"),
                ("Coller", lambda: self.tbl._paste(),
                 "Coller deux colonnes depuis le presse-papiers"),
                ("Vider", self.tbl.clear, "Effacer tout le tableau")):
            tip(ttk.Button(tb, text=txt, width=9, command=cmd),
                hint).pack(side="left", padx=2)
        sv = ttk.Frame(ct.body, style="Card.TFrame")
        sv.pack(fill="x", pady=(10, 0))
        self.lbl_dirty = ttk.Label(sv, style="Card.TLabel", text="")
        self.lbl_dirty.pack(side="left")
        self.btn_save = ttk.Button(sv, text="Enregistrer le jeu de donnees",
                                   style="Success.TButton",
                                   command=self.save_current)
        self.btn_save.pack(side="right")
        tip(self.btn_save,
            "Valide le tableau et les metadonnees. L'enregistrement est aussi "
            "automatique quand vous changez de jeu, lancez un ajustement ou "
            "quittez l'onglet - vos saisies ne sont plus perdues.")
        for _w in (self.e_name, self.e_T, self.e_note):
            _w.bind("<KeyRelease>", self._touch)
        for _w in (self.cb_gas, self.cb_kind, self.cb_up, self.cb_uq):
            _w.bind("<<ComboboxSelected>>", self._touch, add="+")
        self.tbl.on_change = self._touch

        cp = card(cr, "Apercu")
        cp.pack(fill="both", expand=True)
        self.fig = Figure(figsize=(4.6, 3.8), dpi=100, facecolor=UI["card"])
        self.ax = self.fig.add_subplot(111)
        self.canvas = FigureCanvasTkAgg(self.fig, master=cp.body)
        self.canvas.get_tk_widget().pack(fill="both", expand=True)
        if NavigationToolbar2Tk is not None:
            tbar = NavigationToolbar2Tk(self.canvas, cp.body)
            tbar.update()
        set_sash(pan, 0, 330)
        set_sash(mid, 0, 470)
        self.refresh_list()

    def _update_headers(self):
        cols = ["P (%s)" % self.cb_up.get(), "q (%s)" % self.cb_uq.get()]
        for old, new in zip(self.tbl.columns, cols):
            self.tbl.tree.heading(old, text=new)

    def refresh_list(self, select=None):
        if select is None:
            select = self.current_index()          # conserve la selection
        try:
            flt = self.var_filter.get().strip().lower()
        except Exception:
            flt = ""
        self._loading = True
        self.lst.delete(*self.lst.get_children())
        shown = []
        for i, d in enumerate(self.app.datasets):
            txt = "%s  [%s, %.1f K, %s]" % (d["name"], d["gas"], d["T"],
                                            d["kind"])
            if flt and flt not in txt.lower():
                continue
            self.lst.insert("", "end", iid=str(i), text=txt)
            shown.append(i)
        if select is not None and select in shown:
            self.lst.selection_set(str(select))
            self.lst.see(str(select))
        elif shown:
            self.lst.selection_set(str(shown[0]))
        self._loading = False
        self._update_dirty(False)

    def current_index(self):
        sel = self.lst.selection()
        return int(sel[0]) if sel else None

    # ---------------------------------------------------------- enregistrement
    def _update_dirty(self, state=True):
        """Indicateur visuel : le jeu affiche differe-t-il du jeu memorise ?"""
        self._dirty = bool(state) and not getattr(self, "_loading", False)
        if not hasattr(self, "lbl_dirty"):
            return
        if self._dirty:
            self.lbl_dirty.config(
                text="\u25cf  modifications non enregistrees",
                foreground="#B26B00")
            self.btn_save.config(text="Enregistrer le jeu de donnees  (Ctrl+S)")
        else:
            self.lbl_dirty.config(text="\u2713  a jour", foreground="#2E7D32")
            self.btn_save.config(text="Enregistrer le jeu de donnees")

    def _touch(self, *_a):
        self._update_dirty(True)

    def _autosave(self):
        """Enregistre silencieusement le jeu courant s'il a ete modifie."""
        if getattr(self, "_dirty", False) and self._loaded_index is not None:
            try:
                self.save_current(silent=True, index=self._loaded_index)
                return True
            except Exception:
                return False
        return False

    def on_select(self):
        """Change de jeu SANS perdre les modifications en cours."""
        if getattr(self, "_loading", False):
            return
        new = self.current_index()
        if new == self._loaded_index:
            return
        self._autosave()
        self.load_selected()

    def load_selected(self):
        i = self.current_index()
        if i is None:
            return
        self._loading = True
        self._loaded_index = i
        d = self.app.datasets[i]
        self.e_name.delete(0, "end")
        self.e_name.insert(0, d["name"])
        self.cb_gas.set(d["gas"])
        self.e_T.delete(0, "end")
        self.e_T.insert(0, fmt(d["T"], 2))
        self.cb_kind.set(d["kind"])
        self.cb_up.set(d.get("unit_p", "bar"))
        self.cb_uq.set(d.get("unit_q", "mmol/g"))
        self.e_note.delete(0, "end")
        self.e_note.insert(0, d.get("note", ""))
        self._update_headers()
        self.tbl.set_rows(list(zip(d["P_raw"], d["q_raw"])))
        self.plot_preview(d)
        self._loading = False
        self._update_dirty(False)

    def plot_preview(self, d=None):
        self.ax.clear()
        if d:
            self.ax.plot(d["P"], d["q"], "o-", color=PALETTE[0], ms=5)
            self.ax.set_xlabel("P (bar)")
            self.ax.set_ylabel("q (mmol/g)")
            self.ax.set_title(d["name"], fontsize=9)
            self.ax.grid(True, ls="--", alpha=0.5)
        self.fig.tight_layout()
        self.canvas.draw_idle()

    def _collect(self):
        arr = self.tbl.get_numeric(2)
        arr = arr[np.isfinite(arr).all(axis=1)]
        if arr.shape[0] == 0:
            raise ValueError("Aucune donnee numerique valide dans le tableau.")
        return make_dataset(self.e_name.get().strip() or "Jeu %d" % (len(self.app.datasets) + 1),
                            self.cb_gas.get().strip() or "CO2",
                            to_float_or(self.e_T.get(), 298.15),
                            self.cb_kind.get(), self.cb_up.get(),
                            self.cb_uq.get(), arr[:, 0], arr[:, 1],
                            self.e_note.get())

    def save_current(self, silent=False, index=None):
        try:
            d = self._collect()
        except Exception as e:
            if not silent:
                messagebox.showerror("Donnees", str(e), parent=self)
            return
        i = index if index is not None else self.current_index()
        if i is None or not (0 <= i < len(self.app.datasets)):
            self.app.datasets.append(d)
            i = len(self.app.datasets) - 1
        else:
            self.app.datasets[i] = d
        self._loaded_index = i
        self.refresh_list(select=i)
        self.plot_preview(d)
        self.app.notify_datasets()
        self._update_dirty(False)
        self.app.log("Jeu de donnees enregistre : %s (%d points)"
                     % (d["name"], len(d["P"])))

    def new_dataset(self):
        self.lst.selection_remove(self.lst.selection())
        self.e_name.delete(0, "end")
        self.e_name.insert(0, "Jeu %d" % (len(self.app.datasets) + 1))
        self.e_note.delete(0, "end")
        self.tbl.clear()
        for _ in range(12):
            self.tbl.add_row()
        self.plot_preview(None)

    def duplicate(self):
        i = self.current_index()
        if i is None:
            return
        d = copy.deepcopy(self.app.datasets[i])
        d["name"] += " (copie)"
        self.app.datasets.append(d)
        self.refresh_list(select=len(self.app.datasets) - 1)
        self.app.notify_datasets()

    def delete(self):
        i = self.current_index()
        if i is None:
            return
        if messagebox.askyesno("Supprimer", "Supprimer '%s' ?"
                               % self.app.datasets[i]["name"], parent=self):
            self.app.datasets.pop(i)
            self.refresh_list()
            self.app.notify_datasets()

    def import_many(self):
        paths = filedialog.askopenfilenames(
            title="Importer plusieurs isothermes",
            filetypes=[("Excel / CSV", "*.xlsx *.xls *.xlsm *.csv *.txt *.dat"),
                       ("Tous les fichiers", "*.*")])
        if not paths:
            return
        rows, errs = [], []
        Tdef = to_float_or(self.e_T.get(), 303.0)
        for path in paths:
            try:
                P, q, sh, hdr = read_two_columns(path)
            except Exception as e:
                errs.append("%s : %s" % (os.path.basename(path), e))
                continue
            gas = guess_gas_from_name(path) or "CO2"
            samp = guess_sample_from_name(path)
            rows.append(dict(path=path, sheet=sh, gas=gas,
                             name="%s %s" % (samp, gas) if samp else
                                  os.path.splitext(os.path.basename(path))[0],
                             T=Tdef, kind="Exces", unit_p="bar",
                             unit_q="mmol/g", P=P, q=q,
                             note="importe de %s (feuille %s, colonnes %s)"
                                  % (os.path.basename(path), sh, " / ".join(hdr))))
        if errs:
            messagebox.showwarning(
                "Import en lot",
                "Fichier(s) ignore(s) :\n  - " + "\n  - ".join(errs[:12]),
                parent=self)
        if not rows:
            return
        dlg = BatchImportDialog(self, rows, T_default=Tdef)
        if not dlg.result:
            return
        first = len(self.app.datasets)
        for r in dlg.result:
            d = make_dataset(r["name"], r["gas"], r["T"], r["kind"],
                             r["unit_p"], r["unit_q"], r["P"], r["q"],
                             r["note"])
            self.app.datasets.append(d)
        self.refresh_list(select=first)
        self.load_selected()
        self.app.notify_datasets()
        self.app.log("Import en lot : %d jeu(x) ajoute(s)." % len(dlg.result))
        self.app.set_status("%d isotherme(s) importee(s)." % len(dlg.result))

    def import_file(self):
        path = filedialog.askopenfilename(
            title="Importer des donnees d'isotherme",
            filetypes=[("Excel / CSV", "*.xlsx *.xls *.xlsm *.csv *.txt *.dat"),
                       ("Tous les fichiers", "*.*")])
        if not path:
            return
        dlg = ImportDialog(self, path)
        if not dlg.result:
            return
        r = dlg.result
        d = make_dataset(r["name"], r["gas"], r["T"], r["kind"], r["unit_p"],
                         r["unit_q"], r["P_raw"], r["q_raw"], r["note"])
        self.app.datasets.append(d)
        self.refresh_list(select=len(self.app.datasets) - 1)
        self.load_selected()
        self.app.notify_datasets()
        self.app.log("Import : %s (%d points)" % (d["name"], len(d["P"])))


# =============================================================================
# 8. ONGLET 2 : CONVERSION EXCES -> TOTAL
# =============================================================================

# Densites typiques de phase adsorbee (g/cm3) : liquide au point d'ebullition
# normal ou densite de van der Waals, valeurs indicatives de la litterature.
RHO_ADS_TYPICAL = {
    "CO2": 1.023, "CH4": 0.4224, "H2": 0.0708, "N2": 0.8067, "O2": 1.141,
    "Ar": 1.3954, "CO": 0.789, "C2H6": 0.5446, "C2H4": 0.5678,
    "C3H8": 0.5808, "n-C4H10": 0.6011, "Kr": 2.413, "Xe": 2.942,
}


def saturation_pressure(gas, T):
    """Pression de saturation (bar) ou pseudo-saturation si T > Tc.

    T < Tc : correlation de Lee-Kesler.
    T >= Tc: regle de Dubinin  P0 = Pc (T/Tc)^2.
    """
    if gas not in GAS_DB:
        return np.nan, "gaz inconnu"
    g = GAS_DB[gas]
    Tr = T / g["Tc"]
    if Tr >= 1.0:
        return g["Pc"] * Tr ** 2, "pseudo-saturation de Dubinin  Pc(T/Tc)^2"
    if Tr < 0.25:
        return np.nan, "hors domaine de la correlation"
    f0 = (5.92714 - 6.09648 / Tr - 1.28862 * math.log(Tr) + 0.169347 * Tr ** 6)
    f1 = (15.2518 - 15.6875 / Tr - 13.4721 * math.log(Tr) + 0.43577 * Tr ** 6)
    return g["Pc"] * math.exp(f0 + g["w"] * f1), "correlation de Lee-Kesler"


class TabConversion(ttk.Frame):
    def __init__(self, master, app):
        super().__init__(master, padding=10)
        self.app = app
        self.result = None

        top = ttk.Frame(self)
        top.pack(fill="x")
        c1 = card(top, "1. Jeu de donnees et methode")
        c1.pack(side="left", fill="both", expand=True)
        g = c1.body
        ttk.Label(g, text="Jeu de donnees :", style="Card.TLabel").grid(
            row=0, column=0, sticky="w", pady=3)
        self.cb_ds = ttk.Combobox(g, state="readonly", width=44)
        self.cb_ds.grid(row=0, column=1, columnspan=3, sticky="w", padx=6)
        self.cb_ds.bind("<<ComboboxSelected>>", lambda e: self._ds_changed())
        ttk.Label(g, text="Methode :", style="Card.TLabel").grid(
            row=1, column=0, sticky="w", pady=3)
        self.cb_meth = ttk.Combobox(g, state="readonly", width=44,
                                    values=["Volume poreux", "Densite adsorbee",
                                            "Densite fournie"])
        self.cb_meth.set("Volume poreux")
        self.cb_meth.grid(row=1, column=1, columnspan=3, sticky="w", padx=6)
        self.cb_meth.bind("<<ComboboxSelected>>", lambda e: self._meth_changed())
        ttk.Label(g, style="Muted.TLabel", justify="left", text=(
            "Volume poreux      n_tot = n_exc + V_pore x rho_gaz(T,P)\n"
            "Densite adsorbee   n_tot = n_exc / (1 - rho_gaz/rho_ads)\n"
            "Densite fournie    vous saisissez rho_gaz dans la colonne 2 du "
            "tableau (valeurs NIST par exemple)")).grid(
            row=2, column=0, columnspan=4, sticky="w", pady=(8, 0))

        c2 = card(top, "2. Parametres")
        c2.pack(side="left", fill="both", padx=(10, 0))
        h = c2.body
        self.cb_eos = field(h, "Equation d'etat :", ttk.Combobox, 18, 0, 0,
                            "Table NIST : densites de reference tabulees "
                            "(CO2, CH4, H2 a 303 K) - le plus exact, repli "
                            "automatique sur Peng-Robinson hors table. "
                            "Peng-Robinson : EOS cubique translatee Peneloux, "
                            "1 a 3 % d'ecart pres du point critique.",
                            values=(["Table NIST", "Peng-Robinson",
                                     "Gaz parfait"]
                                    + (["CoolProp"] if HAS_COOLPROP else [])),
                            state="readonly")
        self.cb_eos.set("Table NIST")
        self.e_vp = field(h, "V_pore (cm3/g) :", ttk.Entry, 12, 1, 0,
                          "Volume poreux total accessible, mesure par DFT, "
                          "D-R ou pycnometrie helium")
        self.e_vp.insert(0, "0.50")
        ttk.Label(h, text="rho phase adsorbee :", style="Card.TLabel").grid(
            row=2, column=0, sticky="w", pady=3)
        rw = ttk.Frame(h, style="Card.TFrame")
        rw.grid(row=2, column=1, sticky="w")
        self.e_rho = ttk.Entry(rw, width=10)
        self.e_rho.pack(side="left")
        self.cb_rho_u = ttk.Combobox(rw, state="readonly", width=9,
                                     values=["g/cm3", "mmol/cm3"])
        self.cb_rho_u.set("g/cm3")
        self.cb_rho_u.pack(side="left", padx=4)
        tip(ttk.Button(rw, text="Typique", style="Ghost.TButton", width=8,
                       command=self._fill_typical),
            "Densite du liquide au point d'ebullition normal pour ce gaz"
            ).pack(side="left")

        bar = ttk.Frame(self, style="Card.TFrame", padding=10)
        bar.pack(fill="x", pady=10)
        tip(ttk.Button(bar, text="Calculer la conversion", style="Primary.TButton",
                       command=self.compute),
            "Calcule rho_gaz(T,P) et n_total pour chaque point").pack(side="left")
        tip(ttk.Button(bar, text="Creer un jeu 'total'", style="Success.TButton",
                       command=self.create_dataset),
            "Ajoute le resultat comme nouveau jeu de donnees, utilisable pour "
            "l'ajustement").pack(side="left", padx=8)
        tip(ttk.Button(bar, text="Convertir TOUS les jeux en exces",
                       style="Success.TButton", command=self.convert_all),
            "Applique la meme methode et les memes parametres a chaque jeu de "
            "type 'Exces' et cree en une fois tous les jeux 'total' "
            "correspondants.").pack(side="left", padx=(0, 8))
        tip(ttk.Button(bar, text="Exporter en Excel", command=self.export),
            "Enregistre le tableau de conversion").pack(side="left")
        self.lbl_conv = ttk.Label(bar, text="", style="Status.TLabel")
        self.lbl_conv.pack(side="left", padx=14)

        mid = ttk.PanedWindow(self, orient="horizontal")
        mid.pack(fill="both", expand=True)
        fl = ttk.Frame(mid)
        fr = ttk.Frame(mid)
        mid.add(fl, weight=1)
        mid.add(fr, weight=1)
        ct = card(fl, "Resultats")
        ct.pack(fill="both", expand=True)
        cols = ["P (bar)", "rho_gaz (mmol/cm3)", "Z", "n_exces (mmol/g)",
                "n_total (mmol/g)", "Ecart (%)"]
        self.tbl = EditableTable(ct.body, cols, height=15,
                                 widths=[90, 130, 80, 120, 120, 90])
        self.tbl.pack(fill="both", expand=True)
        cg = card(fr, "Comparaison exces / total")
        cg.pack(fill="both", expand=True)
        self.fig = Figure(figsize=(5.4, 4.6), dpi=100, facecolor=UI["card"])
        self.ax = self.fig.add_subplot(111)
        self.canvas = FigureCanvasTkAgg(self.fig, master=cg.body)
        self.canvas.get_tk_widget().pack(fill="both", expand=True)
        if NavigationToolbar2Tk is not None:
            NavigationToolbar2Tk(self.canvas, cg.body).update()
        zoom_button(cg.body, self._render, "Conversion", app).pack(
            anchor="e", pady=(4, 0))
        set_sash(mid, 0, 620)
        self._meth_changed()

    def refresh(self):
        vals = ["%d. %s [%s, %.1f K, %s]" % (i + 1, d["name"], d["gas"],
                                             d["T"], d["kind"])
                for i, d in enumerate(self.app.datasets)]
        cur = self.cb_ds.get()
        self.cb_ds["values"] = vals
        if cur in vals:
            self.cb_ds.set(cur)
        elif vals:
            self.cb_ds.current(0)
            self._ds_changed()
        else:
            self.cb_ds.set("")

    def _ds_changed(self):
        d = self.selected_dataset()
        if d and not self.e_rho.get():
            self._fill_typical()

    def _meth_changed(self):
        m = self.cb_meth.get()
        self.e_vp.config(state="normal" if m in ("Volume poreux",
                                                 "Densite fournie") else "disabled")
        self.e_rho.config(state="normal" if m == "Densite adsorbee" else "disabled")

    def _fill_typical(self):
        d = self.selected_dataset()
        gas = d["gas"] if d else "CO2"
        v = RHO_ADS_TYPICAL.get(gas)
        self.e_rho.config(state="normal")
        self.e_rho.delete(0, "end")
        if v:
            if self.cb_rho_u.get() == "g/cm3":
                self.e_rho.insert(0, "%.4f" % v)
            else:
                self.e_rho.insert(0, "%.4f" % (v * 1000.0 / GAS_DB[gas]["M"]))
        self._meth_changed()

    def selected_dataset(self):
        s = self.cb_ds.get()
        if not s:
            return None
        try:
            i = int(s.split(".")[0]) - 1
            return self.app.datasets[i]
        except Exception:
            return None

    # ------------------------------------------------------------------
    def convert_all(self):
        """Convertit d'un coup tous les jeux 'Exces' avec les parametres courants."""
        src = [d for d in self.app.datasets if d.get("kind") == "Exces"]
        if not src:
            messagebox.showinfo("Conversion en lot",
                                "Aucun jeu de type 'Exces' a convertir.",
                                parent=self)
            return
        meth = self.cb_meth.get()
        eos = self.cb_eos.get()
        Vp = to_float_or(self.e_vp.get(), 0.0)
        if meth == "Volume poreux" and not (Vp > 0):
            messagebox.showwarning("Conversion en lot",
                                   "Renseignez un volume poreux > 0.",
                                   parent=self)
            return
        if not messagebox.askyesno(
                "Conversion en lot",
                "Convertir %d jeu(x) en quantite totale ?\n\n"
                "  methode  : %s\n  EOS      : %s\n  V_pore   : %s cm3/g\n\n"
                "Les jeux 'total' deja presents portant le meme nom seront "
                "remplaces." % (len(src), meth, eos, fmt(Vp, 4)),
                parent=self):
            return
        made, fail = 0, []
        for d in src:
            try:
                rho_ads = None
                if meth == "Densite adsorbee":
                    v = to_float_or(self.e_rho.get(), np.nan)
                    if self.cb_rho_u.get() == "g/cm3" and np.isfinite(v):
                        v = v * 1000.0 / GAS_DB[d["gas"]]["M"]
                    rho_ads = v
                n_tot, rho = excess_to_total(
                    np.asarray(d["P"], float), np.asarray(d["q"], float),
                    d["T"], d["gas"], method=meth, V_pore=Vp,
                    rho_ads=rho_ads, eos=eos)
                if not np.all(np.isfinite(n_tot)):
                    raise ValueError("valeurs non finies (hors table ?)")
                name = d["name"] + " (total)"
                nd_ = make_dataset(name, d["gas"], d["T"], "Total", "bar",
                                   "mmol/g", np.asarray(d["P"], float), n_tot,
                                   "converti de '%s' : %s, V_pore=%s cm3/g, "
                                   "EOS=%s" % (d["name"], meth, fmt(Vp, 4), eos))
                self.app.datasets = [x for x in self.app.datasets
                                     if x["name"] != name]
                self.app.datasets.append(nd_)
                made += 1
            except Exception as e:
                fail.append("%s : %s" % (d["name"], e))
        self.app.notify_datasets()
        self.refresh()
        self.app.log("Conversion en lot : %d jeu(x) cree(s), %d echec(s)."
                     % (made, len(fail)))
        msg = "%d jeu(x) 'total' cree(s)." % made
        if fail:
            msg += "\n\nEchecs :\n  - " + "\n  - ".join(fail[:10])
        messagebox.showinfo("Conversion en lot", msg, parent=self)

    def compute(self):
        d = self.selected_dataset()
        if d is None:
            messagebox.showwarning("Conversion", "Selectionnez un jeu de donnees.",
                                   parent=self)
            return
        P = np.asarray(d["P"], float)
        n_exc = np.asarray(d["q"], float)
        meth = self.cb_meth.get()
        eos = self.cb_eos.get()
        gas, T = d["gas"], d["T"]
        try:
            if meth == "Densite fournie":
                rows = self.tbl.get_numeric(3)
                if rows.shape[0] != len(P):
                    raise ValueError("Saisissez une colonne rho_gaz de meme "
                                     "longueur que les donnees (colonne 2 du "
                                     "tableau), puis relancez le calcul.")
                rho_user = rows[:, 1]
                V_pore = to_float_or(self.e_vp.get(), 0.0)
                n_tot, rho = excess_to_total(P, n_exc, T, gas, "Densite fournie",
                                             V_pore=V_pore,
                                             rho_gas_user=rho_user)
            elif meth == "Densite adsorbee":
                rho_ads = to_float_or(self.e_rho.get(), np.nan)
                if self.cb_rho_u.get() == "g/cm3":
                    rho_ads = rho_ads * 1000.0 / GAS_DB.get(gas, {"M": 44.01})["M"]
                n_tot, rho = excess_to_total(P, n_exc, T, gas, meth,
                                             rho_ads=rho_ads, eos=eos)
            else:
                V_pore = to_float_or(self.e_vp.get(), np.nan)
                if not np.isfinite(V_pore) or V_pore < 0:
                    raise ValueError("Volume poreux invalide.")
                n_tot, rho = excess_to_total(P, n_exc, T, gas, meth,
                                             V_pore=V_pore, eos=eos)
        except Exception as e:
            messagebox.showerror("Conversion", str(e), parent=self)
            return

        Z = [eos_z_effective(T, p, gas, eos) for p in P]
        crit, msg = near_critical(gas, T, float(np.max(P)))
        if crit and eos != "CoolProp":
            self.app.log("ATTENTION : " + msg)
        with np.errstate(divide="ignore", invalid="ignore"):
            ecart = 100.0 * (n_tot - n_exc) / np.where(np.abs(n_exc) > 1e-12,
                                                       n_exc, np.nan)
        rows = [[fmt(P[i], 5), fmt(rho[i], 6), fmt(Z[i], 4), fmt(n_exc[i], 5),
                 fmt(n_tot[i], 5), fmt(ecart[i], 2)] for i in range(len(P))]
        self.tbl.set_rows(rows)
        self.result = dict(dataset=d["name"], gas=gas, T=T, method=meth,
                           eos=eos, P=P, n_exc=n_exc, n_tot=n_tot, rho=rho, Z=Z)

        self._render(self.fig)
        self.canvas.draw_idle()
        self.app.log("Conversion effectuee (%s, %s) sur '%s'."
                     % (meth, eos, d["name"]))
        try:
            dq = float(np.nanmax(n_tot - n_exc))
            crit, msg = near_critical(gas, T, float(np.max(P)))
            self.lbl_conv.config(
                text="Ecart maximal exces/total : %s mmol/g  -  %s"
                     % (fmt(dq, 4), msg if crit else "domaine sur"),
                foreground=(UI["danger"] if crit else UI["ok"]))
        except Exception:
            pass

    def _render(self, fig):
        fig.clear()
        ax = fig.add_subplot(211)
        ax2 = fig.add_subplot(212, sharex=ax)
        self.ax = ax
        r = self.result
        if not r:
            return
        ax.plot(r["P"], r["n_exc"], "o-", color=PALETTE[0], ms=5,
                label="Quantite en exces")
        ax.plot(r["P"], r["n_tot"], "s--", color=PALETTE[1], ms=5,
                label="Quantite totale")
        ax.set_ylabel("q (mmol/g)")
        ax.set_title("%s  -  %s (%s)" % (r["dataset"], r["method"], r["eos"]),
                     fontsize=9)
        ax.legend(fontsize=8)
        ax.grid(True, ls="--", alpha=0.5)
        ax2.plot(r["P"], np.asarray(r["n_tot"]) - np.asarray(r["n_exc"]), "^-",
                 color=PALETTE[2], ms=5, label="Correction appliquee")
        ax2.set_xlabel("Pression (bar)")
        ax2.set_ylabel("n_tot - n_exc")
        ax2.grid(True, ls="--", alpha=0.5)
        ax2.legend(fontsize=8)
        try:
            fig.tight_layout()
        except Exception:
            pass

    def create_dataset(self):
        if not self.result:
            messagebox.showwarning("Conversion", "Lancez d'abord le calcul.",
                                   parent=self)
            return
        r = self.result
        d = make_dataset("%s (n_tot)" % r["dataset"], r["gas"], r["T"], "Total",
                         "bar", "mmol/g", r["P"], r["n_tot"],
                         "Converti par '%s' (%s)" % (r["method"], r["eos"]))
        self.app.datasets.append(d)
        i = len(self.app.datasets) - 1
        self.app.notify_datasets()
        try:
            self.app.tab_data.refresh_list(select=i)
            self.app.tab_data.load_selected()
        except Exception:
            pass
        self.app.log("Nouveau jeu cree : %s (%d points)" % (d["name"],
                                                            len(d["P"])))
        if messagebox.askyesno("Conversion",
                               "Jeu '%s' ajoute (%d points).\n\n"
                               "Aller a l'onglet Donnees pour le voir ?"
                               % (d["name"], len(d["P"])), parent=self):
            try:
                self.app.nb.select(0)
            except Exception:
                pass

    def export(self):
        if not self.result:
            messagebox.showwarning("Conversion", "Lancez d'abord le calcul.",
                                   parent=self)
            return
        path = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                            filetypes=[("Excel", "*.xlsx")],
                                            initialfile="conversion_ntot.xlsx")
        if not path:
            return
        r = self.result
        df = pd.DataFrame({"P (bar)": r["P"], "Z": r["Z"],
                           "rho_gaz (mmol/cm3)": r["rho"],
                           "n_exces (mmol/g)": r["n_exc"],
                           "n_total (mmol/g)": r["n_tot"]})
        with pd.ExcelWriter(path, engine="openpyxl") as w:
            df.to_excel(w, sheet_name="Conversion", index=False)
            pd.DataFrame([{"Jeu": r["dataset"], "Gaz": r["gas"], "T (K)": r["T"],
                           "Methode": r["method"], "EOS": r["eos"]}]
                         ).to_excel(w, sheet_name="Parametres", index=False)
        self.app.log("Export : %s" % path)
        messagebox.showinfo("Export", "Fichier enregistre :\n%s" % path,
                            parent=self)


# =============================================================================
# 9. ONGLET 3 : AJUSTEMENT DES MODELES
# =============================================================================

class TabFit(ttk.Frame):
    def __init__(self, master, app):
        super().__init__(master, padding=10)
        self.app = app
        self.checked = {k: (k in ("langmuir", "sips", "toth"))
                        for k in MODEL_ORDER}
        pan = ttk.PanedWindow(self, orient="horizontal")
        pan.pack(fill="both", expand=True)
        lwrap = ScrollFrame(pan)
        right = ttk.Frame(pan)
        pan.add(lwrap, weight=0)
        pan.add(right, weight=1)
        left = lwrap.inner

        # ---------------------------------------------- jeu et contexte ----
        c1 = card(left, "1. Jeu de donnees et contexte")
        c1.pack(fill="x", pady=(0, 8))
        g = c1.body
        self.cb_ds = ttk.Combobox(g, state="readonly", width=46)
        self.cb_ds.grid(row=0, column=0, columnspan=4, sticky="ew", pady=(0, 6))
        self.cb_ds.bind("<<ComboboxSelected>>", lambda e: self.on_dataset())
        self.e_T = field(g, "T (K) :", ttk.Entry, 12, 1, 0,
                         "Temperature, utilisee par D-R et D-A")
        self.e_T.insert(0, "298.15")
        self.e_P0 = field(g, "P0 (bar) :", ttk.Entry, 12, 1, 2,
                          "Pression de saturation ou pseudo-saturation, "
                          "requise par D-R et D-A a P0 fixe")
        tip(ttk.Button(g, text="Estimer P0 automatiquement",
                       style="Ghost.TButton", command=self.estimate_p0),
            "Lee-Kesler si T < Tc, regle de Dubinin Pc(T/Tc)^2 si T > Tc").grid(
            row=2, column=0, columnspan=2, sticky="w", pady=(6, 0))
        self.lbl_p0 = ttk.Label(g, text="", style="Muted.TLabel")
        self.lbl_p0.grid(row=2, column=2, columnspan=2, sticky="w", pady=(6, 0))

        # ----------------------------------------------------- modeles -----
        c2 = card(left, "2. Modeles a ajuster", "cochez dans la colonne de gauche")
        c2.pack(fill="x", pady=(0, 8))
        self.tv_models = ttk.Treeview(c2.body, columns=("chk", "mod"),
                                      show="headings", height=9,
                                      selectmode="browse")
        self.tv_models.heading("chk", text="")
        self.tv_models.heading("mod", text="Modele")
        self.tv_models.column("chk", width=36, anchor="center", stretch=False)
        self.tv_models.column("mod", width=310, anchor="w")
        self.tv_models.pack(fill="x")
        for k in MODEL_ORDER:
            self.tv_models.insert("", "end", iid=k,
                                  values=("X" if self.checked[k] else "",
                                          MODELS[k].label))
        self.tv_models.bind("<Button-1>", self._click_models)
        self.tv_models.bind("<<TreeviewSelect>>", lambda e: self.load_params())
        bm = ttk.Frame(c2.body, style="Card.TFrame")
        bm.pack(fill="x", pady=(8, 0))
        ttk.Button(bm, text="Tout cocher", width=13,
                   command=lambda: self._check_all(True)).pack(side="left")
        ttk.Button(bm, text="Tout decocher", width=13,
                   command=lambda: self._check_all(False)).pack(side="left",
                                                                padx=5)

        # ------------------------------------------------------ options ----
        c3 = card(left, "3. Options d'ajustement")
        c3.pack(fill="x", pady=(0, 8))
        f4 = c3.body
        ttk.Label(f4, text="Ponderation :", style="Card.TLabel").grid(
            row=0, column=0, sticky="w", pady=3)
        self.cb_w = ttk.Combobox(f4, values=WEIGHT_MODES, state="readonly",
                                 width=30)
        self.cb_w.set(WEIGHT_MODES[0])
        self.cb_w.grid(row=0, column=1, sticky="w", padx=6)
        tip(self.cb_w, "1/q^2 donne le meme poids relatif a tous les points : "
                       "utile quand les faibles pressions comptent autant que "
                       "la saturation")
        self.var_extrap = tk.DoubleVar(value=1.1)
        ttk.Label(f4, text="Extrapolation courbe (x P_max) :",
                  style="Card.TLabel").grid(row=1, column=0, sticky="w", pady=3)
        ttk.Entry(f4, textvariable=self.var_extrap, width=10).grid(
            row=1, column=1, sticky="w", padx=6)
        self.var_ms = tk.BooleanVar(value=True)
        ttk.Checkbutton(f4, text="Multi-depart, essais :", variable=self.var_ms
                        ).grid(row=2, column=0, sticky="w", pady=3)
        self.e_ms = ttk.Entry(f4, width=10)
        self.e_ms.insert(0, "12")
        self.e_ms.grid(row=2, column=1, sticky="w", padx=6)
        tip(self.e_ms, "Relance l'ajustement depuis plusieurs points de depart "
                       "aleatoires et garde le meilleur. Recommande pour Toth, "
                       "Sips et D-A qui ont des minima locaux.")
        self.var_bs = tk.BooleanVar(value=False)
        ttk.Checkbutton(f4, text="Bootstrap IC 95 %, tirages :",
                        variable=self.var_bs).grid(row=3, column=0, sticky="w",
                                                   pady=3)
        self.e_bs = ttk.Entry(f4, width=10)
        self.e_bs.insert(0, "200")
        self.e_bs.grid(row=3, column=1, sticky="w", padx=6)
        tip(self.e_bs, "Reajuste le modele N fois sur des residus "
                       "reechantillonnes. Plus fiable que l'ecart-type "
                       "asymptotique quand les parametres sont correles. "
                       "Comptez quelques secondes par modele.")

        # -------------------------------------- valeurs initiales / bornes -
        c4 = card(left, "4. Valeurs initiales et bornes",
                  "du modele surligne ci-dessus")
        c4.pack(fill="x", pady=(0, 8))
        self.tbl_par = EditableTable(c4.body, ["Parametre", "p0", "min", "max"],
                                     height=6, widths=[120, 90, 90, 90])
        self.tbl_par.pack(fill="x")
        ttk.Button(c4.body, text="Reinitialiser (automatique)",
                   style="Ghost.TButton", command=self.load_params).pack(
            anchor="w", pady=(6, 0))

        act = ttk.Frame(left, style="Card.TFrame", padding=12)
        act.pack(fill="x")
        tip(ttk.Button(act, text="Ajuster les modeles coches",
                       style="Primary.TButton", command=self.run_fit),
            "Lance l'ajustement non lineaire sur le jeu selectionne (F5)"
            ).pack(fill="x")
        tip(ttk.Button(act, text="Ajuster TOUS les jeux x TOUS les modeles "
                                 "coches", style="Primary.TButton",
                       command=self.run_fit_all),
            "Boucle sur l'ensemble des jeux de donnees et lance les modeles "
            "coches sur chacun. Une barre de progression indique l'avancement "
            "(Maj+F5).").pack(fill="x", pady=(6, 0))
        ttk.Button(act, text="Classement des modeles (AIC)",
                   command=self.show_ranking).pack(fill="x", pady=(6, 0))
        tip(ttk.Button(act, text="Matrice de comparaison des modeles",
                       command=self.show_matrix),
            "Tableau croise jeu x modele : R2, RMSE et AIC cote a cote, "
            "avec le meilleur modele signale par jeu.").pack(fill="x",
                                                             pady=(6, 0))
        zoom_button(act, self._render, "Ajustement", app).pack(fill="x",
                                                               pady=(6, 0))

        # ------------------------------------------------- panneau droit ---
        rf = card(right, "Resultats des ajustements",
                  "double-clic sur une ligne pour le detail complet")
        rf.pack(fill="both", expand=False)
        cols = ("ds", "mod", "N", "R2", "R2a", "RMSE", "MSSR", "AIC", "rmax",
                "par")
        heads = ("Jeu", "Modele", "N", "R2", "R2 ajuste", "RMSE", "MSSR",
                 "AIC", "|r|max", "Parametres")
        widths = (140, 175, 40, 80, 80, 80, 80, 80, 70, 430)
        tvf = ttk.Frame(rf.body, style="Card.TFrame")
        tvf.pack(fill="both", expand=True)
        self.tv_fits = ttk.Treeview(tvf, columns=cols, show="headings", height=9)
        for c, h, w in zip(cols, heads, widths):
            self.tv_fits.heading(c, text=h)
            self.tv_fits.column(c, width=w, anchor="w")
        vs = ttk.Scrollbar(tvf, orient="vertical", command=self.tv_fits.yview)
        self.tv_fits.configure(yscrollcommand=vs.set)
        self.tv_fits.pack(side="left", fill="both", expand=True)
        vs.pack(side="right", fill="y")
        self.tv_fits.bind("<Double-1>", lambda e: self.show_details())
        b2 = ttk.Frame(rf.body, style="Card.TFrame")
        b2.pack(fill="x", pady=(8, 0))
        for txt, cmd, sty, hint in (
                ("Details", self.show_details, "Primary.TButton",
                 "Parametres, IC, matrice de correlation et statistiques"),
                ("Exporter les courbes (P, q)...", self.export_curves, "",
                 "Courbe lissee et points experimentaux en .xlsx ou .csv"),
                ("Exporter tout en Excel", self.export_fits, "",
                 "Parametres, courbes et residus de tous les ajustements"),
                ("Vers Graphiques", self.send_to_plot, "",
                 "Envoyer donnees et courbes vers l'onglet Graphiques"),
                ("Supprimer", self.delete_fit, "", "Retirer l'ajustement "
                 "selectionne"),
                ("Tout supprimer", self.clear_fits, "Danger.TButton",
                 "Effacer tous les ajustements")):
            kw = {"style": sty} if sty else {}
            tip(ttk.Button(b2, text=txt, command=cmd, **kw),
                hint).pack(side="left", padx=(0, 5))

        cg = card(right, "Isotherme et residus")
        cg.pack(fill="both", expand=True, pady=(10, 0))
        self.fig = Figure(figsize=(7, 5.4), dpi=100, facecolor=UI["card"])
        self.ax = self.fig.add_subplot(211)
        self.axr = self.fig.add_subplot(212)
        self.canvas = FigureCanvasTkAgg(self.fig, master=cg.body)
        self.canvas.get_tk_widget().pack(fill="both", expand=True)
        if NavigationToolbar2Tk is not None:
            NavigationToolbar2Tk(self.canvas, cg.body).update()
        set_sash(pan, 0, 470)
        self.load_params()

    def refresh(self):
        vals = ["%d. %s [%s, %.1f K, %s]" % (i + 1, d["name"], d["gas"],
                                             d["T"], d["kind"])
                for i, d in enumerate(self.app.datasets)]
        cur = self.cb_ds.get()
        self.cb_ds["values"] = vals
        if cur in vals:
            self.cb_ds.set(cur)
        elif vals:
            self.cb_ds.current(0)
            self.on_dataset()
        self.refresh_fits()

    def selected_dataset(self):
        s = self.cb_ds.get()
        if not s:
            return None
        try:
            return self.app.datasets[int(s.split(".")[0]) - 1]
        except Exception:
            return None

    def on_dataset(self):
        d = self.selected_dataset()
        if not d:
            return
        self.e_T.delete(0, "end")
        self.e_T.insert(0, fmt(d["T"], 2))
        self.estimate_p0()
        self.load_params()
        self.draw()

    def estimate_p0(self):
        d = self.selected_dataset()
        if not d:
            return
        T = to_float_or(self.e_T.get(), d["T"])
        p0, meth = saturation_pressure(d["gas"], T)
        if not np.isfinite(p0):
            p0 = float(np.max(d["P"])) * 1.5
            meth = "1.5 x P_max (par defaut)"
        self.e_P0.delete(0, "end")
        self.e_P0.insert(0, fmt(p0, 4))
        self.lbl_p0.config(text=meth)

    def ctx(self):
        d = self.selected_dataset()
        T = to_float_or(self.e_T.get(), d["T"] if d else 298.15)
        P0 = to_float_or(self.e_P0.get(), np.nan)
        if not np.isfinite(P0) or P0 <= 0:
            P0 = saturation_pressure(d["gas"], T)[0] if d else np.nan
            if not np.isfinite(P0) or P0 <= 0:
                P0 = (float(np.max(d["P"])) * 1.5) if d else 1.0
        return {"T": T, "P0": P0}

    def _click_models(self, event):
        if self.tv_models.identify_region(event.x, event.y) != "cell":
            return
        if self.tv_models.identify_column(event.x) != "#1":
            return
        iid = self.tv_models.identify_row(event.y)
        if not iid:
            return
        self.checked[iid] = not self.checked[iid]
        self.tv_models.set(iid, "chk", "X" if self.checked[iid] else "")
        return "break"

    def _check_all(self, state):
        for k in MODEL_ORDER:
            self.checked[k] = state
            self.tv_models.set(k, "chk", "X" if state else "")

    def load_params(self):
        sel = self.tv_models.selection()
        key = sel[0] if sel else MODEL_ORDER[0]
        m = MODELS[key]
        d = self.selected_dataset()
        rows = []
        if d:
            P = np.asarray(d["P"], float)
            q = np.asarray(d["q"], float)
            msk = P > 0
            try:
                g = m.guess(P[msk], q[msk], self.ctx())
                lb, ub = m.bounds(P[msk], q[msk], self.ctx())
            except Exception:
                g = [1.0] * m.nparam
                lb, ub = [0.0] * m.nparam, [np.inf] * m.nparam
        else:
            g = [1.0] * m.nparam
            lb, ub = [0.0] * m.nparam, [np.inf] * m.nparam
        for i, pn in enumerate(m.pnames):
            rows.append(["%s (%s)" % (pn, m.punits[i]), fmt(g[i], 6),
                         fmt(lb[i], 6), fmt(ub[i], 6) if np.isfinite(ub[i]) else "inf"])
        self.tbl_par.set_rows(rows)
        self._param_model = key

    def _read_params(self, m):
        if getattr(self, "_param_model", None) != m.key:
            return None, None, None
        rows = self.tbl_par.get_rows()
        if len(rows) != m.nparam:
            return None, None, None
        p0, lo, hi = [], [], []
        for r in rows:
            p0.append(to_float(r[1]))
            v = to_float(r[2])
            lo.append(None if np.isnan(v) else v)
            t = str(r[3]).strip().lower()
            hi.append(np.inf if t in ("inf", "+inf", "") else to_float(r[3]))
        if any(np.isnan(v) for v in p0):
            p0 = None
        return p0, lo, hi

    # ------------------------------------------------------------------
    def run_fit_all(self):
        """Ajuste tous les jeux avec les modeles coches."""
        keys = [k for k in MODEL_ORDER if self.checked[k]]
        if not keys:
            messagebox.showwarning("Ajustement", "Cochez au moins un modele.",
                                   parent=self)
            return
        ds = list(self.app.datasets)
        if not ds:
            messagebox.showwarning("Ajustement", "Aucun jeu de donnees.",
                                   parent=self)
            return
        if not messagebox.askyesno(
                "Ajustement en lot",
                "Lancer %d ajustement(s) (%d jeux x %d modeles) ?\n\n"
                "Les ajustements existants portant le meme couple "
                "jeu / modele seront remplaces."
                % (len(ds) * len(keys), len(ds), len(keys)), parent=self):
            return
        n_starts = (int(to_float_or(self.e_ms.get(), 12))
                    if self.var_ms.get() else 1)
        total = len(ds) * len(keys)
        done, nok, nfail, fails = 0, 0, 0, []
        self.app.busy(True, 0)
        for d in ds:
            P0b = saturation_pressure(d["gas"], float(d["T"]))[0]
            if not np.isfinite(P0b) or P0b <= 0:
                P0b = float(np.max(d["P"])) * 1.5
            ctx = {"T": float(d["T"]), "P0": P0b}
            P = np.asarray(d["P"], float)
            q = np.asarray(d["q"], float)
            for k in keys:
                m = MODELS[k]
                done += 1
                self.app.set_status("Ajustement %d / %d  -  %s / %s"
                                    % (done, total, d["name"], m.label))
                self.app.busy(True, 100.0 * done / total)
                try:
                    self.update_idletasks()
                except Exception:
                    pass
                try:
                    res = fit_isotherm(m, P, q, ctx, weights=self.cb_w.get(),
                                       n_starts=n_starts, n_boot=0)
                except Exception as e:
                    nfail += 1
                    fails.append("%s / %s : %s" % (d["name"], m.label, e))
                    continue
                res["dataset"] = d["name"]
                res["gas"] = d["gas"]
                res["label"] = "%s - %s" % (d["name"], m.label)
                self.app.fits = [f for f in self.app.fits
                                 if not (f["dataset"] == d["name"]
                                         and f["model_key"] == k)]
                self.app.fits.append(res)
                nok += 1
        self.app.busy(False)
        self.refresh_fits()
        self.draw()
        self.app.notify_fits()
        self.app.log("Ajustement en lot : %d reussi(s), %d echec(s)."
                     % (nok, nfail))
        msg = "%d ajustement(s) reussi(s), %d echec(s)." % (nok, nfail)
        if fails:
            msg += "\n\nEchecs :\n  - " + "\n  - ".join(fails[:12])
        messagebox.showinfo("Ajustement en lot", msg, parent=self)
        self.app.set_status("Ajustement en lot termine.")

    def show_matrix(self):
        """Tableau croise jeu x modele."""
        if not self.app.fits:
            messagebox.showinfo("Matrice", "Aucun ajustement disponible.",
                                parent=self)
            return
        dss, mods = [], []
        for f in self.app.fits:
            if f["dataset"] not in dss:
                dss.append(f["dataset"])
            if f["model_key"] not in mods:
                mods.append(f["model_key"])
        mods = [k for k in MODEL_ORDER if k in mods]
        lines = []
        lines.append("MATRICE DE COMPARAISON DES MODELES")
        lines.append("=" * 108)
        lines.append("Un bloc par indicateur : R2, RMSE (mmol/g) puis AICc.")
        lines.append("Le modele retenu par jeu (AICc minimal) est marque <<.")
        lines.append("")
        for metric, lab in (("R2", "R2"), ("RMSE", "RMSE (mmol/g)"),
                            ("AICc", "AICc")):
            lines.append("-" * 108)
            lines.append("%-26s" % lab + "".join("%-20s" % MODELS[k].label[:19]
                                                 for k in mods))
            lines.append("-" * 108)
            for dname in dss:
                row = "%-26s" % dname[:25]
                best, bestv = None, np.inf
                for k in mods:
                    f = next((x for x in self.app.fits
                              if x["dataset"] == dname and x["model_key"] == k),
                             None)
                    if f and np.isfinite(f["stats"].get("AICc", np.nan)):
                        if f["stats"]["AICc"] < bestv:
                            bestv, best = f["stats"]["AICc"], k
                for k in mods:
                    f = next((x for x in self.app.fits
                              if x["dataset"] == dname and x["model_key"] == k),
                             None)
                    if f is None:
                        row += "%-20s" % "-"
                    else:
                        v = f["stats"].get(metric, np.nan)
                        txt = fmt(v, 5)
                        if metric == "AICc" and k == best:
                            txt += " <<"
                        row += "%-20s" % txt
                lines.append(row)
            lines.append("")
        show_text_window(self, "Matrice de comparaison des modeles",
                         "\n".join(lines))

    def run_fit(self):
        d = self.selected_dataset()
        if d is None:
            messagebox.showwarning("Ajustement", "Selectionnez un jeu de donnees.",
                                   parent=self)
            return
        keys = [k for k in MODEL_ORDER if self.checked[k]]
        if not keys:
            messagebox.showwarning("Ajustement", "Cochez au moins un modele.",
                                   parent=self)
            return
        ctx = self.ctx()
        P = np.asarray(d["P"], float)
        q = np.asarray(d["q"], float)
        n_starts = (int(to_float_or(self.e_ms.get(), 12))
                    if self.var_ms.get() else 1)
        n_boot = (int(to_float_or(self.e_bs.get(), 200))
                  if self.var_bs.get() else 0)
        nok, nfail = 0, 0
        for k in keys:
            m = MODELS[k]
            p0, lo, hi = self._read_params(m)

            def prog(i, tot, lab=m.label):
                self.app.log("  bootstrap %s : %d / %d" % (lab, i, tot))
                try:
                    self.update_idletasks()
                except Exception:
                    pass
            try:
                if n_boot:
                    self.app.log("Ajustement %s (%d departs, %d tirages "
                                 "bootstrap)..." % (m.label, n_starts, n_boot))
                    try:
                        self.update_idletasks()
                    except Exception:
                        pass
                res = fit_isotherm(m, P, q, ctx, p0=p0, lower=lo, upper=hi,
                                   weights=self.cb_w.get(), n_starts=n_starts,
                                   n_boot=n_boot,
                                   progress=prog if n_boot else None)
            except Exception as e:
                nfail += 1
                self.app.log("[ECHEC] %s / %s : %s" % (d["name"], m.label, e))
                continue
            res["dataset"] = d["name"]
            res["gas"] = d["gas"]
            res["label"] = "%s - %s" % (d["name"], m.label)
            self.app.fits = [f for f in self.app.fits
                             if not (f["dataset"] == d["name"]
                                     and f["model_key"] == k)]
            self.app.fits.append(res)
            nok += 1
        self.refresh_fits()
        self.draw()
        self.app.notify_fits()
        self.app.log("Ajustement : %d reussi(s), %d echec(s) sur '%s'."
                     % (nok, nfail, d["name"]))

    def refresh_fits(self):
        self.tv_fits.delete(*self.tv_fits.get_children())
        self.tv_fits.tag_configure("corr_forte", background="#ffd6d6")
        self.tv_fits.tag_configure("corr_moyenne", background="#fff2cc")
        for i, f in enumerate(self.app.fits):
            m = MODELS[f["model_key"]]
            ps = ", ".join("%s=%s" % (m.pnames[j], fmt(f["params"][j], 4))
                           for j in range(m.nparam))
            s = f["stats"]
            rm = f.get("r_max", np.nan)
            tag = ("corr_forte" if (np.isfinite(rm) and rm > 0.99)
                   else ("corr_moyenne" if (np.isfinite(rm) and rm > 0.95)
                         else ""))
            self.tv_fits.insert("", "end", iid=str(i),
                                tags=((tag,) if tag else ()),
                                values=(f["dataset"], m.label, s["N"],
                                        fmt(s["R2"], 5), fmt(s["R2adj"], 5),
                                        fmt(s["RMSE"], 5), fmt(s["MSSR"], 6),
                                        fmt(s["AIC"], 2), fmt(rm, 4), ps))

    def _sel_fit(self):
        sel = self.tv_fits.selection()
        return int(sel[0]) if sel else None

    def delete_fit(self):
        i = self._sel_fit()
        if i is None:
            return
        self.app.fits.pop(i)
        self.refresh_fits()
        self.draw()
        self.app.notify_fits()

    def clear_fits(self):
        if messagebox.askyesno("Ajustements", "Supprimer tous les ajustements ?",
                               parent=self):
            self.app.fits = []
            self.refresh_fits()
            self.draw()
            self.app.notify_fits()

    def show_details(self):
        i = self._sel_fit()
        if i is None:
            return
        f = self.app.fits[i]
        m = MODELS[f["model_key"]]
        s = f["stats"]
        lines = ["Jeu de donnees : %s" % f["dataset"],
                 "Modele         : %s" % m.label,
                 "Equation       : %s" % m.formula,
                 "Contexte       : T = %s K, P0 = %s bar"
                 % (fmt(f["ctx"].get("T"), 2), fmt(f["ctx"].get("P0"), 4)),
                 "Ponderation    : %s" % f.get("weights", "-"), "",
                 "%-16s %14s %14s %10s" % ("Parametre", "Valeur", "Ecart-type",
                                           "Unite")]
        for j, pn in enumerate(m.pnames):
            lines.append("%-16s %14s %14s %10s"
                         % (pn, fmt(f["params"][j], 6), fmt(f["errors"][j], 6),
                            m.punits[j]))
        if f.get("n_starts", 1) > 1:
            lines.append("")
            lines.append("Multi-depart : %d essais, %d convergents."
                         % (f.get("n_starts", 1), f.get("starts_ok", 1)))
        if f.get("ci_lo"):
            lines += ["", "Intervalles de confiance a 95 %% par bootstrap "
                          "(%d tirages retenus) :" % f.get("n_boot", 0),
                      "%-16s %14s %14s %14s %14s"
                      % ("Parametre", "Valeur", "IC 2.5 %", "IC 97.5 %",
                         "ecart-type")]
            for j, pn in enumerate(m.pnames):
                eb = (f["err_boot"][j] if f.get("err_boot") else np.nan)
                lines.append("%-16s %14s %14s %14s %14s"
                             % (pn, fmt(f["params"][j], 6),
                                fmt(f["ci_lo"][j], 6), fmt(f["ci_hi"][j], 6),
                                fmt(eb, 6)))
            try:
                worst = max(range(m.nparam),
                            key=lambda j: (f["ci_hi"][j] - f["ci_lo"][j])
                            / max(abs(f["params"][j]), 1e-12))
                lg = ((f["ci_hi"][worst] - f["ci_lo"][worst])
                      / max(abs(f["params"][worst]), 1e-12) * 100.0)
                if lg > 50:
                    lines.append("  -> parametre le moins bien determine : %s "
                                 "(largeur de l'IC = %.0f %% de la valeur)."
                                 % (m.pnames[worst], lg))
            except Exception:
                pass
        corr = f.get("corr")
        if corr and m.nparam > 1:
            lines += ["", "Matrice de correlation des parametres :",
                      "%-14s%s" % ("", "".join("%12s" % pn
                                               for pn in m.pnames))]
            for i2, pn in enumerate(m.pnames):
                lines.append("%-14s%s" % (pn, "".join("%12s" % fmt(corr[i2][j2], 4)
                                                      for j2 in range(m.nparam))))
            rm = f.get("r_max", np.nan)
            if np.isfinite(rm):
                if rm > 0.99:
                    lines.append("  -> |r|max = %.4f : parametres quasi "
                                 "indissociables. q_max et b compensent "
                                 "mutuellement ; les valeurs individuelles ne "
                                 "sont pas identifiables de facon fiable, meme "
                                 "avec un R2 excellent." % rm)
                elif rm > 0.95:
                    lines.append("  -> |r|max = %.4f : forte correlation, "
                                 "interpretez les parametres avec prudence."
                                 % rm)
                else:
                    lines.append("  -> |r|max = %.4f : correlation acceptable."
                                 % rm)
        der = m.derived(f["params"], f["ctx"])
        if der:
            lines.append("")
            lines.append("Grandeurs derivees :")
            for k, (v, u) in der.items():
                lines.append("  %-28s %14s %s" % (k, fmt(v, 5), u))
        lines += ["", "Statistiques :",
                  "  N      = %d" % s["N"],
                  "  R2     = %s" % fmt(s["R2"], 6),
                  "  R2 adj = %s" % fmt(s["R2adj"], 6),
                  "  SSR    = %s" % fmt(s["SSR"], 6),
                  "  MSSR   = %s" % fmt(s["MSSR"], 8),
                  "  RMSE   = %s" % fmt(s["RMSE"], 6),
                  "  ARE    = %s %%" % fmt(s["ARE"], 3),
                  "  chi2   = %s" % fmt(s["chi2"], 6),
                  "  AIC    = %s   AICc = %s   BIC = %s"
                  % (fmt(s["AIC"], 3), fmt(s["AICc"], 3), fmt(s["BIC"], 3))]
        if m.note:
            lines += ["", "Remarque : %s" % m.note]
        show_text_window(self, "Details de l'ajustement", "\n".join(lines))

    def show_ranking(self):
        d = self.selected_dataset()
        if d is None:
            return
        fs = [f for f in self.app.fits if f["dataset"] == d["name"]]
        if not fs:
            messagebox.showinfo("Classement", "Aucun ajustement pour ce jeu.",
                                parent=self)
            return
        fs.sort(key=lambda f: f["stats"]["AIC"])
        best = fs[0]["stats"]["AIC"]
        lines = ["Classement des modeles pour '%s'" % d["name"], "",
                 "%-32s %9s %9s %10s %8s %10s"
                 % ("Modele", "R2", "R2 adj", "RMSE", "k", "dAIC")]
        for f in fs:
            m = MODELS[f["model_key"]]
            s = f["stats"]
            lines.append("%-32s %9s %9s %10s %8d %10s"
                         % (m.label, fmt(s["R2"], 5), fmt(s["R2adj"], 5),
                            fmt(s["RMSE"], 5), m.nparam,
                            fmt(s["AIC"] - best, 2)))
        lines += ["", "dAIC = 0 : meilleur compromis qualite / nombre de "
                      "parametres.", "dAIC < 2 : modeles equivalents.",
                  "dAIC > 10 : modele nettement moins bon."]
        show_text_window(self, "Classement des modeles", "\n".join(lines))

    def _extrap(self):
        try:
            v = float(self.var_extrap.get())
        except Exception:
            v = 1.1
        return v if 0.1 <= v <= 100 else 1.1

    # ------------------------------------------------------------------
    def draw(self):
        self._render(self.fig)
        self.canvas.draw_idle()

    def _render(self, fig):
        """Trace donnees, courbes ajustees et residus sur la figure fournie."""
        fig.clear()
        gs = fig.add_gridspec(3, 1)
        ax = fig.add_subplot(gs[0:2, 0])
        axr = fig.add_subplot(gs[2, 0], sharex=ax)
        self.ax, self.axr = ax, axr
        d = self.selected_dataset()
        if d is None:
            return
        P = np.asarray(d["P"], float)
        q = np.asarray(d["q"], float)
        ax.plot(P, q, "o", color="k", ms=6, zorder=5,
                label="Donnees : %s" % d["name"])
        fs = [f for f in self.app.fits if f["dataset"] == d["name"]]
        xmax = float(np.max(P)) * self._extrap()
        Pfit = np.linspace(max(1e-6, float(np.min(P)) * 0.01), xmax, 400)
        for i, f in enumerate(fs):
            m = MODELS[f["model_key"]]
            try:
                y = m.func(Pfit, f["ctx"], *f["params"])
            except Exception:
                continue
            c = PALETTE[i % len(PALETTE)]
            ax.plot(Pfit, y, "-", color=c, lw=1.8,
                    label="%s (R2=%s)" % (m.label, fmt(f["stats"]["R2"], 4)))
            axr.plot(f["P"], f["residuals"], "o", color=c, ms=4)
        axr.axhline(0, color="k", lw=0.8)
        ax.set_ylabel("q (mmol/g)")
        ax.grid(True, ls="--", alpha=0.5)
        ax.legend(fontsize=8)
        axr.set_xlabel("Pression (bar)")
        axr.set_ylabel("Residus")
        axr.grid(True, ls="--", alpha=0.5)
        fig.subplots_adjust(left=0.10, right=0.98, top=0.96, bottom=0.10,
                            hspace=0.30)

    def send_to_plot(self):
        d = self.selected_dataset()
        if d is None:
            return
        self.app.add_series(d["name"], np.asarray(d["P"]), np.asarray(d["q"]),
                            kind="points")
        for f in [f for f in self.app.fits if f["dataset"] == d["name"]]:
            m = MODELS[f["model_key"]]
            xmax = float(np.max(f["P"])) * self._extrap()
            x = np.linspace(max(1e-6, float(np.min(f["P"])) * 0.01), xmax, 400)
            y = m.func(x, f["ctx"], *f["params"])
            self.app.add_series("%s - %s" % (d["name"], m.label), x, y,
                                kind="ligne")
        self.app.tab_plot.refresh_series()
        self.app.log("Series envoyees vers l'onglet Graphiques.")

    def export_fits(self):
        if not self.app.fits:
            messagebox.showwarning("Export", "Aucun ajustement a exporter.",
                                   parent=self)
            return
        path = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                            filetypes=[("Excel", "*.xlsx")],
                                            initialfile="ajustements.xlsx")
        if not path:
            return
        try:
            export_fits_excel(self.app, path)
        except Exception as e:
            messagebox.showerror("Export", "%s\n\n%s"
                                 % (e, traceback.format_exc()), parent=self)
            return
        self.app.log("Export des ajustements : %s" % path)
        messagebox.showinfo("Export",
                            "Fichier enregistre :\n%s\n\nFeuilles : "
                            "Parametres / Courbes ajustees / Residus" % path,
                            parent=self)

    def export_curves(self):
        """Export des donnees d'ajustement : P et q calculee (+ points exp.)."""
        if not self.app.fits:
            messagebox.showwarning("Export", "Aucun ajustement a exporter.",
                                   parent=self)
            return
        d = self.selected_dataset()
        i = self._sel_fit()
        sel = [self.app.fits[i]] if i is not None else []
        of_ds = [f for f in self.app.fits
                 if d is not None and f["dataset"] == d["name"]]
        ref = sel or of_ds or self.app.fits
        P = np.asarray(ref[0]["P"], dtype=float)
        dlg = CurveExportDialog(self, len(sel), len(of_ds), len(self.app.fits),
                                max(1e-9, float(P.min()) * 0.01),
                                float(P.max()) * self._extrap())
        if not dlg.result:
            return
        o = dlg.result
        fits = [sel, of_ds, self.app.fits][o["scope"]]
        if not fits:
            messagebox.showwarning("Export", "Aucun ajustement dans cette "
                                             "selection.", parent=self)
            return
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx", initialfile="courbes_ajustement.xlsx",
            filetypes=[("Excel", "*.xlsx"), ("CSV", "*.csv")])
        if not path:
            return
        try:
            df_c, df_e = build_curve_frames(fits, o["pmin"], o["pmax"],
                                            o["npts"], o["exp"])
            if path.lower().endswith(".csv"):
                base = os.path.splitext(path)[0]
                df_c.to_csv(base + "_courbes.csv", index=False)
                written = [base + "_courbes.csv"]
                if o["exp"] and not df_e.empty:
                    df_e.to_csv(base + "_points.csv", index=False)
                    written.append(base + "_points.csv")
                msg = "\n".join(written)
            else:
                sub = fits_summary_frame(self.app)
                keys = {(f["dataset"], MODELS[f["model_key"]].label)
                        for f in fits}
                if not sub.empty:
                    sub = sub[[(r["Jeu de donnees"], r["Modele"]) in keys
                               for _, r in sub.iterrows()]]
                with pd.ExcelWriter(path, engine="openpyxl") as w:
                    df_c.to_excel(w, sheet_name="Courbes ajustees", index=False)
                    if o["exp"] and not df_e.empty:
                        df_e.to_excel(w, sheet_name="Points et residus",
                                      index=False)
                    if o["par"] and not sub.empty:
                        sub.to_excel(w, sheet_name="Parametres", index=False)
                msg = path
        except Exception as e:
            messagebox.showerror("Export", "%s\n\n%s"
                                 % (e, traceback.format_exc()), parent=self)
            return
        self.app.log("Export des courbes ajustees (%d ajustement(s)) : %s"
                     % (len(fits), path))
        messagebox.showinfo("Export", "Donnees enregistrees :\n%s" % msg,
                            parent=self)


def build_curve_frames(fits, pmin=np.nan, pmax=np.nan, npts=300,
                       include_exp=True):
    """Construit (courbes lissees, points experimentaux) pour une liste de fits."""
    curves, exp = {}, {}
    for f in fits:
        m = MODELS[f["model_key"]]
        P = np.asarray(f["P"], dtype=float)
        a = pmin if np.isfinite(pmin) else max(1e-9, float(P.min()) * 0.01)
        b = pmax if np.isfinite(pmax) else float(P.max()) * 1.1
        if b <= a:
            b = a * 1.1 + 1e-9
        x = np.linspace(a, b, int(max(2, npts)))
        y = np.asarray(m.func(x, f["ctx"], *f["params"]), dtype=float).ravel()
        key = "%s - %s" % (f["dataset"], m.label)
        curves["%s | P (bar)" % key] = x
        curves["%s | q calcule (mmol/g)" % key] = y
        if include_exp:
            exp["%s | P (bar)" % key] = list(f["P"])
            exp["%s | q experimental (mmol/g)" % key] = list(f["q"])
            exp["%s | q calcule (mmol/g)" % key] = list(f["qpred"])
            exp["%s | residu (mmol/g)" % key] = list(f["residuals"])
    df_c = pd.DataFrame(curves) if curves else pd.DataFrame()
    if exp:
        n = max(len(v) for v in exp.values())
        for k in exp:
            exp[k] = list(exp[k]) + [np.nan] * (n - len(exp[k]))
    df_e = pd.DataFrame(exp) if exp else pd.DataFrame()
    return df_c, df_e


class CurveExportDialog(tk.Toplevel):
    """Options d'export des donnees d'ajustement."""

    def __init__(self, master, n_sel, n_ds, n_all, pmin, pmax):
        super().__init__(master)
        self.title("Exporter les donnees d'ajustement")
        self.transient(master)
        self.grab_set()
        self.result = None
        f = ttk.Frame(self, padding=12)
        f.pack(fill="both", expand=True)
        ttk.Label(f, text="Que faut-il exporter ?").grid(row=0, column=0,
                                                          columnspan=4,
                                                          sticky="w")
        self.cb_scope = ttk.Combobox(f, state="readonly", width=52, values=[
            "Ajustement selectionne (%d)" % n_sel,
            "Tous les ajustements du jeu de donnees courant (%d)" % n_ds,
            "Tous les ajustements (%d)" % n_all])
        self.cb_scope.current(0 if n_sel else 1)
        self.cb_scope.grid(row=1, column=0, columnspan=4, sticky="w", pady=4)

        ttk.Label(f, text="P min (bar) :").grid(row=2, column=0, sticky="w")
        self.e_min = ttk.Entry(f, width=12)
        self.e_min.insert(0, fmt(pmin, 5))
        self.e_min.grid(row=2, column=1, sticky="w", padx=3)
        ttk.Label(f, text="P max (bar) :").grid(row=2, column=2, sticky="e")
        self.e_max = ttk.Entry(f, width=12)
        self.e_max.insert(0, fmt(pmax, 5))
        self.e_max.grid(row=2, column=3, sticky="w", padx=3)
        ttk.Label(f, text="Nombre de points :").grid(row=3, column=0, sticky="w")
        self.e_np = ttk.Entry(f, width=12)
        self.e_np.insert(0, "300")
        self.e_np.grid(row=3, column=1, sticky="w", padx=3)
        self.v_exp = tk.BooleanVar(value=True)
        ttk.Checkbutton(f, text="Inclure les points experimentaux, la valeur "
                               "calculee et le residu",
                        variable=self.v_exp).grid(row=4, column=0, columnspan=4,
                                                  sticky="w", pady=3)
        self.v_par = tk.BooleanVar(value=True)
        ttk.Checkbutton(f, text="Inclure une feuille avec les parametres et "
                               "les statistiques",
                        variable=self.v_par).grid(row=5, column=0, columnspan=4,
                                                  sticky="w")
        ttk.Label(f, foreground="#555", justify="left",
                  text="La courbe lissee est echantillonnee entre P min et "
                       "P max.\nLaisser vide pour utiliser la gamme des "
                       "donnees experimentales.").grid(row=6, column=0,
                                                       columnspan=4, sticky="w",
                                                       pady=(6, 2))
        bar = ttk.Frame(f)
        bar.grid(row=7, column=0, columnspan=4, sticky="e", pady=(8, 0))
        ttk.Button(bar, text="Annuler", command=self.destroy).pack(side="right",
                                                                   padx=4)
        ttk.Button(bar, text="Exporter...", command=self._ok).pack(side="right")
        self.wait_window(self)

    def _ok(self):
        self.result = dict(scope=self.cb_scope.current(),
                           pmin=to_float_or(self.e_min.get(), np.nan),
                           pmax=to_float_or(self.e_max.get(), np.nan),
                           npts=int(to_float_or(self.e_np.get(), 300)),
                           exp=bool(self.v_exp.get()),
                           par=bool(self.v_par.get()))
        self.destroy()


def show_text_window(master, title, text):
    top = tk.Toplevel(master)
    top.title(title)
    top.geometry("820x560")
    txt = tk.Text(top, wrap="none", font=("Consolas", 10))
    vs = ttk.Scrollbar(top, orient="vertical", command=txt.yview)
    txt.configure(yscrollcommand=vs.set)
    txt.pack(side="left", fill="both", expand=True)
    vs.pack(side="right", fill="y")
    txt.insert("1.0", text)
    txt.config(state="disabled")
    return top


# =============================================================================
# 10. ONGLET 4 : IAST
# =============================================================================

class ParamDialog(tk.Toplevel):
    """Saisie manuelle d'un modele et de ses parametres."""

    def __init__(self, master, spec=None):
        super().__init__(master)
        self.title("Parametres du constituant")
        self.transient(master)
        self.grab_set()
        self.result = None
        frm = ttk.Frame(self, padding=10)
        frm.pack(fill="both", expand=True)
        ttk.Label(frm, text="Modele :").grid(row=0, column=0, sticky="w")
        self.cb = ttk.Combobox(frm, state="readonly", width=34,
                               values=[MODELS[k].label for k in MODEL_ORDER])
        self.cb.grid(row=0, column=1, columnspan=3, sticky="w", pady=3)
        self.cb.bind("<<ComboboxSelected>>", lambda e: self._model_changed())
        ttk.Label(frm, text="T (K) :").grid(row=1, column=0, sticky="w")
        self.e_T = ttk.Entry(frm, width=12)
        self.e_T.grid(row=1, column=1, sticky="w")
        ttk.Label(frm, text="P0 (bar) :").grid(row=1, column=2, sticky="e")
        self.e_P0 = ttk.Entry(frm, width=12)
        self.e_P0.grid(row=1, column=3, sticky="w")
        ttk.Label(frm, text="P experimentale max (bar) :").grid(
            row=4, column=0, columnspan=2, sticky="w")
        self.e_pexp = ttk.Entry(frm, width=12)
        self.e_pexp.grid(row=4, column=2, sticky="w")
        ttk.Label(frm, foreground="#555", justify="left",
                  text="Sert uniquement a signaler l'extrapolation du modele "
                       "pendant le calcul IAST.").grid(row=5, column=0,
                                                       columnspan=4, sticky="w")
        self.tbl = EditableTable(frm, ["Parametre", "Valeur"], height=6,
                                 widths=[160, 140])
        self.tbl.grid(row=2, column=0, columnspan=4, sticky="nsew", pady=6)
        frm.rowconfigure(2, weight=1)
        bar = ttk.Frame(frm)
        bar.grid(row=3, column=0, columnspan=4, sticky="e")
        ttk.Button(bar, text="Annuler", command=self.destroy).pack(side="right",
                                                                   padx=4)
        ttk.Button(bar, text="Valider", command=self._ok).pack(side="right")

        spec = spec or dict(model_key="langmuir", params=[10.0, 0.1],
                            ctx={"T": 298.15, "P0": 1.0})
        self.cb.set(MODELS[spec["model_key"]].label)
        self.e_T.insert(0, fmt(spec["ctx"].get("T", 298.15), 2))
        self.e_P0.insert(0, fmt(spec["ctx"].get("P0", 1.0), 4))
        if spec.get("P_exp_max"):
            self.e_pexp.insert(0, fmt(spec["P_exp_max"], 4))
        m = MODELS[spec["model_key"]]
        self.tbl.set_rows([["%s (%s)" % (m.pnames[i], m.punits[i]),
                            fmt(spec["params"][i], 6)] for i in range(m.nparam)])
        self.wait_window(self)

    def _key(self):
        lab = self.cb.get()
        for k in MODEL_ORDER:
            if MODELS[k].label == lab:
                return k
        return "langmuir"

    def _model_changed(self):
        m = MODELS[self._key()]
        self.tbl.set_rows([["%s (%s)" % (m.pnames[i], m.punits[i]), ""]
                           for i in range(m.nparam)])

    def _ok(self):
        k = self._key()
        m = MODELS[k]
        rows = self.tbl.get_rows()
        params = [to_float(r[1]) for r in rows[:m.nparam]]
        if any(np.isnan(v) for v in params):
            messagebox.showerror("Parametres", "Toutes les valeurs doivent etre "
                                               "renseignees.", parent=self)
            return
        pexp = to_float_or(self.e_pexp.get(), np.nan)
        self.result = dict(model_key=k, params=params,
                           ctx={"T": to_float_or(self.e_T.get(), 298.15),
                                "P0": to_float_or(self.e_P0.get(), 1.0)},
                           P_exp_max=(float(pexp) if np.isfinite(pexp)
                                      and pexp > 0 else None))
        self.destroy()


class CompRow(ttk.Frame):
    """Ligne de definition d'un constituant du melange."""

    def __init__(self, master, app, idx, default_name="CO2"):
        super().__init__(master)
        self.app = app
        self.idx = idx
        self.spec = None
        ttk.Label(self, text="Constituant %d :" % (idx + 1), width=13
                  ).grid(row=0, column=0, sticky="w")
        self.e_name = ttk.Entry(self, width=10)
        self.e_name.insert(0, default_name)
        self.e_name.grid(row=0, column=1, sticky="w", padx=2)
        self.cb_src = ttk.Combobox(self, state="readonly", width=42)
        self.cb_src.grid(row=0, column=2, sticky="w", padx=2)
        self.cb_src.bind("<<ComboboxSelected>>", lambda e: self._src_changed())
        ttk.Label(self, text="y =").grid(row=0, column=3, sticky="e", padx=(8, 0))
        self.e_y = ttk.Entry(self, width=9)
        self.e_y.insert(0, "0.5")
        self.e_y.grid(row=0, column=4, sticky="w", padx=2)
        ttk.Button(self, text="Editer...", width=9,
                   command=self.edit).grid(row=0, column=5, padx=2)
        self.lbl = ttk.Label(self, text="-", foreground="#333")
        self.lbl.grid(row=1, column=1, columnspan=5, sticky="w", pady=(0, 4))

    def refresh_sources(self):
        vals = ["%d. %s" % (i + 1, f["label"]) for i, f in
                enumerate(self.app.fits)] + ["Saisie manuelle..."]
        cur = self.cb_src.get()
        self.cb_src["values"] = vals
        if cur in vals:
            self.cb_src.set(cur)
        elif len(vals) > 1:
            self.cb_src.set(vals[min(self.idx, len(vals) - 2)])
            self._src_changed()

    def _src_changed(self):
        s = self.cb_src.get()
        if s == "Saisie manuelle...":
            self.edit()
            return
        try:
            f = self.app.fits[int(s.split(".")[0]) - 1]
        except Exception:
            return
        self.spec = dict(model_key=f["model_key"], params=list(f["params"]),
                         ctx=dict(f["ctx"]),
                         P_exp_max=f.get("P_exp_max",
                                         float(np.max(f["P"])) if f.get("P")
                                         else None))
        self._update_label()

    def edit(self):
        dlg = ParamDialog(self, self.spec)
        if dlg.result:
            self.spec = dlg.result
            self.cb_src.set("Saisie manuelle...")
            self._update_label()

    def _update_label(self):
        if not self.spec:
            self.lbl.config(text="-")
            return
        m = MODELS[self.spec["model_key"]]
        txt = "%s : %s" % (m.label, ", ".join(
            "%s=%s" % (m.pnames[i], fmt(self.spec["params"][i], 4))
            for i in range(m.nparam)))
        if m.needs:
            txt += "   [T=%s K, P0=%s bar]" % (fmt(self.spec["ctx"].get("T"), 1),
                                               fmt(self.spec["ctx"].get("P0"), 3))
        pm = self.spec.get("P_exp_max")
        txt += "   P_exp_max=%s bar" % (fmt(pm, 2) if pm else "non renseignee")
        self.lbl.config(text=txt)

    def y_value(self):
        return to_float(self.e_y.get())

    def name(self):
        return self.e_name.get().strip() or "C%d" % (self.idx + 1)


MIX_PRESETS = [
    ("CO2/CH4  50:50", [("CO2", 0.50), ("CH4", 0.50)]),
    ("CO2/CH4  25:75", [("CO2", 0.25), ("CH4", 0.75)]),
    ("CO2/CH4  75:25", [("CO2", 0.75), ("CH4", 0.25)]),
    ("CH4/H2   50:50", [("CH4", 0.50), ("H2", 0.50)]),
    ("CO2/CH4/H2 1:1:1", [("CO2", 1 / 3.), ("CH4", 1 / 3.), ("H2", 1 / 3.)]),
]


class TabIAST(ttk.Frame):
    def __init__(self, master, app):
        super().__init__(master, padding=10)
        self.app = app
        self.df = None

        top = card(self, "1. Constituants du melange",
                   "chaque constituant recoit un modele de corps pur")
        top.pack(fill="x")
        hdr = ttk.Frame(top.body, style="Card.TFrame")
        hdr.pack(fill="x", pady=(0, 6))
        ttk.Label(hdr, text="Nombre de constituants :",
                  style="Card.TLabel").pack(side="left")
        self.var_n = tk.IntVar(value=3)
        for n in (2, 3):
            ttk.Radiobutton(hdr, text="  %d  " % n, value=n, variable=self.var_n,
                            command=self._n_changed).pack(side="left", padx=3)
        tip(ttk.Button(hdr, text="Normaliser les fractions molaires",
                       style="Ghost.TButton", command=self.normalize),
            "Ramene la somme des y a 1").pack(side="left", padx=16)
        self.rows = []
        for i, nm in enumerate(("CO2", "CH4", "H2")):
            r = CompRow(top.body, app, i, nm)
            r.pack(fill="x", pady=2)
            self.rows.append(r)

        pre = ttk.Frame(top.body, style="Card.TFrame")
        pre.pack(fill="x", pady=(8, 0))
        ttk.Label(pre, text="Melanges types :", style="Card.TLabel").pack(
            side="left", padx=(0, 6))
        for lab, spec in MIX_PRESETS:
            tip(ttk.Button(pre, text=lab, style="Ghost.TButton", width=15,
                           command=lambda sp=spec, lb=lab:
                           self.apply_preset(sp, lb)),
                "Configure les constituants et leurs fractions molaires, puis "
                "associe automatiquement l'ajustement disponible pour chaque "
                "gaz.").pack(side="left", padx=2)

        cfg = card(self, "2. Conditions de calcul")
        cfg.pack(fill="x", pady=10)
        g = cfg.body
        ttk.Label(g, text="Mode :", style="Card.TLabel").grid(row=0, column=0,
                                                              sticky="w", pady=3)
        self.cb_mode = ttk.Combobox(g, state="readonly", width=32,
                                    values=["Balayage en pression totale",
                                            "Balayage en composition (y1)"])
        self.cb_mode.set("Balayage en pression totale")
        self.cb_mode.grid(row=0, column=1, sticky="w", padx=6)
        self.e_pmin = field(g, "P min (bar) :", ttk.Entry, 10, 0, 2,
                            "Borne basse du balayage en pression")
        self.e_pmin.insert(0, "0.1")
        self.e_pmax = field(g, "P max (bar) :", ttk.Entry, 10, 0, 4,
                            "Borne haute. Au-dela de la pression mesuree, "
                            "le calcul repose sur l'extrapolation du modele.")
        self.e_pmax.insert(0, "60")
        self.e_np = field(g, "Points :", ttk.Entry, 8, 0, 6,
                          "Nombre de points calcules")
        self.e_np.insert(0, "40")
        self.var_log = tk.BooleanVar(value=False)
        ttk.Checkbutton(g, text="Echelle logarithmique en pression",
                        variable=self.var_log).grid(row=1, column=1, sticky="w",
                                                    pady=3)
        self.e_pfix = field(g, "P fixe pour balayage en y (bar) :", ttk.Entry,
                            10, 1, 2, "Pression totale imposee lors du "
                            "balayage en composition")
        self.e_pfix.insert(0, "1.0")

        bar = ttk.Frame(self, style="Card.TFrame", padding=10)
        bar.pack(fill="x")
        tip(ttk.Button(bar, text="Lancer le calcul IAST",
                       style="Primary.TButton", command=self.run),
            "Resolution sur la pression d'etalement (F5)").pack(side="left")
        ttk.Button(bar, text="Exporter en Excel",
                   command=self.export).pack(side="left", padx=8)
        ttk.Button(bar, text="Vers Graphiques",
                   command=self.send_to_plot).pack(side="left")
        self.lbl_warn = ttk.Label(self, text="", style="MutedBg.TLabel",
                                  wraplength=1200, justify="left",
                                  foreground=UI["warn"])
        self.lbl_warn.pack(fill="x", pady=(6, 0))

        mid = ttk.PanedWindow(self, orient="horizontal")
        mid.pack(fill="both", expand=True, pady=(6, 0))
        self._pan = mid
        self._table_visible = True
        self._tf = ttk.Frame(mid)
        pf = ttk.Frame(mid)
        mid.add(self._tf, weight=1)
        mid.add(pf, weight=2)
        ct = card(self._tf, "Resultats",
                  "orange : P0 au-dela du domaine mesure  -  rouge : au-dela "
                  "du double")
        ct.pack(fill="both", expand=True)
        tvf = ttk.Frame(ct.body, style="Card.TFrame")
        tvf.pack(fill="both", expand=True)
        self.tv = ttk.Treeview(tvf, show="headings", height=13)
        vs = ttk.Scrollbar(tvf, orient="vertical", command=self.tv.yview)
        hs = ttk.Scrollbar(ct.body, orient="horizontal", command=self.tv.xview)
        self.tv.configure(yscrollcommand=vs.set, xscrollcommand=hs.set)
        self.tv.pack(side="left", fill="both", expand=True)
        vs.pack(side="right", fill="y")
        hs.pack(fill="x")

        cg = card(pf, "Graphiques")
        cg.pack(fill="both", expand=True)
        opt = ttk.Frame(cg.body, style="Card.TFrame")
        opt.pack(fill="x", pady=(0, 6))
        ttk.Label(opt, text="Afficher :", style="Card.TLabel").pack(side="left",
                                                                    padx=(0, 6))
        self.var_pq = tk.BooleanVar(value=True)
        self.var_ps = tk.BooleanVar(value=True)
        self.var_px = tk.BooleanVar(value=False)
        self.var_pp = tk.BooleanVar(value=False)
        for txt, var, hint in (
                ("quantites", self.var_pq, "q_i et q_total"),
                ("selectivites", self.var_ps, "S_i/j"),
                ("fractions x", self.var_px,
                 "composition de la phase adsorbee"),
                ("P0", self.var_pp,
                 "pressions de reference exigees par l'IAST")):
            cb = ttk.Checkbutton(opt, text=txt, variable=var,
                                 command=lambda: self.draw())
            cb.pack(side="left", padx=4)
            tip(cb, hint)
        self.btn_table = ttk.Button(opt, text="Masquer le tableau",
                                    style="Ghost.TButton",
                                    command=self.toggle_table)
        self.btn_table.pack(side="right")
        tip(self.btn_table, "Donne toute la largeur aux graphiques")
        zoom_button(opt, self._render, "IAST", app).pack(side="right", padx=6)

        self.fig = Figure(figsize=(7.2, 6.4), dpi=100, facecolor=UI["card"])
        self.ax1 = None
        self.ax2 = None
        self.canvas = FigureCanvasTkAgg(self.fig, master=cg.body)
        self.canvas.get_tk_widget().pack(fill="both", expand=True)
        if NavigationToolbar2Tk is not None:
            NavigationToolbar2Tk(self.canvas, cg.body).update()
        set_sash(mid, 0, 520)
        self._n_changed()

    def apply_preset(self, spec, label=""):
        """Applique un melange type : noms, fractions et modeles associes."""
        self.var_n.set(len(spec))
        self._n_changed()
        missing = []
        for i, (gas, y) in enumerate(spec):
            if i >= len(self.rows):
                break
            r = self.rows[i]
            r.e_name.delete(0, "end")
            r.e_name.insert(0, gas)
            r.e_y.delete(0, "end")
            r.e_y.insert(0, "%.6g" % y)
            r.refresh_sources()
            # choisir l'ajustement dont le gaz correspond, le plus recent
            cand = [(j, f) for j, f in enumerate(self.app.fits)
                    if str(f.get("gas", "")).upper() == gas.upper()]
            if not cand:
                missing.append(gas)
                continue
            j, f = cand[-1]
            r.cb_src.set("%d. %s" % (j + 1, f["label"]))
            r._src_changed()
        self.app.log("Melange type applique : %s" % (label or spec))
        if missing:
            messagebox.showinfo(
                "Melange type",
                "Fractions molaires configurees.\n\n"
                "Aucun ajustement disponible pour : %s.\n"
                "Ajustez d'abord ces isothermes dans l'onglet Ajustement, "
                "ou saisissez les parametres a la main (bouton 'Editer...')."
                % ", ".join(missing), parent=self)
        else:
            self.app.set_status("Melange %s pret - lancez le calcul IAST."
                                % (label or ""))

    def toggle_table(self):
        """Masque ou reaffiche le tableau pour elargir les graphiques."""
        try:
            if self._table_visible:
                self._pan.forget(self._tf)
                self._table_visible = False
                self.btn_table.config(text="Afficher le tableau")
            else:
                self._pan.insert(0, self._tf, weight=1)
                self._table_visible = True
                self.btn_table.config(text="Masquer le tableau")
                set_sash(self._pan, 0, 520)
        except Exception:
            pass

    def _n_changed(self):
        n = self.var_n.get()
        for i, r in enumerate(self.rows):
            if i < n:
                r.pack(fill="x", pady=1)
            else:
                r.pack_forget()
        self.normalize()

    def refresh(self):
        for r in self.rows:
            r.refresh_sources()

    def normalize(self):
        n = self.var_n.get()
        ys = [self.rows[i].y_value() for i in range(n)]
        ys = [0.0 if (v is None or np.isnan(v) or v < 0) else v for v in ys]
        s = sum(ys)
        if s <= 0:
            ys = [1.0 / n] * n
        else:
            ys = [v / s for v in ys]
        for i in range(n):
            self.rows[i].e_y.delete(0, "end")
            self.rows[i].e_y.insert(0, "%.4f" % ys[i])
        return ys

    def build_components(self):
        n = self.var_n.get()
        comps, names, specs = [], [], []
        warn = []
        for i in range(n):
            r = self.rows[i]
            if not r.spec:
                raise ValueError("Constituant %d : aucun modele defini "
                                 "(choisissez un ajustement ou saisissez les "
                                 "parametres)." % (i + 1))
            m = MODELS[r.spec["model_key"]]
            comps.append(SpreadingPressure(m, r.spec["ctx"], r.spec["params"],
                                           P_exp_max=r.spec.get("P_exp_max")))
            names.append(r.name())
            specs.append(r.spec)
            if m.key in ("dr", "dr_p0", "da", "da_p0"):
                warn.append(m.label)
            if m.key == "freundlich" and r.spec["params"][1] >= 1.0:
                warn.append("Freundlich n>=1")
        self._warn_models = ("Modeles %s : constante de Henry nulle ou non "
                             "definie, IAST discutable a tres basse pression."
                             % ", ".join(sorted(set(warn)))) if warn else ""
        self._no_pexp = [names[i] for i in range(n)
                         if not specs[i].get("P_exp_max")]
        self.lbl_warn.config(text=self._warn_models)
        return comps, names, specs

    def run(self):
        try:
            comps, names, specs = self.build_components()
        except Exception as e:
            messagebox.showerror("IAST", str(e), parent=self)
            return
        ys = self.normalize()
        try:
            npts = max(3, int(to_float_or(self.e_np.get(), 40)))
            if self.cb_mode.get().startswith("Balayage en pression"):
                pmin = to_float_or(self.e_pmin.get(), np.nan)
                pmax = to_float_or(self.e_pmax.get(), np.nan)
                if not (pmin > 0 and pmax > pmin):
                    raise ValueError("Gamme de pression invalide.")
                P = (np.logspace(math.log10(pmin), math.log10(pmax), npts)
                     if self.var_log.get() else np.linspace(pmin, pmax, npts))
                df = iast_scan(P, ys, comps, names)
                self.mode = "P"
            else:
                pfix = to_float_or(self.e_pfix.get(), 1.0)
                y1 = np.linspace(0.001, 0.999, npts)
                rest = None
                if len(comps) == 3:
                    rest = [max(ys[1], 1e-9), max(ys[2], 1e-9)]
                df = iast_scan_composition(pfix, y1, comps, names, rest)
                self.mode = "y"
        except Exception as e:
            messagebox.showerror("IAST", "%s\n\n%s" % (e, traceback.format_exc()),
                                 parent=self)
            return
        self.df = df
        self.names = names
        self.specs = specs
        self.fill_table(df)
        self.draw(df, names)
        nbad = int((~df["Convergence"]).sum())
        self.app.log("IAST : %d points calcules, %d non convergents."
                     % (len(df), nbad))
        self._update_warning(df, names)
        self.app.mark_dirty()

    def _update_warning(self, df, names):
        """Signale l'extrapolation implicite des modeles de corps purs."""
        msgs = []
        if getattr(self, "_warn_models", ""):
            msgs.append(self._warn_models)
        col = "Extrapolation max (P0/P_exp)"
        r = df[col].values if col in df else np.array([np.nan])
        if np.any(np.isfinite(r)):
            rmax = float(np.nanmax(r))
            nb = int(np.sum(r > 2.0))
            worst = {}
            for nm in names:
                c = "P0/P_exp_%s" % nm
                if c in df and np.any(np.isfinite(df[c].values)):
                    worst[nm] = float(np.nanmax(df[c].values))
            det = ", ".join("%s x%.1f" % (k, v) for k, v in
                            sorted(worst.items(), key=lambda kv: -kv[1]))
            if rmax > 2.0:
                msgs.append("EXTRAPOLATION : les P0 requis depassent jusqu'a "
                            "x%.1f la pression mesuree (%s) sur %d point(s). "
                            "Ces points reposent sur l'extrapolation des "
                            "isothermes de corps purs, pas sur des mesures."
                            % (rmax, det, nb))
            elif rmax > 1.0:
                msgs.append("Extrapolation legere : P0 max = x%.1f la pression "
                            "mesuree (%s)." % (rmax, det))
            else:
                msgs.append("Domaine valide : tous les P0 restent dans la "
                            "gamme mesuree (max x%.2f)." % rmax)
        if getattr(self, "_no_pexp", None):
            msgs.append("P experimentale max non renseignee pour : %s "
                        "(extrapolation non controlee)."
                        % ", ".join(self._no_pexp))
        txt = "  |  ".join(msgs)
        self.lbl_warn.config(text=txt[:400])
        col_fg = "#a00000" if "EXTRAPOLATION" in txt else "#a06000"
        try:
            self.lbl_warn.config(foreground=col_fg)
        except Exception:
            pass
        for m in msgs:
            self.app.log("IAST - " + m)

    def fill_table(self, df):
        cols = list(df.columns)
        self.tv["columns"] = cols
        self.tv.delete(*self.tv.get_children())
        for c in cols:
            self.tv.heading(c, text=c)
            self.tv.column(c, width=max(80, min(160, 9 * len(c))), anchor="e")
        self.tv.tag_configure("extrap_fort", background="#ffd6d6")
        self.tv.tag_configure("extrap_leger", background="#fff2cc")
        self.tv.tag_configure("nonconv", background="#dddddd")
        rcol = "Extrapolation max (P0/P_exp)"
        for _, row in df.iterrows():
            tag = ""
            if not bool(row.get("Convergence", True)):
                tag = "nonconv"
            else:
                r = row.get(rcol, np.nan)
                if np.isfinite(r):
                    if r > 2.0:
                        tag = "extrap_fort"
                    elif r > 1.0:
                        tag = "extrap_leger"
            self.tv.insert("", "end", tags=((tag,) if tag else ()),
                           values=[fmt(v, 5) if isinstance(v, (int, float,
                                                               np.floating))
                                   else str(v) for v in row.tolist()])

    def draw(self, df=None, names=None):
        if df is not None:
            self._df_plot, self._names_plot = df, list(names or [])
        self._render(self.fig)
        self.canvas.draw_idle()

    def _render(self, fig):
        """Trace les panneaux demandes sur la figure fournie."""
        fig.clear()
        df = getattr(self, "_df_plot", None)
        names = getattr(self, "_names_plot", None)
        if df is None or not names:
            return
        mode = getattr(self, "mode", "P")
        xcol = "P_total (bar)" if mode == "P" else "y_%s" % names[0]
        if xcol not in df:
            return
        x = df[xcol].values
        panels = [k for k, v in (("q", self.var_pq), ("S", self.var_ps),
                                 ("x", self.var_px), ("P0", self.var_pp))
                  if v.get()]
        if not panels:
            panels = ["q"]
        axes = fig.subplots(len(panels), 1, sharex=True, squeeze=False)[:, 0]

        # zone d'extrapolation
        col = "Extrapolation max (P0/P_exp)"
        x0 = None
        if col in df and mode == "P":
            v = df[col].values
            bad = np.isfinite(v) & (v > 2.0)
            if np.any(bad):
                x0 = float(np.min(x[bad]))

        for ax, kind in zip(axes, panels):
            if x0 is not None:
                ax.axvspan(x0, float(np.max(x)), color="#c00000", alpha=0.07,
                           zorder=0)
                ax.axvline(x0, color="#c00000", ls=":", lw=1.2)
            if kind == "q":
                for i, nm in enumerate(names):
                    c = "q_%s (mmol/g)" % nm
                    if c in df:
                        ax.plot(x, df[c].values, "-o", ms=3.5,
                                color=PALETTE[i % len(PALETTE)], label=nm)
                if "q_total (mmol/g)" in df:
                    ax.plot(x, df["q_total (mmol/g)"].values, "--k", lw=1.2,
                            label="total")
                ax.set_ylabel("q adsorbee (mmol/g)")
            elif kind == "S":
                k = 0
                for i in range(len(names)):
                    for j in range(i + 1, len(names)):
                        c = "S_%s/%s" % (names[i], names[j])
                        if c in df:
                            ax.plot(x, df[c].values, "-s", ms=3.5,
                                    color=PALETTE[k % len(PALETTE)], label=c)
                            k += 1
                ax.set_ylabel("Selectivite")
            elif kind == "x":
                for i, nm in enumerate(names):
                    c = "x_%s" % nm
                    if c in df:
                        ax.plot(x, df[c].values, "-^", ms=3.5,
                                color=PALETTE[i % len(PALETTE)], label=nm)
                ax.set_ylabel("Fraction adsorbee x")
                ax.set_ylim(0, 1)
            else:
                for i, nm in enumerate(names):
                    c = "P0_%s (bar)" % nm
                    if c in df:
                        ax.plot(x, df[c].values, "-v", ms=3.5,
                                color=PALETTE[i % len(PALETTE)],
                                label="P0 %s" % nm)
                    pm = None
                    if getattr(self, "specs", None) and i < len(self.specs):
                        pm = self.specs[i].get("P_exp_max")
                    if pm:
                        ax.axhline(pm, color=PALETTE[i % len(PALETTE)], ls=":",
                                   lw=1)
                ax.set_ylabel("P0 (bar)")
                ax.set_yscale("log")
            ax.grid(True, ls="--", alpha=0.5)
            ax.legend(fontsize=8, ncol=2)
            if mode == "P" and self.var_log.get():
                ax.set_xscale("log")
        axes[-1].set_xlabel(xcol if mode == "P"
                            else "Fraction molaire gazeuse y_%s" % names[0])
        try:
            fig.tight_layout()
        except Exception:
            pass

    def export(self):
        if self.df is None:
            messagebox.showwarning("IAST", "Lancez d'abord un calcul.", parent=self)
            return
        path = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                            filetypes=[("Excel", "*.xlsx")],
                                            initialfile="resultats_IAST.xlsx")
        if not path:
            return
        meta = []
        for i, nm in enumerate(self.names):
            m = MODELS[self.specs[i]["model_key"]]
            d = {"Constituant": nm, "Modele": m.label}
            for j, pn in enumerate(m.pnames):
                d[pn] = self.specs[i]["params"][j]
            d["T (K)"] = self.specs[i]["ctx"].get("T")
            d["P0 (bar)"] = self.specs[i]["ctx"].get("P0")
            meta.append(d)
        with pd.ExcelWriter(path, engine="openpyxl") as w:
            self.df.to_excel(w, sheet_name="IAST", index=False)
            pd.DataFrame(meta).to_excel(w, sheet_name="Modeles", index=False)
        self.app.log("Export IAST : %s" % path)
        messagebox.showinfo("Export", "Fichier enregistre :\n%s" % path,
                            parent=self)

    def send_to_plot(self):
        if self.df is None:
            return
        xcol = "P_total (bar)" if self.mode == "P" else "y_%s" % self.names[0]
        x = self.df[xcol].values
        for nm in self.names:
            self.app.add_series("IAST q %s" % nm, x,
                                self.df["q_%s (mmol/g)" % nm].values, kind="ligne")
        for i in range(len(self.names)):
            for j in range(i + 1, len(self.names)):
                c = "S_%s/%s" % (self.names[i], self.names[j])
                if c in self.df:
                    self.app.add_series("IAST %s" % c, x, self.df[c].values,
                                        kind="ligne")
        self.app.tab_plot.refresh_series()
        self.app.log("Series IAST envoyees vers l'onglet Graphiques.")


# =============================================================================
# 11. ONGLET 5 : GRAPHIQUES PERSONNALISABLES
# =============================================================================

class TabPlot(ttk.Frame):
    def __init__(self, master, app):
        super().__init__(master, padding=10)
        self.app = app
        self.annotations = []
        pan = ttk.PanedWindow(self, orient="horizontal")
        pan.pack(fill="both", expand=True)
        left = ScrollFrame(pan)
        right = ttk.Frame(pan)
        pan.add(left, weight=0)
        pan.add(right, weight=1)
        L = left.inner

        cs = card(L, "Series", "clic sur la case pour afficher ou masquer")
        cs.pack(fill="x", pady=(0, 8))
        self.tv = ttk.Treeview(cs.body, columns=("v", "lab", "typ"),
                               show="headings", height=9, selectmode="browse")
        for c, h, w in (("v", "", 30), ("lab", "Serie", 250),
                        ("typ", "Type", 95)):
            self.tv.heading(c, text=h)
            self.tv.column(c, width=w, anchor="w")
        self.tv.pack(fill="x")
        self.tv.bind("<Button-1>", self._click)
        self.tv.bind("<<TreeviewSelect>>", lambda e: self.load_style())
        bs = ttk.Frame(cs.body, style="Card.TFrame")
        bs.pack(fill="x", pady=(8, 0))
        for txt, cmd, sty in (("Monter", lambda: self.move(-1), ""),
                              ("Descendre", lambda: self.move(1), ""),
                              ("Supprimer", self.del_series, ""),
                              ("Tout vider", self.clear_series, "Danger.TButton")):
            kw = {"style": sty} if sty else {}
            ttk.Button(bs, text=txt, width=11, command=cmd, **kw).pack(
                side="left", padx=2)

        st_ = card(L, "Style de la serie selectionnee")
        st_.pack(fill="x", pady=(0, 8))
        st = st_.body
        st.columnconfigure(1, weight=1)
        self.e_lab = field(st, "Legende :", ttk.Entry, 34, 0, 0,
                           "Texte affiche dans la legende")
        rowc = ttk.Frame(st, style="Card.TFrame")
        ttk.Label(st, text="Type / couleur :", style="Card.TLabel").grid(
            row=1, column=0, sticky="w", pady=3)
        rowc.grid(row=1, column=1, sticky="w")
        self.cb_kind = ttk.Combobox(rowc, width=15, state="readonly",
                                    values=["points", "ligne", "points+ligne"])
        self.cb_kind.pack(side="left")
        self.btn_col = tk.Button(rowc, text="      ", bg="#1f77b4",
                                 relief="flat", bd=0, command=self.pick_color)
        self.btn_col.pack(side="left", padx=6)
        tip(self.btn_col, "Choisir la couleur de la serie")
        rowm = ttk.Frame(st, style="Card.TFrame")
        ttk.Label(st, text="Marqueur / ligne :", style="Card.TLabel").grid(
            row=2, column=0, sticky="w", pady=3)
        rowm.grid(row=2, column=1, sticky="w")
        self.cb_mk = ttk.Combobox(rowm, width=8, state="readonly",
                                  values=MARKERS)
        self.cb_mk.pack(side="left")
        self.cb_ls = ttk.Combobox(rowm, width=8, state="readonly",
                                  values=LINESTYLES)
        self.cb_ls.pack(side="left", padx=6)
        rowe = ttk.Frame(st, style="Card.TFrame")
        ttk.Label(st, text="Trait / point / opacite :",
                  style="Card.TLabel").grid(row=3, column=0, sticky="w", pady=3)
        rowe.grid(row=3, column=1, sticky="w")
        self.e_lw = ttk.Entry(rowe, width=7)
        self.e_lw.pack(side="left")
        tip(self.e_lw, "Epaisseur du trait")
        self.e_ms = ttk.Entry(rowe, width=7)
        self.e_ms.pack(side="left", padx=6)
        tip(self.e_ms, "Taille des marqueurs")
        self.e_al = ttk.Entry(rowe, width=7)
        self.e_al.pack(side="left")
        tip(self.e_al, "Opacite entre 0 et 1")
        self.var_y2 = tk.BooleanVar(value=False)
        ttk.Checkbutton(st, text="Tracer sur l'axe Y secondaire",
                        variable=self.var_y2).grid(row=4, column=0,
                                                   columnspan=2, sticky="w",
                                                   pady=3)
        ttk.Button(st, text="Appliquer le style", style="Primary.TButton",
                   command=self.apply_style).grid(row=5, column=0, columnspan=2,
                                                  sticky="ew", pady=(8, 0))

        ca = card(L, "Axes et mise en forme")
        ca.pack(fill="x", pady=(0, 8))
        ax = ca.body
        self.v_title = tk.StringVar(value="")
        self.v_xlab = tk.StringVar(value="Pression (bar)")
        self.v_ylab = tk.StringVar(value="Quantite adsorbee (mmol/g)")
        self.v_y2lab = tk.StringVar(value="Selectivite")
        for i, (lab, var) in enumerate((("Titre :", self.v_title),
                                        ("Axe X :", self.v_xlab),
                                        ("Axe Y :", self.v_ylab),
                                        ("Axe Y2 :", self.v_y2lab))):
            ttk.Label(ax, text=lab, style="Card.TLabel").grid(row=i, column=0,
                                                              sticky="w", pady=2)
            ttk.Entry(ax, textvariable=var, width=32).grid(row=i, column=1,
                                                           columnspan=3,
                                                           sticky="ew", padx=6)
        self.v_xmin = tk.StringVar()
        self.v_xmax = tk.StringVar()
        self.v_ymin = tk.StringVar()
        self.v_ymax = tk.StringVar()
        ttk.Label(ax, text="X min / max :", style="Card.TLabel").grid(
            row=4, column=0, sticky="w", pady=2)
        ttk.Entry(ax, textvariable=self.v_xmin, width=11).grid(row=4, column=1,
                                                               sticky="w", padx=6)
        ttk.Entry(ax, textvariable=self.v_xmax, width=11).grid(row=4, column=2,
                                                               sticky="w")
        ttk.Label(ax, text="Y min / max :", style="Card.TLabel").grid(
            row=5, column=0, sticky="w", pady=2)
        ttk.Entry(ax, textvariable=self.v_ymin, width=11).grid(row=5, column=1,
                                                               sticky="w", padx=6)
        ttk.Entry(ax, textvariable=self.v_ymax, width=11).grid(row=5, column=2,
                                                               sticky="w")
        self.v_logx = tk.BooleanVar(value=False)
        self.v_logy = tk.BooleanVar(value=False)
        self.v_grid = tk.BooleanVar(value=True)
        self.v_leg = tk.BooleanVar(value=True)
        box = ttk.Frame(ax, style="Card.TFrame")
        box.grid(row=6, column=0, columnspan=4, sticky="w", pady=3)
        for txt, var in (("Log X", self.v_logx), ("Log Y", self.v_logy),
                         ("Grille", self.v_grid), ("Legende", self.v_leg)):
            ttk.Checkbutton(box, text=txt, variable=var).pack(side="left",
                                                              padx=(0, 12))
        self.cb_loc = field(ax, "Position legende :", ttk.Combobox, 14, 7, 0,
                            None, values=LEGEND_LOCS, state="readonly")
        self.cb_loc.set("best")
        self.v_fs = tk.StringVar(value="11")
        self.v_w = tk.StringVar(value="8.0")
        self.v_h = tk.StringVar(value="6.0")
        self.v_dpi = tk.StringVar(value="300")
        ttk.Label(ax, text="Police :", style="Card.TLabel").grid(row=8, column=0,
                                                                 sticky="w",
                                                                 pady=2)
        ttk.Entry(ax, textvariable=self.v_fs, width=7).grid(row=8, column=1,
                                                            sticky="w", padx=6)
        ttk.Label(ax, text="Export : largeur x hauteur (pouces) :",
                  style="Card.TLabel").grid(row=9, column=0, sticky="w", pady=2)
        ttk.Entry(ax, textvariable=self.v_w, width=7).grid(row=9, column=1,
                                                           sticky="w", padx=6)
        ttk.Entry(ax, textvariable=self.v_h, width=7).grid(row=9, column=2,
                                                           sticky="w")
        ttk.Label(ax, text="DPI export :", style="Card.TLabel").grid(
            row=10, column=0, sticky="w", pady=2)
        ttk.Entry(ax, textvariable=self.v_dpi, width=7).grid(row=10, column=1,
                                                             sticky="w", padx=6)

        an = card(L, "Annotations",
                  "x et y vides : placement automatique en bas a droite")
        an.pack(fill="x", pady=(0, 8))
        self.tbl_ann = EditableTable(an.body, ["Texte", "x", "y", "Taille"],
                                     height=4, widths=[150, 70, 70, 60])
        self.tbl_ann.pack(fill="x")
        ab = ttk.Frame(an.body, style="Card.TFrame")
        ab.pack(fill="x", pady=(6, 0))
        ttk.Button(ab, text="+ Annotation", width=14,
                   command=lambda: self.tbl_ann.add_row(["Texte", "", "", "10"])
                   ).pack(side="left", padx=2)
        ttk.Button(ab, text="- Annotation", width=14,
                   command=self.tbl_ann.delete_selected).pack(side="left", padx=2)

        bt = ttk.Frame(L, style="Card.TFrame", padding=12)
        bt.pack(fill="x")
        tip(ttk.Button(bt, text="Actualiser le graphique",
                       style="Primary.TButton", command=self.draw),
            "Applique tous les reglages ci-dessus (F5)").pack(fill="x")
        zoom_button(bt, self._plot_on, "Graphique", app).pack(fill="x",
                                                              pady=(6, 0))
        tip(ttk.Button(bt, text="Exporter l'image...", style="Success.TButton",
                       command=self.export_image),
            "PNG, PDF, SVG, EPS, TIFF ou JPEG a la resolution choisie").pack(
            fill="x", pady=6)
        ttk.Button(bt, text="Exporter les donnees tracees...",
                   command=self.export_data).pack(fill="x")

        cf = card(right, "Figure",
                  "molette et barre d'outils pour zoomer  -  bouton Agrandir "
                  "pour une fenetre dediee")
        cf.pack(fill="both", expand=True)
        self.fig = Figure(figsize=(8, 6), dpi=100, facecolor=UI["card"])
        self.ax = self.fig.add_subplot(111)
        self.ax2 = None
        self.canvas = FigureCanvasTkAgg(self.fig, master=cf.body)
        self.canvas.get_tk_widget().pack(fill="both", expand=True)
        if NavigationToolbar2Tk is not None:
            NavigationToolbar2Tk(self.canvas, cf.body).update()
        set_sash(pan, 0, 470)

    def refresh_series(self):
        sel = self.tv.selection()
        self.tv.delete(*self.tv.get_children())
        for i, s in enumerate(self.app.series):
            self.tv.insert("", "end", iid=str(i),
                           values=("X" if s["visible"] else "", s["label"],
                                   s["kind"]))
        if sel and sel[0] in self.tv.get_children():
            self.tv.selection_set(sel[0])
        self.draw()
        try:
            self.app.mark_dirty()
        except Exception:
            pass

    def _click(self, event):
        if self.tv.identify_region(event.x, event.y) != "cell":
            return
        if self.tv.identify_column(event.x) != "#1":
            return
        iid = self.tv.identify_row(event.y)
        if not iid:
            return
        s = self.app.series[int(iid)]
        s["visible"] = not s["visible"]
        self.tv.set(iid, "v", "X" if s["visible"] else "")
        self.draw()
        return "break"

    def _sel(self):
        sel = self.tv.selection()
        return int(sel[0]) if sel else None

    def load_style(self):
        i = self._sel()
        if i is None:
            return
        s = self.app.series[i]
        self.e_lab.delete(0, "end")
        self.e_lab.insert(0, s["label"])
        self.cb_kind.set(s["kind"])
        self.btn_col.config(bg=s["color"])
        self.cb_mk.set(s["marker"])
        self.cb_ls.set(s["ls"])
        for e, v in ((self.e_lw, s["lw"]), (self.e_ms, s["ms"]),
                     (self.e_al, s["alpha"])):
            e.delete(0, "end")
            e.insert(0, str(v))
        self.var_y2.set(s.get("y2", False))

    def pick_color(self):
        c = colorchooser.askcolor(color=self.btn_col.cget("bg"), parent=self)
        if c and c[1]:
            self.btn_col.config(bg=c[1])

    def apply_style(self):
        i = self._sel()
        if i is None:
            return
        s = self.app.series[i]
        s["label"] = self.e_lab.get()
        s["kind"] = self.cb_kind.get()
        s["color"] = self.btn_col.cget("bg")
        s["marker"] = self.cb_mk.get()
        s["ls"] = self.cb_ls.get()
        s["lw"] = to_float_or(self.e_lw.get(), 1.5)
        s["ms"] = to_float_or(self.e_ms.get(), 6.0)
        s["alpha"] = min(max(to_float_or(self.e_al.get(), 1.0), 0.05), 1.0)
        s["y2"] = bool(self.var_y2.get())
        self.refresh_series()

    def del_series(self):
        i = self._sel()
        if i is None:
            return
        self.app.series.pop(i)
        self.refresh_series()

    def clear_series(self):
        self.app.series = []
        self.refresh_series()

    def move(self, d):
        i = self._sel()
        if i is None:
            return
        j = i + d
        if 0 <= j < len(self.app.series):
            self.app.series[i], self.app.series[j] = (self.app.series[j],
                                                      self.app.series[i])
            self.refresh_series()
            self.tv.selection_set(str(j))

    # ------------------------------------------------------------------
    def _style_axes(self, ax, ax2):
        fs = to_float_or(self.v_fs.get(), 11)
        ax.set_xlabel(self.v_xlab.get(), fontsize=fs)
        ax.set_ylabel(self.v_ylab.get(), fontsize=fs)
        if self.v_title.get():
            ax.set_title(self.v_title.get(), fontsize=fs + 1)
        ax.tick_params(labelsize=fs - 1)
        if self.v_logx.get():
            ax.set_xscale("log")
        if self.v_logy.get():
            ax.set_yscale("log")
        lims = [to_float(v.get()) for v in (self.v_xmin, self.v_xmax,
                                            self.v_ymin, self.v_ymax)]
        if np.isfinite(lims[0]) or np.isfinite(lims[1]):
            cur = ax.get_xlim()
            ax.set_xlim(lims[0] if np.isfinite(lims[0]) else cur[0],
                        lims[1] if np.isfinite(lims[1]) else cur[1])
        if np.isfinite(lims[2]) or np.isfinite(lims[3]):
            cur = ax.get_ylim()
            ax.set_ylim(lims[2] if np.isfinite(lims[2]) else cur[0],
                        lims[3] if np.isfinite(lims[3]) else cur[1])
        ax.grid(self.v_grid.get(), ls="--", alpha=0.5)
        if ax2 is not None:
            ax2.set_ylabel(self.v_y2lab.get(), fontsize=fs)
            ax2.tick_params(labelsize=fs - 1)
        if self.v_leg.get():
            h, l = ax.get_legend_handles_labels()
            if ax2 is not None:
                h2, l2 = ax2.get_legend_handles_labels()
                h, l = h + h2, l + l2
            if h:
                ax.legend(h, l, loc=(self.cb_loc.get() or "best"),
                          fontsize=max(fs - 2, 1))
        for row in self.tbl_ann.get_rows():
            txt = row[0].strip()
            if not txt:
                continue
            x, y = to_float(row[1]), to_float(row[2])
            sz = to_float_or(row[3], 10)
            if np.isfinite(x) and np.isfinite(y):
                ax.text(x, y, txt, fontsize=sz,
                        bbox=dict(boxstyle="round", fc="#FFFACD", alpha=0.7))
            else:
                ax.text(0.97, 0.05, txt, transform=ax.transAxes, fontsize=sz,
                        ha="right", va="bottom", multialignment="left",
                        bbox=dict(boxstyle="round", fc="#FFFACD", alpha=0.7))

    def _plot_on(self, fig):
        fig.clear()
        ax = fig.add_subplot(111)
        ax2 = None
        if any(s.get("y2") and s["visible"] for s in self.app.series):
            ax2 = ax.twinx()
        for s in self.app.series:
            if not s["visible"]:
                continue
            tgt = ax2 if s.get("y2") else ax
            mk = s["marker"] if s["kind"] in ("points", "points+ligne") else "None"
            ls = s["ls"] if s["kind"] in ("ligne", "points+ligne") else "None"
            tgt.plot(s["x"], s["y"], marker=(None if mk == "None" else mk),
                     linestyle=("None" if ls == "None" else ls),
                     color=s["color"], lw=s["lw"], ms=s["ms"],
                     alpha=s["alpha"], label=s["label"])
        self._style_axes(ax, ax2)
        try:
            fig.tight_layout()
        except Exception:
            pass
        return ax, ax2

    def _figsize(self):
        w = min(max(to_float_or(self.v_w.get(), 8.0), 2.0), 40.0)
        h = min(max(to_float_or(self.v_h.get(), 6.0), 2.0), 40.0)
        return (w, h)

    def draw(self):
        # La taille en pouces ne s'applique qu'a l'export : imposer une taille
        # a la figure affichee entre en conflit avec le gestionnaire de
        # placement Tk et corrompt le rendu.
        self._plot_on(self.fig)
        self.canvas.draw_idle()

    def export_image(self):
        path = filedialog.asksaveasfilename(
            defaultextension=".png", initialfile="figure.png",
            filetypes=[("PNG", "*.png"), ("PDF vectoriel", "*.pdf"),
                       ("SVG vectoriel", "*.svg"), ("EPS", "*.eps"),
                       ("TIFF", "*.tif"), ("JPEG", "*.jpg")])
        if not path:
            return
        dpi = to_float_or(self.v_dpi.get(), 300)
        fig = Figure(figsize=self._figsize(), dpi=dpi)
        self._plot_on(fig)
        try:
            fig.savefig(path, dpi=dpi, bbox_inches="tight",
                        facecolor="white")
        except Exception as e:
            messagebox.showerror("Export", str(e), parent=self)
            return
        self.app.log("Figure exportee : %s" % path)
        messagebox.showinfo("Export", "Figure enregistree :\n%s" % path,
                            parent=self)

    def export_data(self):
        vis = [s for s in self.app.series if s["visible"]]
        if not vis:
            messagebox.showwarning("Export", "Aucune serie affichee.", parent=self)
            return
        path = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                            initialfile="donnees_graphique.xlsx",
                                            filetypes=[("Excel", "*.xlsx"),
                                                       ("CSV", "*.csv")])
        if not path:
            return
        n = max(len(s["x"]) for s in vis)
        data = {}
        for s in vis:
            x = list(s["x"]) + [np.nan] * (n - len(s["x"]))
            y = list(s["y"]) + [np.nan] * (n - len(s["y"]))
            data["%s | x" % s["label"]] = x
            data["%s | y" % s["label"]] = y
        df = pd.DataFrame(data)
        if path.lower().endswith(".csv"):
            df.to_csv(path, index=False)
        else:
            df.to_excel(path, index=False, sheet_name="Graphique")
        self.app.log("Donnees du graphique exportees : %s" % path)
        messagebox.showinfo("Export", "Fichier enregistre :\n%s" % path,
                            parent=self)


# =============================================================================
# 12. EXPORTS GLOBAUX ET PROJET
# =============================================================================

def fits_summary_frame(app):
    rows = []
    for f in app.fits:
        m = MODELS[f["model_key"]]
        d = {"Jeu de donnees": f["dataset"], "Modele": m.label,
             "Equation": m.formula, "Ponderation": f.get("weights", "")}
        for j, pn in enumerate(m.pnames):
            d["%s (%s)" % (pn, m.punits[j])] = f["params"][j]
            d["sigma(%s)" % pn] = f["errors"][j]
            if f.get("ci_lo"):
                d["%s IC 2.5%%" % pn] = f["ci_lo"][j]
                d["%s IC 97.5%%" % pn] = f["ci_hi"][j]
        for k, (v, u) in m.derived(f["params"], f["ctx"]).items():
            d["%s [%s]" % (k, u)] = v
        d["T (K)"] = f["ctx"].get("T")
        d["P0 (bar)"] = f["ctx"].get("P0")
        d["|r| max (correlation)"] = f.get("r_max", np.nan)
        d["Departs"] = f.get("n_starts", 1)
        d["Tirages bootstrap"] = f.get("n_boot", 0)
        d["P exp. max (bar)"] = f.get("P_exp_max", np.nan)
        d.update({("R2" if k == "R2" else k): v for k, v in f["stats"].items()})
        rows.append(d)
    return pd.DataFrame(rows)


def export_fits_excel(app, path):
    with pd.ExcelWriter(path, engine="openpyxl") as w:
        fits_summary_frame(app).to_excel(w, sheet_name="Parametres", index=False)
        # courbes ajustees
        curves = {}
        for f in app.fits:
            m = MODELS[f["model_key"]]
            P = np.asarray(f["P"], float)
            x = np.linspace(max(1e-6, P.min() * 0.01), P.max() * 1.1, 300)
            y = m.func(x, f["ctx"], *f["params"])
            key = "%s - %s" % (f["dataset"], m.label)
            curves["%s | P (bar)" % key] = x
            curves["%s | q (mmol/g)" % key] = np.asarray(y).ravel()
        if curves:
            pd.DataFrame(curves).to_excel(w, sheet_name="Courbes ajustees",
                                          index=False)
        res = {}
        for f in app.fits:
            m = MODELS[f["model_key"]]
            key = "%s - %s" % (f["dataset"], m.label)
            n = len(f["P"])
            res["%s | P" % key] = f["P"]
            res["%s | q exp" % key] = f["q"]
            res["%s | q calc" % key] = f["qpred"]
            res["%s | residu" % key] = f["residuals"]
        if res:
            n = max(len(v) for v in res.values())
            for k in res:
                res[k] = list(res[k]) + [np.nan] * (n - len(res[k]))
            pd.DataFrame(res).to_excel(w, sheet_name="Residus", index=False)


def export_all_excel(app, path):
    with pd.ExcelWriter(path, engine="openpyxl") as w:
        # 1. donnees brutes
        rows = []
        for d in app.datasets:
            for i in range(len(d["P"])):
                rows.append({"Jeu": d["name"], "Gaz": d["gas"], "T (K)": d["T"],
                             "Type": d["kind"], "P (bar)": d["P"][i],
                             "q (mmol/g)": d["q"][i],
                             "P (%s)" % d["unit_p"]: d["P_raw"][i],
                             "q (%s)" % d["unit_q"]: d["q_raw"][i],
                             "Commentaire": d.get("note", "")})
        pd.DataFrame(rows).to_excel(w, sheet_name="Donnees", index=False)

        # 2. ajustements
        if app.fits:
            fits_summary_frame(app).to_excel(w, sheet_name="Ajustements",
                                             index=False)
            curves = {}
            for f in app.fits:
                m = MODELS[f["model_key"]]
                P = np.asarray(f["P"], float)
                x = np.linspace(max(1e-6, P.min() * 0.01), P.max() * 1.1, 300)
                y = np.asarray(m.func(x, f["ctx"], *f["params"])).ravel()
                key = "%s - %s" % (f["dataset"], m.label)
                curves["%s | P (bar)" % key] = x
                curves["%s | q (mmol/g)" % key] = y
            pd.DataFrame(curves).to_excel(w, sheet_name="Courbes ajustees",
                                          index=False)

        # 3. IAST
        df = getattr(app.tab_iast, "df", None)
        if df is not None:
            df.to_excel(w, sheet_name="IAST", index=False)

        # 4. series tracees
        vis = [s for s in app.series]
        if vis:
            n = max(len(s["x"]) for s in vis)
            data = {}
            for s in vis:
                data["%s | x" % s["label"]] = (list(s["x"])
                                               + [np.nan] * (n - len(s["x"])))
                data["%s | y" % s["label"]] = (list(s["y"])
                                               + [np.nan] * (n - len(s["y"])))
            pd.DataFrame(data).to_excel(w, sheet_name="Graphique", index=False)

        # 5. informations
        pd.DataFrame([{"Logiciel": "%s v%s" % (APP_NAME, APP_VERSION),
                       "Date": datetime.datetime.now().strftime("%Y-%m-%d %H:%M"),
                       "Jeux de donnees": len(app.datasets),
                       "Ajustements": len(app.fits)}]
                     ).to_excel(w, sheet_name="Informations", index=False)



def export_origin(app, path, npts=400):
    """Classeur au format attendu par Origin :
       ligne 1 = Long Name, ligne 2 = Units, donnees a partir de la ligne 3.
    Les series sont mises en colonnes cote a cote (format 'large'), ce qui est
    la disposition directement exploitable par Origin (une colonne = une
    courbe), contrairement au format 'long' d'un export classique.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter

    H1 = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    H2 = Font(name="Arial", italic=True, size=9, color="404040")
    BD = Font(name="Arial", size=10)
    FH = PatternFill("solid", fgColor="2F4F6F")
    FU = PatternFill("solid", fgColor="E8EDF2")

    def block(ws, cols, units):
        """cols : liste (nom, sequence). units : liste de chaines."""
        nmax = max((len(v) for _n, v in cols), default=0)
        for j, ((name, vals), un) in enumerate(zip(cols, units), start=1):
            c = ws.cell(row=1, column=j, value=name)
            c.font, c.fill = H1, FH
            c.alignment = Alignment(horizontal="center", wrap_text=True)
            u = ws.cell(row=2, column=j, value=un)
            u.font, u.fill = H2, FU
            u.alignment = Alignment(horizontal="center")
            for i in range(nmax):
                v = vals[i] if i < len(vals) else None
                if v is not None:
                    try:
                        v = float(v)
                        if not np.isfinite(v):
                            v = None
                    except Exception:
                        pass
                cc = ws.cell(row=3 + i, column=j, value=v)
                cc.font = BD
                if isinstance(v, float):
                    cc.number_format = "0.######"
            ws.column_dimensions[get_column_letter(j)].width = \
                max(11, min(22, len(name) + 3))
        ws.freeze_panes = "A3"

    wb = Workbook()

    # ---------------------------------------------------- notice
    ws = wb.active
    ws.title = "LISEZMOI"
    notes = [
        "%s %s - export Origin" % (APP_NAME, APP_VERSION),
        datetime.datetime.now().strftime("genere le %d/%m/%Y a %H:%M"),
        "",
        "STRUCTURE DES FEUILLES",
        "  ligne 1 : Long Name (nom de colonne Origin)",
        "  ligne 2 : Units (unite)",
        "  ligne 3 et suivantes : donnees",
        "",
        "IMPORT DANS ORIGIN",
        "  File > Import > Excel...  puis dans les options d'import :",
        "     Main Header Lines = 0",
        "     Long Names        = 1",
        "     Units             = 2",
        "     Data begins on    = 3",
        "  Chaque colonne devient une colonne Origin nommee et unitee ;",
        "  il ne reste qu'a designer les colonnes X et Y.",
        "",
        "FEUILLES",
        "  Isothermes      : points experimentaux, un bloc de colonnes par jeu",
        "  Courbes_modeles : courbes lissees des ajustements (%d points)" % npts,
        "  Parametres      : parametres et statistiques de chaque ajustement",
        "  IAST            : dernier calcul IAST effectue",
        "",
        "Les colonnes vides en bas d'un bloc sont normales : les jeux n'ont pas",
        "tous le meme nombre de points.",
    ]
    for i, t in enumerate(notes, start=1):
        c = ws.cell(row=i, column=1, value=t)
        c.font = Font(name="Arial", size=10,
                      bold=t.isupper() and len(t) > 3)
    ws.column_dimensions["A"].width = 88

    # ---------------------------------------------------- isothermes
    ws = wb.create_sheet("Isothermes")
    cols, units = [], []
    for d in app.datasets:
        tag = d["name"].replace(" ", "_")
        cols.append(("P_%s" % tag, list(np.asarray(d["P"], float))))
        units.append("bar")
        cols.append(("q_%s" % tag, list(np.asarray(d["q"], float))))
        units.append("mmol/g")
    if cols:
        block(ws, cols, units)
    else:
        ws.cell(row=1, column=1, value="aucun jeu de donnees")

    # ---------------------------------------------------- courbes de modeles
    ws = wb.create_sheet("Courbes_modeles")
    cols, units = [], []
    for f in app.fits:
        m = MODELS[f["model_key"]]
        P = np.asarray(f["P"], float)
        if P.size < 2:
            continue
        grid = np.linspace(max(float(np.min(P)), 1e-4), float(np.max(P)), npts)
        try:
            q = np.asarray(m.func(grid, f["ctx"], *f["params"]),
                           dtype=float).ravel()
        except Exception:
            continue
        tag = f["label"].replace(" ", "_").replace("-", "_")
        cols.append(("P_%s" % tag, list(grid)))
        units.append("bar")
        cols.append(("q_%s" % tag, list(q)))
        units.append("mmol/g")
    if cols:
        block(ws, cols, units)
    else:
        ws.cell(row=1, column=1, value="aucun ajustement")

    # ---------------------------------------------------- parametres
    ws = wb.create_sheet("Parametres")
    rows = []
    for f in app.fits:
        m = MODELS[f["model_key"]]
        r = {"Jeu": f["dataset"], "Gaz": f.get("gas", ""), "Modele": m.label}
        for i in range(m.nparam):
            r["p%d_nom" % (i + 1)] = m.pnames[i]
            r["p%d" % (i + 1)] = f["params"][i]
            r["p%d_err" % (i + 1)] = f["errors"][i]
            r["p%d_unite" % (i + 1)] = m.punits[i]
        st = f["stats"]
        r.update({"N": st["N"], "R2": st["R2"], "R2_ajuste": st["R2adj"],
                  "RMSE": st["RMSE"], "AICc": st["AICc"], "BIC": st["BIC"],
                  "T": f["ctx"].get("T"), "P0": f["ctx"].get("P0")})
        rows.append(r)
    if rows:
        keys = []
        for r in rows:
            for k in r:
                if k not in keys:
                    keys.append(k)
        U = {"RMSE": "mmol/g", "T": "K", "P0": "bar"}
        block(ws, [(k, [r.get(k) for r in rows]) for k in keys],
              [U.get(k, "") for k in keys])
    else:
        ws.cell(row=1, column=1, value="aucun ajustement")

    # ---------------------------------------------------- IAST
    ws = wb.create_sheet("IAST")
    df = getattr(app.tab_iast, "df", None)
    if df is not None and len(df):
        block(ws, [(str(c), list(df[c].values)) for c in df.columns],
              ["bar" if str(c).lower().startswith("p")
               else ("mmol/g" if str(c).lower().startswith(("n", "q")) else "")
               for c in df.columns])
    else:
        ws.cell(row=1, column=1,
                value="aucun calcul IAST - lancez d'abord l'onglet IAST")

    wb.save(path)
    return path


def save_project(app, path):
    data = dict(version=APP_VERSION, datasets=app.datasets,
                fits=app.fits,
                series=[{k: (list(map(float, v)) if k in ("x", "y") else v)
                         for k, v in s.items()} for s in app.series])
    with open(path, "w", encoding="utf-8") as fh:
        json.dump(data, fh, indent=1, ensure_ascii=False, default=float)


def load_project(app, path):
    with open(path, "r", encoding="utf-8") as fh:
        data = json.load(fh)
    app.datasets = data.get("datasets", [])
    app.fits = data.get("fits", [])
    app.series = [{k: (np.asarray(v, float) if k in ("x", "y") else v)
                   for k, v in s.items()} for s in data.get("series", [])]


HELP_TEXT = """\
%s v%s - aide rapide
================================================================================

1) ONGLET DONNEES
   - "Nouveau" cree un jeu vide ; tapez les valeurs ou collez deux colonnes
     depuis Excel avec Ctrl+V (selectionnez d'abord la premiere ligne cible).
   - "Importer fichier..." lit un .xlsx / .xls / .csv et laisse choisir la
     feuille, les colonnes et les unites.
   - Les unites sont converties en interne en bar et mmol/g.
   - Cliquez toujours sur "Enregistrer le jeu" apres une modification.

2) ONGLET CONVERSION  (exces -> total)
   - Volume poreux    : n_tot = n_exc + V_pore x rho_gaz(T,P)
     rho_gaz est calculee par Peng-Robinson (ou gaz parfait, ou CoolProp si
     la bibliotheque est installee).
   - Densite adsorbee : n_tot = n_exc / (1 - rho_gaz/rho_ads)
   - Densite fournie  : vous saisissez vous-meme rho_gaz (colonne 2 du tableau)
     par exemple depuis NIST REFPROP.
   - "Creer un jeu de donnees 'total'" ajoute le resultat comme nouveau jeu,
     directement utilisable pour l'ajustement.

3) ONGLET AJUSTEMENT
   - Modeles : Langmuir, Langmuir bi-site, Freundlich, Sips, Toth,
     Dubinin-Radushkevich et Dubinin-Astakhov (P0 fixe ou ajuste).
   - P0 peut etre estimee automatiquement : correlation de Lee-Kesler si
     T < Tc, regle de Dubinin Pc(T/Tc)^2 si T > Tc.
   - Les valeurs initiales et les bornes sont modifiables dans le tableau.
   - Ponderation : moindres carres, 1/q ou 1/q^2 (erreur relative).
   - Multi-depart : relance l'ajustement depuis plusieurs points de depart
     aleatoires et conserve le meilleur. Utile pour Toth, Sips et D-A qui
     possedent des minima locaux.
   - Bootstrap : reajuste le modele N fois sur des residus reechantillonnes et
     fournit un intervalle de confiance a 95 %% par parametre. Plus honnete que
     l'ecart-type asymptotique de la matrice de covariance.
   - Colonne |r|max : plus forte correlation entre deux parametres. Au-dela de
     0,99 (ligne rouge), q_max et b se compensent : leurs valeurs individuelles
     ne sont pas identifiables, meme avec un R2 de 0,9999. La capacite
     q_max issue d'un ajustement de Sips ou Toth doit alors etre citee avec
     son intervalle de confiance, pas comme une valeur exacte.
   - "Classement (AIC)" compare les modeles en penalisant le nombre de
     parametres ; dAIC < 2 signifie que deux modeles sont equivalents.

4) ONGLET IAST
   - Chaque constituant recoit un modele : soit un ajustement realise dans
     l'onglet precedent, soit une saisie manuelle des parametres.
   - La pression d'etalement reduite pi* = integrale de q/P dP est analytique
     pour Langmuir, Langmuir bi-site, Freundlich, Sips, D-R et D-A, et
     numerique pour Toth. Le solveur travaille a une seule inconnue (pi*),
     ce qui est nettement plus robuste qu'un systeme a N equations.
   - Deux modes : balayage en pression totale, ou balayage en composition a
     pression fixee.
   - Sorties : P0_i, x_i, q_i, q_total, selectivites S_i/j, et la quantite
     adsorbee du corps pur a la meme pression partielle (reference).
   - Avertissement : les modeles D-R et D-A ont une constante de Henry nulle ;
     l'IAST y est thermodynamiquement discutable a tres basse pression.
   - CONTROLE D'EXTRAPOLATION : pour calculer un melange a P_totale, l'IAST a
     besoin des isothermes de corps purs jusqu'a P0_i, souvent bien au-dela de
     la pression mesuree (typiquement x5 pour un constituant peu adsorbe comme
     H2 a 60 bar). Les colonnes "P0/P_exp_i" et "Extrapolation max" donnent ce
     rapport ; les lignes concernees sont surlignees (orange au-dela de 1,
     rouge au-dela de 2) et la zone correspondante est grisee sur le graphique.
     Au-dela de 2, le resultat depend surtout du modele choisi, pas des mesures.

5) ONGLET GRAPHIQUES
   - Toute serie envoyee depuis les onglets Ajustement ou IAST est listee ici.
   - Couleur, marqueur, style de ligne, epaisseur, opacite, axe Y secondaire,
     bornes, echelles log, legende, annotations : tout est modifiable.
   - Export PNG / PDF / SVG / EPS / TIFF / JPEG a la resolution choisie, et
     export Excel des donnees effectivement tracees.

6) ONGLET SESSION / EXPORT
   - "Enregistrer la session" ecrit UN SEUL fichier .adsp qui contient tout :
     jeux de donnees, ajustements et leurs statistiques, constituants et
     resultats IAST, series et mise en forme des graphiques, annotations.
     Il suffit de le rouvrir pour retrouver l'integralite du travail, y compris
     sur un autre ordinateur. Le format est du JSON lisible.
   - Une copie .adsp.bak de la version precedente est conservee a chaque
     enregistrement, et le programme propose d'enregistrer avant de quitter si
     des modifications sont en cours.
   - "Verification interne" rejoue les controles numeriques du programme
     (modeles, pression d'etalement, IAST, equation d'etat) et affiche un
     rapport. Equivalent en ligne de commande : AdsorpSuite.exe --selftest
   - Export Excel global : donnees, ajustements, courbes, IAST, graphique.

GRAPHIQUES ET LISIBILITE
   - Chaque graphique possede un bouton "Agrandir le graphique" : il s'ouvre
     dans une fenetre separee et redimensionnable, avec zoom, deplacement et
     export haute resolution.
   - Onglet IAST : cases "quantites / selectivites / fractions x / P0" pour
     choisir les panneaux traces, et bouton "Masquer le tableau" pour donner
     toute la largeur aux courbes.
   - Les separateurs entre panneaux se deplacent a la souris ; les panneaux
     lateraux defilent a la molette et affichent une barre horizontale si la
     fenetre est trop etroite.
   - La taille en pouces reglee dans l'onglet Graphiques ne concerne que
     l'export : la figure affichee s'adapte a la fenetre.

RACCOURCIS CLAVIER
   Ctrl+N nouvelle session      Ctrl+O ouvrir       Ctrl+S enregistrer
   Ctrl+Maj+S enregistrer sous  Ctrl+1..6 onglets   F1 aide
   F5 lance l'action principale de l'onglet courant
   Dans les tableaux : double-clic pour editer, Ctrl+V coller, Suppr effacer

RAPPELS D'UNITES
   P : bar    q : mmol/g    V_pore : cm3/g    rho : mmol/cm3
   E (D-A) : J/mol          K (D-R) : mol2/J2       E = 1/sqrt(K)
""" % (APP_NAME, APP_VERSION)


def models_text():
    lines = ["MODELES DISPONIBLES", "=" * 78, ""]
    for k in MODEL_ORDER:
        m = MODELS[k]
        lines.append("%s" % m.label)
        lines.append("   Equation   : %s" % m.formula)
        lines.append("   Parametres : %s" % ", ".join(
            "%s [%s]" % (p, u) for p, u in zip(m.pnames, m.punits)))
        lines.append("   pi* (IAST) : %s"
                     % ("analytique" if m.pi else "integration numerique"))
        if m.needs:
            lines.append("   Contexte   : %s" % ", ".join(m.needs))
        if m.note:
            lines.append("   Remarque   : %s" % m.note)
        lines.append("")
    return "\n".join(lines)


# =============================================================================
# 13. ONGLET 6 : EXPORT ET PROJET
# =============================================================================

class TabExport(ttk.Frame):
    def __init__(self, master, app):
        super().__init__(master, padding=10)
        self.app = app
        top = ttk.Frame(self)
        top.pack(fill="x")

        c1 = card(top, "Session de travail",
                  "un seul fichier contient tout : donnees, ajustements, IAST, "
                  "graphiques")
        c1.pack(side="left", fill="both", expand=True)
        g = c1.body
        tip(ttk.Button(g, text="Enregistrer la session...",
                       style="Primary.TButton", command=self.save),
            "Ecrit un fichier .adsp contenant l'integralite du travail "
            "(Ctrl+S)").pack(fill="x", pady=3)
        tip(ttk.Button(g, text="Enregistrer sous un nouveau nom...",
                       command=lambda: self.save(as_new=True)),
            "Utile pour garder des versions successives").pack(fill="x", pady=3)
        tip(ttk.Button(g, text="Ouvrir une session...", style="Success.TButton",
                       command=self.open),
            "Recharge integralement une session enregistree (Ctrl+O)").pack(
            fill="x", pady=3)
        ttk.Label(g, style="Muted.TLabel", justify="left", wraplength=430,
                  text="Le fichier .adsp est du texte JSON : il se copie d'un "
                       "PC a l'autre, se sauvegarde et reste lisible. Une copie "
                       ".adsp.bak de la version precedente est conservee a "
                       "chaque enregistrement.").pack(anchor="w", pady=(8, 0))

        c2 = card(top, "Exports")
        c2.pack(side="left", fill="both", expand=True, padx=(10, 0))
        h = c2.body
        tip(ttk.Button(h, text="Export Excel global", style="Primary.TButton",
                       command=self.export_all),
            "Donnees, ajustements, courbes, IAST et graphique dans un seul "
            "classeur").pack(fill="x", pady=3)
        ttk.Button(h, text="Ajustements uniquement",
                   command=self.export_fits).pack(fill="x", pady=3)
        tip(ttk.Button(h, text="Export Origin (large, Long Name + Units)",
                       style="Success.TButton", command=self.export_origin),
            "Classeur au format attendu par Origin : ligne 1 = Long Name, "
            "ligne 2 = Units, donnees ligne 3. Une colonne par courbe.").pack(
            fill="x", pady=3)
        ttk.Button(h, text="Tous les jeux en CSV",
                   command=self.export_csv).pack(fill="x", pady=3)
        ttk.Button(h, text="Verification interne du programme",
                   command=self.run_selftest).pack(fill="x", pady=(12, 3))
        ttk.Label(h, style="Muted.TLabel", justify="left", wraplength=430,
                  text="La verification interne rejoue les tests numeriques "
                       "(modeles, IAST, equation d'etat) et affiche un rapport."
                  ).pack(anchor="w")

        c3 = card(self, "Resume de la session")
        c3.pack(fill="both", expand=True, pady=(10, 0))
        self.txt = tk.Text(c3.body, font=("Consolas", 9), wrap="none",
                           relief="flat", background="#fbfdff")
        vs = ttk.Scrollbar(c3.body, orient="vertical", command=self.txt.yview)
        self.txt.configure(yscrollcommand=vs.set)
        self.txt.pack(side="left", fill="both", expand=True)
        vs.pack(side="right", fill="y")
        ttk.Button(self, text="Actualiser le resume",
                   command=self.refresh).pack(anchor="w", pady=(8, 0))

    def refresh(self):
        a = self.app
        lines = ["%s v%s   -   %s" % (APP_NAME, APP_VERSION,
                                      datetime.datetime.now().strftime("%d/%m/%Y %H:%M")),
                 "", "JEUX DE DONNEES (%d)" % len(a.datasets)]
        for d in a.datasets:
            lines.append("  - %-28s %-5s %8.2f K  %-6s %3d points  "
                         "P: %s-%s bar   q_max=%s mmol/g"
                         % (d["name"], d["gas"], d["T"], d["kind"], len(d["P"]),
                            fmt(min(d["P"]), 3), fmt(max(d["P"]), 3),
                            fmt(max(d["q"]), 3)))
        lines += ["", "AJUSTEMENTS (%d)" % len(a.fits)]
        for f in a.fits:
            m = MODELS[f["model_key"]]
            lines.append("  - %-28s %-30s R2=%s  RMSE=%s"
                         % (f["dataset"], m.label, fmt(f["stats"]["R2"], 5),
                            fmt(f["stats"]["RMSE"], 5)))
            lines.append("      %s" % ", ".join(
                "%s = %s +/- %s %s" % (m.pnames[j], fmt(f["params"][j], 5),
                                       fmt(f["errors"][j], 5), m.punits[j])
                for j in range(m.nparam)))
        df = getattr(a.tab_iast, "df", None)
        lines += ["", "IAST : %s" % ("aucun calcul" if df is None
                                     else "%d points, colonnes : %s"
                                     % (len(df), ", ".join(df.columns[:6]) + "..."))]
        lines += ["", "SERIES GRAPHIQUES (%d)" % len(a.series)]
        for s in a.series:
            lines.append("  - %-40s %s  (%d points)"
                         % (s["label"], s["kind"], len(s["x"])))
        self.txt.delete("1.0", "end")
        self.txt.insert("1.0", "\n".join(lines))

    def export_all(self):
        if not self.app.datasets:
            messagebox.showwarning("Export", "Aucune donnee.", parent=self)
            return
        path = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                            initialfile="AdsorpSuite_export.xlsx",
                                            filetypes=[("Excel", "*.xlsx")])
        if not path:
            return
        try:
            export_all_excel(self.app, path)
        except Exception as e:
            messagebox.showerror("Export", "%s\n%s" % (e, traceback.format_exc()),
                                 parent=self)
            return
        self.app.log("Export global : %s" % path)
        messagebox.showinfo("Export", "Fichier enregistre :\n%s" % path,
                            parent=self)

    def export_fits(self):
        if not self.app.fits:
            messagebox.showwarning("Export", "Aucun ajustement.", parent=self)
            return
        path = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                            initialfile="ajustements.xlsx",
                                            filetypes=[("Excel", "*.xlsx")])
        if path:
            export_fits_excel(self.app, path)
            messagebox.showinfo("Export", "Fichier enregistre :\n%s" % path,
                                parent=self)

    def export_origin(self):
        if not self.app.datasets and not self.app.fits:
            messagebox.showinfo("Export Origin", "Rien a exporter.",
                                parent=self)
            return
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx", title="Export Origin",
            initialfile="export_origin.xlsx",
            filetypes=[("Classeur Excel", "*.xlsx")])
        if not path:
            return
        try:
            export_origin(self.app, path)
        except Exception as e:
            messagebox.showerror("Export Origin", str(e), parent=self)
            self.app.log("[ECHEC] export Origin : %s" % e)
            return
        self.app.log("Export Origin : %s" % path)
        self.app.set_status("Export Origin ecrit.")
        messagebox.showinfo(
            "Export Origin",
            "Fichier ecrit :\n%s\n\nDans Origin : File > Import > Excel, "
            "avec Long Names = 1, Units = 2, Data begins on = 3." % path,
            parent=self)

    def export_csv(self):
        if not self.app.datasets:
            return
        d = filedialog.askdirectory(title="Dossier de destination")
        if not d:
            return
        for ds in self.app.datasets:
            fn = "".join(c if c.isalnum() or c in "-_ " else "_"
                         for c in ds["name"]).strip()
            pd.DataFrame({"P (bar)": ds["P"], "q (mmol/g)": ds["q"]}).to_csv(
                os.path.join(d, "%s.csv" % fn), index=False)
        messagebox.showinfo("Export", "%d fichiers CSV ecrits dans :\n%s"
                            % (len(self.app.datasets), d), parent=self)

    def save(self, as_new=False):
        self.app.save_session(as_new=as_new)

    def open(self):
        self.app.open_session()

    def run_selftest(self):
        show_text_window(self, "Verification interne", run_selftest())


# =============================================================================
# 14. FENETRE PRINCIPALE
# =============================================================================

# =============================================================================
# 14. VERIFICATION INTERNE
# =============================================================================

def run_selftest():
    """Rejoue les controles numeriques du noyau scientifique."""
    out = []
    ok_n = [0]
    ko_n = [0]

    def chk(name, cond, detail=""):
        (ok_n if cond else ko_n).__getitem__(0)
        if cond:
            ok_n[0] += 1
        else:
            ko_n[0] += 1
        out.append("  [%s] %-52s %s" % ("OK " if cond else "ECHEC", name, detail))

    t0 = datetime.datetime.now()
    out.append("VERIFICATION INTERNE  -  %s v%s" % (APP_NAME, APP_VERSION))
    out.append("=" * 78)
    ctx = {"T": 303.15, "P0": 64.0}
    P = np.linspace(0.05, 50, 40)

    out.append("\n1. Ajustement : recuperation de parametres exacts")
    cases = [("langmuir", [14.5, 0.08]), ("dsl", [8.0, 0.5, 7.0, 0.02]),
             ("freundlich", [3.0, 0.45]), ("sips", [16.0, 0.06, 0.85]),
             ("toth", [18.0, 0.05, 0.7]), ("dr", [12.0, 2.2e-9]),
             ("da", [12.0, 18000.0, 2.4])]
    for key, true in cases:
        m = MODELS[key]
        q = np.asarray(m.func(P, ctx, *true)).ravel()
        try:
            r = fit_isotherm(m, P, q, ctx)
            e = max(abs(a - b) / max(abs(b), 1e-12)
                    for a, b in zip(r["params"], true))
            chk(m.label, r["stats"]["R2"] > 0.9999 and e < 1e-4,
                "R2=%.8f  erreur relative max=%.1e" % (r["stats"]["R2"], e))
        except Exception as ex:
            chk(m.label, False, str(ex))

    out.append("\n2. Pression d'etalement : formule analytique vs integration")
    def pi_num(m, par, Pt):
        y = np.linspace(math.log(Pt) - 40, math.log(Pt), 200000)
        return float(_TRAPZ(np.asarray(m.func(np.exp(y), ctx, *par)).ravel(), y))
    for key, par in cases:
        m = MODELS[key]
        sp = SpreadingPressure(m, ctx, par)
        errs = []
        for Pt in (0.5, 5.0, 40.0):
            a = float(sp.value(Pt)[0])
            n = pi_num(m, par, Pt)
            errs.append(abs(a - n) / max(abs(n), 1e-12))
        chk("pi* " + m.label, max(errs) < 1e-5, "ecart relatif max=%.1e"
            % max(errs))

    out.append("\n3. IAST : coherence thermodynamique")
    L = MODELS["langmuir"]
    qm, b1, b2 = 12.0, 0.30, 0.05
    c2 = [SpreadingPressure(L, ctx, [qm, b1]), SpreadingPressure(L, ctx, [qm, b2])]
    err = 0.0
    for Pt in (0.01, 0.5, 5, 50, 200):
        r = iast_point(Pt, [0.4, 0.6], c2)
        p1, p2 = 0.4 * Pt, 0.6 * Pt
        den = 1 + b1 * p1 + b2 * p2
        ex = [qm * b1 * p1 / den, qm * b2 * p2 / den]
        err = max(err, max(abs(r["q"][i] - ex[i]) / ex[i] for i in (0, 1)))
    chk("q_max identiques -> Langmuir competitif exact", err < 1e-9,
        "ecart relatif max=%.1e" % err)

    c3 = [SpreadingPressure(MODELS["toth"], ctx, [18.0, 0.05, 0.7]),
          SpreadingPressure(MODELS["sips"], ctx, [9.0, 0.03, 0.85]),
          SpreadingPressure(MODELS["da"], ctx, [6.0, 12000.0, 2.0])]
    dpi_max, dx_max = 0.0, 0.0
    for Pt in (0.5, 5.0, 30.0, 60.0):
        r = iast_point(Pt, [0.5, 0.3, 0.2], c3)
        pis = [float(c3[i].value(r["P0"][i])[0]) for i in range(3)]
        dpi_max = max(dpi_max, max(pis) - min(pis))
        dx_max = max(dx_max, abs(sum(r["x"]) - 1.0))
    chk("melange Toth+Sips+D-A : pi identiques et somme x = 1",
        dpi_max < 1e-6 and dx_max < 1e-9,
        "ecart pi max=%.1e   |somme x - 1|=%.1e" % (dpi_max, dx_max))

    r = iast_point(1e-6, [0.4, 0.6], c2)
    e1 = abs(r["q"][0] / (0.4e-6) - qm * b1) / (qm * b1)
    e2 = abs(r["q"][1] / (0.6e-6) - qm * b2) / (qm * b2)
    chk("limite basse pression -> constantes de Henry", max(e1, e2) < 1e-5,
        "ecart relatif max=%.1e" % max(e1, e2))

    sp = [SpreadingPressure(L, ctx, p, P_exp_max=60.0)
          for p in ([14.8885, 0.0806], [7.4225, 0.0577], [12.0235, 0.0090])]
    df = iast_scan(np.array([10.0, 60.0]), [1 / 3, 1 / 3, 1 / 3], sp,
                   ["CO2", "CH4", "H2"])
    rr = float(df["Extrapolation max (P0/P_exp)"].iloc[-1])
    chk("detection de l'extrapolation des modeles", rr > 2.0,
        "P0 max = %.1f x la pression mesuree a 60 bar" % rr)

    out.append("\n4. Equation d'etat et pressions de saturation")
    z1 = eos_z_effective(298.15, 1.0, "CO2")
    z2 = eos_z_effective(298.15, 100.0, "CH4")
    z3 = eos_z_effective(298.15, 100.0, "H2")
    chk("Z(CO2, 1 bar) proche de 1", 0.99 < z1 < 1.0, "Z=%.4f" % z1)
    chk("Z(CH4, 100 bar) entre 0.80 et 0.88", 0.80 < z2 < 0.88, "Z=%.4f" % z2)
    chk("Z(H2, 100 bar) superieur a 1", 1.0 < z3 < 1.10, "Z=%.4f" % z3)
    ps, meth = saturation_pressure("CO2", 298.15)
    chk("P_sat(CO2, 298.15 K) proche de 64.2 bar", abs(ps - 64.2) < 1.5,
        "%.2f bar (%s)" % (ps, meth))
    ps2, meth2 = saturation_pressure("CH4", 298.15)
    chk("CH4 supercritique -> regle de Dubinin", "Dubinin" in meth2,
        "%.1f bar" % ps2)
    nt, rho = excess_to_total(np.array([1.0, 30.0]), np.array([2.0, 12.0]),
                              298.15, "CO2", "Volume poreux", V_pore=0.6)
    chk("conversion exces -> total croissante avec P",
        nt[0] > 2.0 and nt[1] > 12.0 and (nt[1] - 12.0) > (nt[0] - 2.0),
        "n_tot = %.4f puis %.4f mmol/g" % (nt[0], nt[1]))

    out.append("\n5. Unites et robustesse")
    q0 = np.array([1.0, 5.0, 12.0])
    rt = mmol_g_to_unit(q_to_mmol_g(q0, "cm3(STP)/g", 44.01), "cm3(STP)/g", 44.01)
    chk("aller-retour d'unites cm3(STP)/g", np.allclose(rt, q0, rtol=1e-12))
    chk("lecture de nombres a virgule decimale",
        abs(to_float("1,25") - 1.25) < 1e-12 and np.isnan(to_float("abc")))
    chk("valeur de repli sur saisie invalide",
        to_float_or("abc", 7.0) == 7.0 and to_float_or("", 3.0) == 3.0)

    out.append("\n6. Robustesse statistique")
    rng = np.random.default_rng(4)
    qn = np.asarray(MODELS["sips"].func(P, ctx, 16.0, 0.06, 0.85)).ravel()
    qn = qn * (1 + rng.normal(0, 0.012, len(P)))
    r = fit_isotherm(MODELS["sips"], P, qn, ctx, n_starts=8, n_boot=60)
    chk("multi-depart et bootstrap operationnels",
        r["n_boot"] >= 10 and r["ci_lo"] is not None
        and r["ci_lo"][0] < r["params"][0] < r["ci_hi"][0],
        "%d tirages, IC q_max = [%.2f ; %.2f]"
        % (r["n_boot"], r["ci_lo"][0], r["ci_hi"][0]))
    chk("correlation des parametres calculee",
        np.isfinite(r.get("r_max", np.nan)), "|r|max = %.4f" % r["r_max"])
    bad = [200.0, 1e-4, 2.5]
    r1 = fit_isotherm(MODELS["toth"], P, qn, ctx, p0=bad, n_starts=1)
    r2 = fit_isotherm(MODELS["toth"], P, qn, ctx, p0=bad, n_starts=20)
    chk("multi-depart jamais moins bon qu'un depart unique",
        r2["stats"]["SSR"] <= r1["stats"]["SSR"] * (1 + 1e-9),
        "SSR %.4e -> %.4e" % (r1["stats"]["SSR"], r2["stats"]["SSR"]))

    # ---------------------------------------------------- 7. ajouts v3.0
    out.append("")
    out.append("7. Densites de reference et conversion (v3.0)")
    chk("table NIST : 3 gaz a 303 K",
        set(g for g, _T in NIST_TABLES) == {"CO2", "CH4", "H2"},
        ", ".join(sorted(g for g, _T in NIST_TABLES)))
    _v = nist_density_mmol_cm3(303.0, 40.0, "CH4")
    chk("rho(CH4, 303 K, 40 bar) = 1.69394 mol/L", abs(_v - 1.69394) < 1e-5,
        "%.5f mol/L" % _v)
    _v = nist_density_mmol_cm3(303.0, 40.0, "CO2")
    chk("rho(CO2, 303 K, 40 bar) = 2.04183 mol/L", abs(_v - 2.04183) < 1e-5,
        "%.5f mol/L" % _v)
    chk("CO2 au-dela de Psat(303 K) : hors table",
        not np.isfinite(nist_density_mmol_cm3(303.0, 80.0, "CO2")),
        "Psat = 71.89 bar")
    _a = gas_density_mmol_cm3(303.0, 80.0, "CO2", eos="Table NIST")
    _b = gas_density_mmol_cm3(303.0, 80.0, "CO2", eos="Peng-Robinson")
    chk("repli automatique sur Peng-Robinson hors table",
        abs(_a - _b) < 1e-12, "rho = %.5f mol/L" % _a)
    _e = 100.0 * abs(gas_density_mmol_cm3(303.0, 40.0, "CH4",
                                          eos="Peng-Robinson") / 1.69394 - 1)
    chk("ecart Peng-Robinson / NIST sur CH4 inferieur a 3 %", _e < 3.0,
        "%.2f %% a 40 bar" % _e)
    chk("deduction du gaz depuis le nom de fichier",
        guess_gas_from_name("ZTC850-CO2.xlsx") == "CO2"
        and guess_gas_from_name("P800_CH4.csv") == "CH4"
        and guess_gas_from_name("ZTC-H2.xlsx") == "H2")
    chk("deduction de l'echantillon depuis le nom de fichier",
        guess_sample_from_name("ZTC850-CO2.xlsx") == "ZTC850",
        guess_sample_from_name("ZTC850-CO2.xlsx"))
    _P = np.array([0.0, 10.0, 40.0])
    _nt, _rho = excess_to_total(_P, np.array([0.0, 3.0, 5.0]), 303.0, "CH4",
                                method="Volume poreux", V_pore=0.54,
                                eos="Table NIST")
    chk("conversion exces -> total via la table NIST",
        abs(_nt[2] - (5.0 + 1.69394 * 0.54)) < 1e-6,
        "n_tot(40 bar) = %.5f mmol/g" % _nt[2])

    dt = (datetime.datetime.now() - t0).total_seconds()
    out.append("\n" + "=" * 78)
    out.append("RESULTAT : %d controles reussis, %d echecs   (%.1f s)"
               % (ok_n[0], ko_n[0], dt))
    if ko_n[0] == 0:
        out.append("Le noyau scientifique se comporte comme attendu.")
    else:
        out.append("ATTENTION : au moins un controle a echoue, voir le detail.")
    return "\n".join(out)


# =============================================================================
# 15. FENETRE PRINCIPALE  (v2)
# =============================================================================

STEPS = ["Donnees", "Conversion", "Ajustement", "IAST", "Graphiques",
         "Session / Export"]


class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.datasets = []
        self.fits = []
        self.series = []
        self.notes = ""
        self.session_path = None
        self.dirty = False
        self._last_tab = 0
        self.settings = load_settings()

        apply_theme(self)
        self.title("%s v%s" % (APP_NAME, APP_VERSION))
        for ic in ("adsorpsuite.ico", "adsorpsuite.png"):
            try:
                path = resource_path(ic)
                if not os.path.exists(path):
                    continue
                if ic.endswith(".ico"):
                    self.iconbitmap(path)
                else:
                    self.iconphoto(True, tk.PhotoImage(file=path))
                break
            except Exception:
                continue
        geo = self.settings.get("geometry")
        self.geometry(geo if geo else "1500x960")
        self.minsize(1150, 720)
        if not geo:
            try:
                self.state("zoomed")
            except Exception:
                pass

        self._build_menu()
        self._build_toolbar()

        self.steps = StepBar(self, STEPS, command=self.goto)
        self.steps.pack(fill="x")
        tk.Frame(self, height=1, bg=UI["line"]).pack(fill="x")

        body = ttk.Frame(self, padding=(8, 8, 8, 4))
        body.pack(fill="both", expand=True)
        self.nb = ttk.Notebook(body, style="Hidden.TNotebook")
        self.nb.pack(fill="both", expand=True)
        self.tab_data = TabData(self.nb, self)
        self.tab_conv = TabConversion(self.nb, self)
        self.tab_fit = TabFit(self.nb, self)
        self.tab_iast = TabIAST(self.nb, self)
        self.tab_plot = TabPlot(self.nb, self)
        self.tab_export = TabExport(self.nb, self)
        for t, n in zip((self.tab_data, self.tab_conv, self.tab_fit,
                         self.tab_iast, self.tab_plot, self.tab_export), STEPS):
            self.nb.add(t, text=n)
        self.nb.bind("<<NotebookTabChanged>>", self._tab_changed)

        self._build_logbar()
        self._build_status()
        self._bind_keys()
        self.protocol("WM_DELETE_WINDOW", self.on_close)
        self.steps.set_active(0)

        self.log("%s v%s pret." % (APP_NAME, APP_VERSION))
        self.log("CoolProp %s. Astuce : F1 ouvre l'aide, F5 lance l'action de "
                 "l'onglet courant."
                 % ("detecte" if HAS_COOLPROP else "absent, Peng-Robinson "
                    "utilise"))
        last = self.settings.get("last_session")
        if last and os.path.exists(last):
            self.set_status("Derniere session : %s  (menu Fichier pour la "
                            "rouvrir)" % os.path.basename(last))

    # ------------------------------------------------------------ interface
    def _build_toolbar(self):
        bar = ttk.Frame(self, style="Bar.TFrame", padding=(10, 7))
        bar.pack(fill="x")
        ttk.Label(bar, text=APP_NAME, style="Title.TLabel").pack(side="left",
                                                                 padx=(0, 4))
        ttk.Label(bar, text="v%s" % APP_VERSION,
                  style="Muted.TLabel").pack(side="left", padx=(0, 18))

        def sep():
            tk.Frame(bar, width=1, bg=UI["line"]).pack(side="left", fill="y",
                                                       padx=8, pady=2)

        def btn(txt, cmd, hint, style="Tool.TButton"):
            b = ttk.Button(bar, text=txt, style=style, command=cmd)
            b.pack(side="left", padx=2)
            tip(b, hint)
            return b

        btn("Nouvelle session", self.new_session,
            "Repartir de zero (Ctrl+N)")
        btn("Ouvrir...", self.open_session,
            "Recharger une session enregistree (Ctrl+O)")
        self.btn_save = btn("Enregistrer", self.save_session,
                            "Enregistrer toute la session dans un fichier "
                            ".adsp (Ctrl+S)", "Primary.TButton")
        sep()
        btn("Importer des donnees", lambda: (self.goto(0),
                                             self.tab_data.import_file()),
            "Lire un fichier Excel ou CSV d'isotherme")
        btn("Ajuster", lambda: (self.goto(2), self.tab_fit.run_fit()),
            "Aller a l'ajustement et lancer les modeles coches")
        btn("IAST", lambda: (self.goto(3), self.tab_iast.run()),
            "Aller au calcul de melange et le lancer")
        sep()
        btn("Export Excel", self.tab_export_all,
            "Classeur complet : donnees, ajustements, courbes, IAST")
        btn("Aide", self.show_help, "Aide rapide (F1)")

    def _build_logbar(self):
        wrap = ttk.Frame(self, padding=(8, 0, 8, 0))
        wrap.pack(fill="x")
        self.log_open = tk.BooleanVar(value=True)
        head = ttk.Frame(wrap, style="Card.TFrame", padding=(10, 4))
        head.pack(fill="x")
        self.btn_log = ttk.Button(head, text="Journal  -  masquer",
                                  style="Ghost.TButton", command=self._toggle_log)
        self.btn_log.pack(side="left")
        ttk.Button(head, text="Effacer", style="Ghost.TButton",
                   command=lambda: self.txt_log.delete("1.0", "end")).pack(
            side="left", padx=6)
        self.logbox = ttk.Frame(wrap, style="Card.TFrame", padding=(10, 0, 10, 8))
        self.logbox.pack(fill="x")
        self.txt_log = tk.Text(self.logbox, height=5, font=("Consolas", 9),
                               relief="flat", background="#fbfdff",
                               foreground=UI["ink"])
        vs = ttk.Scrollbar(self.logbox, orient="vertical",
                           command=self.txt_log.yview)
        self.txt_log.configure(yscrollcommand=vs.set)
        self.txt_log.pack(side="left", fill="x", expand=True)
        vs.pack(side="right", fill="y")

    def _toggle_log(self):
        if self.log_open.get():
            self.logbox.pack_forget()
            self.btn_log.config(text="Journal  -  afficher")
        else:
            self.logbox.pack(fill="x")
            self.btn_log.config(text="Journal  -  masquer")
        self.log_open.set(not self.log_open.get())

    def _build_status(self):
        st = ttk.Frame(self, style="Bar.TFrame", padding=(10, 5))
        st.pack(fill="x", side="bottom")
        self.lbl_status = ttk.Label(st, text="Pret", style="Status.TLabel")
        self.lbl_status.pack(side="left")
        self.pb = ttk.Progressbar(st, mode="determinate", length=170)
        self.lbl_session = ttk.Label(st, text="Session non enregistree",
                                     style="Status.TLabel")
        self.lbl_session.pack(side="right")

    def _build_menu(self):
        mb = tk.Menu(self)
        m1 = tk.Menu(mb, tearoff=0)
        m1.add_command(label="Nouvelle session", accelerator="Ctrl+N",
                       command=self.new_session)
        m1.add_command(label="Ouvrir une session...", accelerator="Ctrl+O",
                       command=self.open_session)
        m1.add_command(label="Enregistrer la session", accelerator="Ctrl+S",
                       command=self.save_session)
        m1.add_command(label="Enregistrer sous...", accelerator="Ctrl+Maj+S",
                       command=lambda: self.save_session(as_new=True))
        self.menu_recent = tk.Menu(m1, tearoff=0)
        m1.add_cascade(label="Sessions recentes", menu=self.menu_recent)
        m1.add_separator()
        m1.add_command(label="Importer plusieurs isothermes...",
                       accelerator="Ctrl+I", command=self._import_many)
        m1.add_command(label="Export Origin (Long Name + Units)...",
                       accelerator="Ctrl+E",
                       command=lambda: self.tab_export.export_origin())
        m1.add_separator()
        m1.add_command(label="Export Excel global...",
                       command=self.tab_export_all)
        m1.add_separator()
        m1.add_command(label="Quitter", command=self.on_close)
        mb.add_cascade(label="Fichier", menu=m1)
        m2 = tk.Menu(mb, tearoff=0)
        for i, nm in enumerate(STEPS):
            m2.add_command(label="%d. %s" % (i + 1, nm),
                           accelerator="Ctrl+%d" % (i + 1),
                           command=lambda k=i: self.goto(k))
        mb.add_cascade(label="Aller a", menu=m2)
        m3 = tk.Menu(mb, tearoff=0)
        m3.add_command(label="Aide rapide", accelerator="F1",
                       command=self.show_help)
        m3.add_command(label="Modeles et equations",
                       command=lambda: show_text_window(self, "Modeles",
                                                        models_text()))
        m3.add_command(label="Verification interne du programme",
                       command=lambda: show_text_window(
                           self, "Verification interne", run_selftest()))
        m3.add_command(label="A propos", command=self.about)
        mb.add_cascade(label="Aide", menu=m3)
        self.config(menu=mb)
        self._refresh_recent()

    def _bind_keys(self):
        self.bind_all("<Control-s>", lambda e: self.save_session())
        self.bind_all("<Control-S>", lambda e: self.save_session(as_new=True))
        self.bind_all("<Control-o>", lambda e: self.open_session())
        self.bind_all("<Control-n>", lambda e: self.new_session())
        self.bind_all("<F1>", lambda e: self.show_help())
        self.bind_all("<F5>", lambda e: self.run_current())
        self.bind_all("<Shift-F5>", lambda e: self._run_all())
        self.bind_all("<Control-i>", lambda e: self._import_many())
        self.bind_all("<Control-e>", lambda e: self.tab_export.export_origin())
        for i in range(6):
            self.bind_all("<Control-Key-%d>" % (i + 1),
                          lambda e, k=i: self.goto(k))

    def _run_all(self):
        """Maj+F5 : action 'en lot' de l'onglet actif."""
        try:
            i = self.nb.index(self.nb.select())
        except Exception:
            return
        if i == 1:
            self.tab_conv.convert_all()
        elif i == 2:
            self.tab_fit.run_fit_all()
        else:
            self.set_status("Maj+F5 : disponible dans les onglets Conversion "
                            "et Ajustement.")

    def _import_many(self):
        self.goto(0)
        self.tab_data.import_many()

    # ------------------------------------------------------------ navigation
    def goto(self, idx):
        try:
            self.nb.select(idx)
        except Exception:
            pass

    def run_current(self):
        i = self._last_tab
        try:
            (self.tab_data.save_current, self.tab_conv.compute,
             self.tab_fit.run_fit, self.tab_iast.run, self.tab_plot.draw,
             self.tab_export.refresh)[i]()
        except Exception as e:
            self.log("F5 : %s" % e)

    def _tab_changed(self, event=None):
        try:
            if getattr(self, "tab_data", None) is not None:
                self.tab_data._autosave()
        except Exception:
            pass
        try:
            idx = self.nb.index(self.nb.select())
        except Exception:
            return
        self._last_tab = idx
        self.steps.set_active(idx)
        if idx == 1:
            self.tab_conv.refresh()
        elif idx == 2:
            self.tab_fit.refresh()
        elif idx == 3:
            self.tab_iast.refresh()
        elif idx == 5:
            self.tab_export.refresh()

    # --------------------------------------------------------------- session
    def _title(self):
        nm = (os.path.basename(self.session_path) if self.session_path
              else "session non enregistree")
        self.title("%s v%s  -  %s%s" % (APP_NAME, APP_VERSION, nm,
                                        " *" if self.dirty else ""))
        try:
            self.lbl_session.config(
                text="%s%s" % (nm, "  (modifie)" if self.dirty else ""))
        except Exception:
            pass

    def mark_dirty(self, state=True):
        self.dirty = bool(state)
        self._title()

    def save_session(self, as_new=False):
        path = self.session_path
        if as_new or not path:
            init = os.path.basename(path) if path else "session_adsorption.adsp"
            path = filedialog.asksaveasfilename(
                defaultextension=SESSION_EXT, initialfile=init,
                initialdir=self.settings.get("last_dir", ""),
                title="Enregistrer la session",
                filetypes=[("Session AdsorpSuite", "*" + SESSION_EXT),
                           ("JSON", "*.json")])
            if not path:
                return False
        try:
            write_session(self, path)
        except Exception as e:
            messagebox.showerror("Session", "Enregistrement impossible :\n%s"
                                 % e, parent=self)
            return False
        self.session_path = path
        self.settings["last_session"] = path
        self.settings["last_dir"] = os.path.dirname(path)
        self._push_recent(path)
        save_settings(self.settings)
        self.mark_dirty(False)
        self.log("Session enregistree : %s" % path)
        self.set_status("Session enregistree (%d jeux, %d ajustements)"
                        % (len(self.datasets), len(self.fits)))
        return True

    def open_session(self, path=None):
        if not self._confirm_discard():
            return
        if not path:
            path = filedialog.askopenfilename(
                title="Ouvrir une session",
                initialdir=self.settings.get("last_dir", ""),
                filetypes=[("Session AdsorpSuite", "*" + SESSION_EXT),
                           ("JSON", "*.json"), ("Tous les fichiers", "*.*")])
        if not path:
            return
        try:
            data = read_session(path)
            info = restore_session(self, data)
        except Exception as e:
            messagebox.showerror("Session", "Lecture impossible :\n%s\n\n%s"
                                 % (e, traceback.format_exc()), parent=self)
            return
        self.session_path = path
        self.settings["last_session"] = path
        self.settings["last_dir"] = os.path.dirname(path)
        self._push_recent(path)
        save_settings(self.settings)
        self.refresh_all()
        self.mark_dirty(False)
        msg = ("Session chargee : %d jeux, %d ajustements, %d series, "
               "%d points IAST" % (info["datasets"], info["fits"],
                                   info["series"], info["iast"]))
        self.log(msg + "  (%s)" % path)
        self.set_status(msg)

    def new_session(self):
        if not self._confirm_discard():
            return
        self.datasets, self.fits, self.series = [], [], []
        self.notes = ""
        self.session_path = None
        self.tab_iast.df = None
        self.refresh_all()
        self.mark_dirty(False)
        self.log("Nouvelle session.")

    def _confirm_discard(self):
        if not self.dirty:
            return True
        r = messagebox.askyesnocancel(
            "Session en cours",
            "La session en cours contient des modifications non "
            "enregistrees.\n\nVoulez-vous l'enregistrer maintenant ?",
            parent=self)
        if r is None:
            return False
        if r:
            return self.save_session()
        return True

    def _push_recent(self, path):
        rec = [p for p in self.settings.get("recent", []) if p != path]
        rec.insert(0, path)
        self.settings["recent"] = rec[:8]
        self._refresh_recent()

    def _refresh_recent(self):
        try:
            self.menu_recent.delete(0, "end")
        except Exception:
            return
        rec = [p for p in self.settings.get("recent", []) if os.path.exists(p)]
        if not rec:
            self.menu_recent.add_command(label="(aucune)", state="disabled")
            return
        for p in rec:
            self.menu_recent.add_command(
                label=os.path.basename(p),
                command=lambda q=p: self.open_session(q))

    def on_close(self):
        if not self._confirm_discard():
            return
        try:
            self.settings["geometry"] = self.geometry()
            save_settings(self.settings)
        except Exception:
            pass
        self.destroy()

    # ------------------------------------------------------------- services
    def refresh_all(self):
        for fn in (self.tab_data.refresh_list, self.tab_fit.refresh_fits,
                   self.tab_plot.refresh_series, self.tab_conv.refresh,
                   self.tab_fit.refresh, self.tab_iast.refresh,
                   self.tab_export.refresh):
            try:
                fn()
            except Exception:
                pass

    def notify_datasets(self):
        try:
            self.tab_data.refresh_list()
        except Exception:
            pass
        for t in (self.tab_conv, self.tab_fit):
            try:
                t.refresh()
            except Exception:
                pass
        self.mark_dirty()

    def notify_fits(self):
        try:
            self.tab_iast.refresh()
        except Exception:
            pass
        self.mark_dirty()

    def add_series(self, label, x, y, kind="ligne"):
        i = len(self.series)
        self.series.append(dict(label=label, x=np.asarray(x, float),
                                y=np.asarray(y, float), kind=kind,
                                color=PALETTE[i % len(PALETTE)],
                                marker=MARKERS[i % 6], ls="-",
                                lw=1.8, ms=6.0, alpha=1.0, visible=True,
                                y2=False))
        self.mark_dirty()

    def tab_export_all(self):
        self.goto(5)
        self.tab_export.export_all()

    def show_help(self):
        show_text_window(self, "Aide rapide", HELP_TEXT)

    def about(self):
        messagebox.showinfo(
            "A propos",
            "%s v%s\n\nAnalyse d'isothermes d'adsorption : ajustement de "
            "modeles, conversion exces/total, IAST binaire et ternaire, "
            "graphiques et exports.\n\nPython %s\nnumpy %s - pandas %s - "
            "matplotlib %s\nCoolProp : %s\n\nConfiguration : %s"
            % (APP_NAME, APP_VERSION, sys.version.split()[0], np.__version__,
               pd.__version__, matplotlib.__version__,
               "installe" if HAS_COOLPROP else "non installe", config_dir()),
            parent=self)

    def set_status(self, txt):
        try:
            self.lbl_status.config(text=txt)
        except Exception:
            pass

    def busy(self, on=True, value=None):
        """Affiche ou masque la barre de progression."""
        try:
            if on:
                self.pb.pack(side="right", padx=10)
                if value is not None:
                    self.pb["value"] = value
                self.config(cursor="watch")
            else:
                self.pb.pack_forget()
                self.config(cursor="")
            self.update_idletasks()
        except Exception:
            pass

    def log(self, msg):
        ts = datetime.datetime.now().strftime("%H:%M:%S")
        try:
            self.txt_log.insert("end", "[%s] %s\n" % (ts, msg))
            self.txt_log.see("end")
        except Exception:
            print(msg)
        self.set_status(msg[:150])


def main():
    if "--selftest" in sys.argv:
        print(run_selftest())
        return
    app = App()
    app.mainloop()


if __name__ == "__main__":
    main()
